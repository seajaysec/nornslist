#!/usr/bin/env python3
"""
Norns Community Script Scraper

This script scrapes script details from norns.community and outputs them to an Excel file.
"""

import argparse
import logging
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import pandas as pd
import questionary
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from playwright.sync_api import sync_playwright

# Set up logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class NornsScraper:
    # Field mapping: Excel column -> internal key
    FIELD_MAP = {
        "Name": "project_name",
        "Author": "author",
        "Tags": "tags",
        "Description": "description",
        "Demo": "demo",
        "Discussion URL": "discussion_url",
        "Project URL": "project_url",
        "Community URL": "community_url",
        "Playwright Status": "playwright_status",
        "Last Updated": "last_updated",
        "Out of Sync": "out_of_sync",
    }

    # Playwright status values
    RESOLVED_STATUSES = [
        "No Conflict",
        "Playwright Preferred",
        "Extract Preferred",
        "Manual Override",
        "Missing Demo",
    ]

    def __init__(
        self,
        base_url="https://norns.community",
        max_workers=10,
        demo_delay=0.5,
    ):
        self.base_url = base_url
        self.max_workers = max_workers
        self.session = requests.Session()

        # Configure connection pool to handle high concurrency
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=self.max_workers,
            pool_maxsize=self.max_workers * 2,
            max_retries=3,
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
        )
        self.script_data = []
        self.demo_delay = demo_delay
        self.failed_demo_requests = []  # Track failed requests for retry
        self.playwright_conflicts = []  # Store conflicts for user resolution
        self.summary_stats = {
            "scripts_added": 0,
            "scripts_updated": 0,
            "scripts_preserved": 0,
            "total_scripts": 0,
            "added_details": [],
            "updated_details": [],
        }

        # GitHub integration
        self.github_token = self._load_github_token()
        self.github_session = self._init_github_session()

    def discover_demo_video(self, discussion_url):
        """Discover demo video from discussion URL by looking for YouTube, Vimeo, or SoundCloud links"""
        if not discussion_url:
            return ""

        try:
            logger.debug(f"Checking discussion URL for demo: {discussion_url}")
            # Add a small delay to avoid rate limiting
            import time

            time.sleep(self.demo_delay)  # Configurable delay between requests
            response = self.session.get(
                discussion_url, timeout=10, allow_redirects=True
            )
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")

            # Look for links in the page content
            links = soup.find_all("a", href=True)

            # Also look for iframe embeds (like Vimeo, YouTube embeds)
            iframes = soup.find_all("iframe")

            # Patterns for video/audio platforms
            patterns = [
                r"https?://(?:www\.)?youtube\.com/watch\?v=([^&\s]+)",
                r"https?://(?:www\.)?youtu\.be/([^&\s]+)",
                r"https?://(?:www\.)?vimeo\.com/(\d+)",
                r"https?://(?:www\.)?soundcloud\.com/[^/\s]+/[^/\s]+",
                r"https?://(?:www\.)?instagram\.com/[^/\s]+/p/[^/\s]+",
            ]

            # Check regular links first
            for link in links:
                href = link.get("href", "")
                if not href:
                    continue

                # HTML-decode href to handle &amp; etc.
                from html import unescape

                href = unescape(href)

                # Normalize URL for better matching (lowercase, remove www, handle protocols)
                href_lower = href.lower()
                href_normalized = (
                    href_lower.replace("www.", "")
                    .replace("http://", "")
                    .replace("https://", "")
                )

                # Check if it's a YouTube link (handle various formats)
                if any(
                    pattern in href_lower
                    for pattern in [
                        "youtube.com/watch",
                        "youtu.be/",
                        "m.youtube.com/watch",
                        "music.youtube.com/watch",
                    ]
                ):
                    logger.info(f"Found YouTube demo: {href}")
                    return href

                # Check if it's a Vimeo link (handle various formats)
                if any(
                    pattern in href_lower
                    for pattern in ["vimeo.com/", "player.vimeo.com/"]
                ):
                    logger.info(f"Found Vimeo demo: {href}")
                    return href

                # Check if it's a SoundCloud link (handle various formats)
                if any(
                    pattern in href_lower
                    for pattern in ["soundcloud.com/", "m.soundcloud.com/"]
                ):
                    logger.info(f"Found SoundCloud demo: {href}")
                    return href

                # Check if it's an Instagram link (handle various formats)
                if (
                    any(
                        pattern in href_lower
                        for pattern in ["instagram.com/", "www.instagram.com/"]
                    )
                    and "/p/" in href_lower
                ):
                    logger.info(f"Found Instagram demo: {href}")
                    return href

            # Check iframe embeds (for embedded videos)
            for iframe in iframes:
                src = iframe.get("src", "")
                data_original = iframe.get("data-original-href", "")

                # HTML-decode iframe src and data attributes to handle &amp; etc.
                if src:
                    from html import unescape

                    src = unescape(src)
                if data_original:
                    data_original = unescape(data_original)

                # Normalize iframe src for better matching
                src_lower = src.lower()

                # Check iframe src for YouTube embeds (handle various formats)
                if any(
                    pattern in src_lower
                    for pattern in [
                        "youtube.com/embed/",
                        "youtu.be/",
                        "m.youtube.com/embed/",
                        "music.youtube.com/embed/",
                    ]
                ):
                    # Extract video ID and construct watch URL
                    if "youtube.com/embed/" in src:
                        video_id = src.split("youtube.com/embed/")[1].split("?")[0]
                        href = f"https://www.youtube.com/watch?v={video_id}"
                    else:
                        href = src.replace(
                            "youtu.be/", "https://www.youtube.com/watch?v="
                        )
                    logger.info(f"Found YouTube embed demo: {href}")
                    return href

                # Check iframe src for Vimeo embeds (handle various formats)
                if any(
                    pattern in src_lower
                    for pattern in ["vimeo.com/video/", "player.vimeo.com/video/"]
                ):
                    video_id = src.split("vimeo.com/video/")[1].split("?")[0]
                    href = f"https://vimeo.com/{video_id}"
                    logger.info(f"Found Vimeo embed demo: {href}")
                    return href

                # Check data-original-href for Vimeo (common pattern)
                if data_original and "vimeo.com/" in data_original:
                    logger.info(
                        f"Found Vimeo embed demo (data-original): {data_original}"
                    )
                    return data_original

                # Check iframe src for SoundCloud embeds (handle various formats)
                if any(
                    pattern in src_lower
                    for pattern in [
                        "soundcloud.com/player",
                        "w.soundcloud.com",
                        "player.soundcloud.com",
                    ]
                ):
                    # Extract track ID from SoundCloud player URL
                    import urllib.parse

                    try:
                        # Parse the URL to get query parameters
                        parsed_url = urllib.parse.urlparse(src)
                        query_params = urllib.parse.parse_qs(parsed_url.query)

                        # Look for 'url' parameter which contains the SoundCloud API URL
                        if "url" in query_params:
                            api_url = query_params["url"][0]
                            decoded_url = urllib.parse.unquote(api_url)

                            # Extract track ID from API URL like https://api.soundcloud.com/tracks/970754032
                            if "api.soundcloud.com/tracks/" in decoded_url:
                                track_id = decoded_url.split(
                                    "api.soundcloud.com/tracks/"
                                )[1]
                                # Construct public SoundCloud URL (we'll use a generic format)
                                href = f"https://soundcloud.com/track/{track_id}"
                                logger.info(f"Found SoundCloud embed demo: {href}")
                                return href
                    except Exception as e:
                        logger.debug(f"Error parsing SoundCloud iframe: {e}")
                        continue

            # Check for video embedding divs and other containers
            video_containers = soup.find_all(
                ["div", "span"],
                class_=lambda x: x
                and any(
                    keyword in x.lower()
                    for keyword in ["youtube", "vimeo", "video", "embed"]
                ),
            )

            for container in video_containers:
                # Check for data-video-id attributes (common in YouTube embeds)
                video_id = container.get("data-video-id", "")
                if video_id:
                    # HTML-decode data attributes to handle &amp; etc.
                    from html import unescape

                    video_id = unescape(video_id)
                    # Check if it's YouTube (most common)
                    if "youtube" in container.get("class", []) or "youtube" in str(
                        container.get("class", "")
                    ):
                        href = f"https://www.youtube.com/watch?v={video_id}"
                        logger.info(f"Found YouTube demo (data-video-id): {href}")
                        return href
                    # Could be other platforms, but YouTube is most common
                    elif (
                        len(video_id) == 11
                    ):  # YouTube video IDs are typically 11 characters
                        href = f"https://www.youtube.com/watch?v={video_id}"
                        logger.info(
                            f"Found YouTube demo (data-video-id, assumed): {href}"
                        )
                        return href

                # Check for data-provider-name attributes
                provider = container.get("data-provider-name", "").lower()
                if provider == "youtube" and video_id:
                    href = f"https://www.youtube.com/watch?v={video_id}"
                    logger.info(f"Found YouTube demo (data-provider): {href}")
                    return href

                # Check for Vimeo data attributes
                if "vimeo" in container.get("class", []) or "vimeo" in str(
                    container.get("class", "")
                ):
                    vimeo_id = container.get("data-video-id", "")
                    if vimeo_id:
                        href = f"https://vimeo.com/{vimeo_id}"
                        logger.info(f"Found Vimeo demo (data-video-id): {href}")
                        return href

            # Check for JSON-LD structured data (common in modern websites)
            json_scripts = soup.find_all("script", type="application/ld+json")
            for script in json_scripts:
                try:
                    import json

                    data = json.loads(script.string)

                    # Handle single objects or arrays
                    if isinstance(data, list):
                        for item in data:
                            if self._extract_from_json_ld(item):
                                return self._extract_from_json_ld(item)
                    elif isinstance(data, dict):
                        result = self._extract_from_json_ld(data)
                        if result:
                            return result
                except (json.JSONDecodeError, AttributeError):
                    continue

            logger.debug(f"No demo video found in discussion: {discussion_url}")
            return ""

        except requests.exceptions.HTTPError as e:
            if e.response.status_code in [429, 403]:  # Rate limited or forbidden
                logger.warning(
                    f"Access denied (HTTP {e.response.status_code}) when checking discussion URL {discussion_url}"
                )
                # Track for retry
                self.failed_demo_requests.append(discussion_url)
                return ""
            else:
                logger.warning(
                    f"HTTP error checking discussion URL {discussion_url}: {e}"
                )
                return ""
        except Exception as e:
            logger.warning(f"Error checking discussion URL {discussion_url}: {e}")
            return ""

    def _extract_from_json_ld(self, data):
        """Extract video URL from JSON-LD structured data"""
        if not isinstance(data, dict):
            return None

        # Check for VideoObject type
        if data.get("@type") == "VideoObject":
            # Look for contentUrl (direct video URL)
            content_url = data.get("contentUrl", "")
            if content_url:
                # HTML-decode URLs to handle &amp; etc.
                from html import unescape

                content_url = unescape(content_url)
                if "youtube.com/watch" in content_url or "youtu.be/" in content_url:
                    logger.info(
                        f"Found YouTube demo (JSON-LD contentUrl): {content_url}"
                    )
                    return content_url
                elif "vimeo.com/" in content_url:
                    logger.info(f"Found Vimeo demo (JSON-LD contentUrl): {content_url}")
                    return content_url
                elif "soundcloud.com/" in content_url:
                    logger.info(
                        f"Found SoundCloud demo (JSON-LD contentUrl): {content_url}"
                    )
                    return content_url
                elif "instagram.com/" in content_url and "/p/" in content_url:
                    logger.info(
                        f"Found Instagram demo (JSON-LD contentUrl): {content_url}"
                    )
                    return content_url

            # Look for embedUrl (embedded video URL)
            embed_url = data.get("embedUrl", "")
            if embed_url:
                # HTML-decode URLs to handle &amp; etc.
                from html import unescape

                embed_url = unescape(embed_url)
                if "youtube.com/embed/" in embed_url:
                    video_id = embed_url.split("youtube.com/embed/")[1].split("?")[0]
                    href = f"https://www.youtube.com/watch?v={video_id}"
                    logger.info(f"Found YouTube demo (JSON-LD embedUrl): {href}")
                    return href
                elif "vimeo.com/video/" in embed_url:
                    video_id = embed_url.split("vimeo.com/video/")[1].split("?")[0]
                    href = f"https://vimeo.com/{video_id}"
                    logger.info(f"Found Vimeo demo (JSON-LD embedUrl): {href}")
                    return href

        return None

    def retry_failed_demo_requests(self):
        """Retry demo discovery for URLs that failed due to rate limiting"""
        if not self.failed_demo_requests:
            return

        logger.info(
            f"Retrying {len(self.failed_demo_requests)} failed demo requests..."
        )

        # Increase delay for retries to be more respectful
        original_delay = self.demo_delay
        self.demo_delay = max(2.0, original_delay * 2)  # At least 2 seconds for retries

        retry_successes = 0
        for discussion_url in self.failed_demo_requests[
            :
        ]:  # Copy list to modify during iteration
            try:
                logger.info(f"Retrying demo discovery for: {discussion_url}")
                demo_url = self.discover_demo_video(discussion_url)

                if demo_url:
                    # Find the script in our data and update it
                    for script in self.script_data:
                        if script.get("discussion_url") == discussion_url:
                            script["demo"] = demo_url
                            retry_successes += 1
                            logger.info(f"Successfully found demo on retry: {demo_url}")
                            break

                    # Remove from failed list
                    self.failed_demo_requests.remove(discussion_url)

            except Exception as e:
                logger.warning(f"Retry failed for {discussion_url}: {e}")

        # Restore original delay
        self.demo_delay = original_delay

        logger.info(
            f"Retry completed: {retry_successes} demos found, {len(self.failed_demo_requests)} still failed"
        )

    def get_main_page(self):
        """Fetch the main norns.community page to get list of scripts"""
        try:
            logger.info("Fetching main page...")
            response = self.session.get(self.base_url)
            response.raise_for_status()
            return response.text
        except requests.RequestException as e:
            logger.error(f"Error fetching main page: {e}")
            return None

    def extract_script_links(self, html_content):
        """Extract script links from the main page"""
        soup = BeautifulSoup(html_content, "html.parser")
        script_links = []
        seen_paths = set()

        # Find all script links in the main list
        script_items = soup.find_all("li")

        for item in script_items:
            link = item.find("a")
            if link and link.get("href"):
                href = link.get("href")
                # Skip if it's not a script link (like external links)
                if (
                    href.startswith("/")
                    and not href.startswith("/author")
                    and not href.startswith("/explore")
                    and not href.startswith("/about")
                ):
                    script_name = href.strip("/")
                    if script_name in seen_paths:
                        continue
                    seen_paths.add(script_name)
                    script_url = urljoin(self.base_url, href)
                    script_links.append({"name": script_name, "url": script_url})

        logger.info(f"Found {len(script_links)} script links")
        return script_links

    def scrape_script_details(
        self, script_url, script_name, existing_data=None, discover_demo=True
    ):
        """Scrape detailed information from a single script page"""
        # Create a new session for each thread to avoid conflicts
        session = requests.Session()

        # Configure connection pool to handle high concurrency
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=self.max_workers,
            pool_maxsize=self.max_workers * 2,
            max_retries=3,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
        )

        try:
            logger.info(f"Scraping {script_name} from {script_url}...")
            response = session.get(script_url)
            response.raise_for_status()

            # Use html.parser for better compatibility
            soup = BeautifulSoup(response.text, "html.parser")

            # Initialize data structure - create fresh dict for each thread
            script_data = {
                "project_name": script_name,
                "author": "",
                "tags": [],
                "description": "",
                "demo": "",
                "discussion_url": "",
                "project_url": "",
                "community_url": script_url,
            }

            # Check which fields we need to scrape based on existing data
            fields_to_scrape = {
                "project_name": True,
                "author": True,
                "tags": True,
                "description": True,
                "discussion_url": True,
                "project_url": True,
            }

            if existing_data:
                # Skip fields that are already populated
                if (
                    pd.notna(existing_data.get("Author"))
                    and existing_data.get("Author") != ""
                ):
                    fields_to_scrape["author"] = False
                    script_data["author"] = existing_data["Author"]
                if (
                    pd.notna(existing_data.get("Description"))
                    and existing_data.get("Description") != ""
                ):
                    fields_to_scrape["description"] = False
                    script_data["description"] = existing_data["Description"]
                if (
                    pd.notna(existing_data.get("Tags"))
                    and existing_data.get("Tags") != ""
                ):
                    fields_to_scrape["tags"] = False
                    script_data["tags"] = existing_data["Tags"]
                if (
                    pd.notna(existing_data.get("Discussion URL"))
                    and existing_data.get("Discussion URL") != ""
                ):
                    fields_to_scrape["discussion_url"] = False
                    script_data["discussion_url"] = existing_data["Discussion URL"]
                if (
                    pd.notna(existing_data.get("Project URL"))
                    and existing_data.get("Project URL") != ""
                ):
                    fields_to_scrape["project_url"] = False
                    script_data["project_url"] = existing_data["Project URL"]

                # Log which fields we're skipping
                skipped_fields = [
                    field
                    for field, should_scrape in fields_to_scrape.items()
                    if not should_scrape
                ]
                if skipped_fields:
                    logger.info(
                        f"Skipping already populated fields for {script_name}: {skipped_fields}"
                    )

            # Extract data from the table
            table = soup.find("table")
            if table:
                rows = table.find_all("tr")
                for row in rows:
                    cells = row.find_all("td")
                    if len(cells) >= 2:
                        key = cells[0].get_text().strip().lower()
                        value_cell = cells[1]

                        if key == "project name:" and fields_to_scrape["project_name"]:
                            script_data["project_name"] = value_cell.get_text().strip()
                        elif key == "project url:" and fields_to_scrape["project_url"]:
                            link = value_cell.find("a")
                            if link:
                                script_data["project_url"] = link.get(
                                    "href", ""
                                ).strip()
                        elif (
                            key in ["author:", "authors:"]
                            and fields_to_scrape["author"]
                        ):
                            # Find all links in the author cell and join them with commas
                            links = value_cell.find_all("a")
                            if links:
                                # Extract all author names and join with commas
                                author_names = [
                                    link.get_text().strip() for link in links
                                ]
                                author_text = ", ".join(author_names)
                                script_data["author"] = author_text
                                logger.info(
                                    f"Extracted author(s) for {script_name}: '{author_text}' (from {len(links)} links)"
                                )
                            else:
                                logger.warning(
                                    f"No author link found for {script_name}"
                                )
                                script_data["author"] = ""
                        elif key == "description:" and fields_to_scrape["description"]:
                            script_data["description"] = value_cell.get_text().strip()
                        elif (
                            key == "discussion url:"
                            and fields_to_scrape["discussion_url"]
                        ):
                            link = value_cell.find("a")
                            if link:
                                script_data["discussion_url"] = link.get(
                                    "href", ""
                                ).strip()
                        elif key == "tags:" and fields_to_scrape["tags"]:
                            tag_links = value_cell.find_all("a", class_="project-tag")
                            script_data["tags"] = [
                                tag.get_text().strip() for tag in tag_links
                            ]

            # If no project URL found in table, use the script URL
            if not script_data["project_url"]:
                script_data["project_url"] = script_url

            # Discover demo video if enabled and demo field is empty
            if (
                discover_demo
                and script_data["discussion_url"]
                and not script_data["demo"]
            ):
                demo_url = self.discover_demo_video(script_data["discussion_url"])
                if demo_url:
                    script_data["demo"] = demo_url
                    # Only set status if there was a conflict
                    # Regular demo discovery without conflicts should not set status
                    logger.info(f"Discovered demo for {script_name}: {demo_url}")

            logger.debug(f"Final data for {script_name}: {script_data}")
            return script_data

        except requests.RequestException as e:
            logger.error(f"Error scraping {script_name}: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error scraping {script_name}: {e}")
            return None
        finally:
            session.close()

    def scrape_all_scripts(self, test_filter=None):
        """Scrape all scripts from norns.community using parallel processing"""
        # Get main page
        main_html = self.get_main_page()
        if not main_html:
            logger.error("Failed to fetch main page")
            return

        # Extract script links
        script_links = self.extract_script_links(main_html)
        if not script_links:
            logger.error("No script links found")
            return

        # Filter to single script if in test mode
        if test_filter:
            original_count = len(script_links)
            script_links = [
                link for link in script_links if link["name"] == test_filter
            ]
            if not script_links:
                logger.error(
                    f"Test script '{test_filter}' not found in {original_count} available scripts"
                )
                return
            logger.info(
                f"Test mode: Filtered to 1 script out of {original_count} available"
            )

        # Load existing data to check which scripts we already have
        existing_df = self.load_existing_data()

        # Create efficient lookup for existing scripts
        # Map both project names and URL paths to existing scripts
        existing_scripts = {}
        url_to_project_name = {}

        if existing_df is not None:
            for _, row in existing_df.iterrows():
                project_name = row["Name"]
                community_url = row.get("Community URL", "")

                # Store by project name
                existing_scripts[project_name] = {
                    "community_url": community_url,
                    "has_author": pd.notna(row["Author"]) and row["Author"] != "",
                    "has_description": pd.notna(row["Description"])
                    and row["Description"] != "",
                    "has_tags": pd.notna(row["Tags"]) and row["Tags"] != "",
                    "has_demo": pd.notna(row["Demo"]) and row["Demo"] != "",
                    "has_discussion_url": pd.notna(row["Discussion URL"])
                    and row["Discussion URL"] != "",
                    "has_project_url": pd.notna(row["Project URL"])
                    and row["Project URL"] != "",
                    "playwright_status": (
                        row["Playwright Status"]
                        if "Playwright Status" in existing_df.columns
                        and pd.notna(row["Playwright Status"])
                        else ""
                    ),
                }

                # Create URL path to project name mapping
                if community_url:
                    url_path = community_url.replace(
                        "https://norns.community/", ""
                    ).strip("/")
                    url_to_project_name[url_path] = project_name

            logger.info(f"Found {len(existing_scripts)} existing scripts in Excel file")

        # Determine which scripts need scraping
        scripts_to_scrape = []
        scripts_skipped = 0

        for script_link in script_links:
            script_name = script_link["name"]  # URL path (e.g., "mxsynths")
            community_url = script_link["url"]

            # Use reverse lookup to find actual project name
            actual_project_name = url_to_project_name.get(script_name)

            if actual_project_name is None:
                # New script - needs full scraping
                scripts_to_scrape.append(script_link)
                logger.debug(f"New script '{script_name}' needs full scraping")
            else:
                existing = existing_scripts[actual_project_name]

                # Check if community URL differs (script moved/renamed)
                if existing["community_url"] != community_url:
                    scripts_to_scrape.append(script_link)
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) has different community URL - needs scraping"
                    )
                    continue

                # Check if all fields have data AND Playwright Status is not blank
                # Note: We need Playwright Status even if demo exists, to track resolution
                all_fields_complete = (
                    existing["has_author"]
                    and existing["has_description"]
                    and existing["has_tags"]
                    and existing["has_discussion_url"]
                    and existing["has_project_url"]
                )

                playwright_status_resolved = existing[
                    "playwright_status"
                ] != "" and not pd.isna(existing["playwright_status"])

                if all_fields_complete and playwright_status_resolved:
                    # Script is complete - skip entirely
                    scripts_skipped += 1
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) is complete - skipping"
                    )
                else:
                    # Script needs scraping for missing fields
                    scripts_to_scrape.append(script_link)
                    missing_fields = []
                    if not existing["has_author"]:
                        missing_fields.append("Author")
                    if not existing["has_description"]:
                        missing_fields.append("Description")
                    if not existing["has_tags"]:
                        missing_fields.append("Tags")
                    if not existing["has_discussion_url"]:
                        missing_fields.append("Discussion URL")
                    if not existing["has_project_url"]:
                        missing_fields.append("Project URL")
                    if not existing["has_demo"] and existing["playwright_status"] == "":
                        missing_fields.append("Demo")
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) needs scraping for: {', '.join(missing_fields)}"
                    )

        # Log efficiency summary
        total_scripts = len(script_links)
        scraping_count = len(scripts_to_scrape)

        logger.info(f"Efficiency analysis:")
        logger.info(f"  Total scripts found: {total_scripts}")
        logger.info(f"  Scripts needing scraping: {scraping_count}")
        logger.info(f"  Scripts skipped (already complete): {scripts_skipped}")

        if not scripts_to_scrape:
            logger.info("No scripts need scraping - all are complete")
            return

        logger.info(
            f"Starting parallel scraping of {scraping_count} scripts with {self.max_workers} workers..."
        )

        # Use ThreadPoolExecutor for parallel scraping
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks with existing data
            future_to_script = {}
            for script_link in scripts_to_scrape:
                script_name = script_link["name"]
                existing_data = None

                # Get existing data for this script if it exists
                if existing_df is not None:
                    # Prefer matching by Community URL path (most reliable)
                    try:
                        community_url_series = existing_df.get("Community URL")
                        if community_url_series is not None:
                            # Normalize to URL path for comparison
                            url_path_series = (
                                community_url_series.fillna("")
                                .astype(str)
                                .str.replace(
                                    "https://norns.community/", "", regex=False
                                )
                                .str.strip("/")
                            )
                            existing_row = existing_df[url_path_series == script_name]
                        else:
                            existing_row = existing_df[
                                existing_df["Name"] == script_name
                            ]
                    except Exception:
                        existing_row = existing_df[existing_df["Name"] == script_name]
                    if not existing_row.empty:
                        existing_data = existing_row.iloc[0].to_dict()

                future_to_script[
                    executor.submit(
                        self.scrape_script_details,
                        script_link["url"],
                        script_name,
                        existing_data,
                    )
                ] = script_link

            # Process completed tasks
            completed = 0
            for future in as_completed(future_to_script):
                script_link = future_to_script[future]
                completed += 1

                try:
                    script_data = future.result()
                    if script_data:
                        self.script_data.append(script_data)
                        logger.info(
                            f"Completed {completed}/{scraping_count}: {script_link['name']}"
                        )
                    else:
                        logger.warning(f"Failed to scrape {script_link['name']}")
                except Exception as e:
                    logger.error(f"Error processing {script_link['name']}: {e}")

        logger.info(
            f"Parallel scraping completed. Successfully scraped {len(self.script_data)} scripts."
        )

        # Handle demo discovery for scripts that need it
        self.discover_demos_unified(existing_df)

        # Retry failed demo requests if demo discovery was enabled
        if self.failed_demo_requests:
            self.retry_failed_demo_requests()

        # Resolve playwright conflicts if in playwright mode
        if self.playwright_conflicts:
            self.resolve_demo_conflicts()

    # ---------------------------
    # GitHub Last Updated helpers
    # ---------------------------
    def _load_github_token(self):
        """Load GitHub token from env var GITHUB_TOKEN or local gh.api file."""
        token = os.getenv("GITHUB_TOKEN", "").strip()
        if token:
            return token
        candidates = [
            "gh.api",
            os.path.join(os.path.dirname(__file__), "gh.api"),
        ]
        for path in candidates:
            try:
                if os.path.exists(path):
                    with open(path, "r", encoding="utf-8", errors="ignore") as f:
                        token = f.read().strip()
                        if token:
                            return token
            except Exception:
                continue
        return ""

    def _init_github_session(self):
        """Create a requests session for GitHub API with optional auth."""
        s = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=self.max_workers,
            pool_maxsize=self.max_workers * 2,
            max_retries=3,
        )
        s.mount("http://", adapter)
        s.mount("https://", adapter)
        headers = {
            "Accept": "application/vnd.github+json",
            "User-Agent": "NornsScraper-GitHub/1.0",
        }
        if self.github_token:
            headers["Authorization"] = f"Bearer {self.github_token}"
        s.headers.update(headers)
        return s

    def _parse_github_repo(self, url: str):
        """Return (owner, repo) if url points to a GitHub repo; else (None, None)."""
        try:
            from urllib.parse import urlparse

            parsed = urlparse(str(url))
            host = (parsed.netloc or "").lower()
            if not host.endswith("github.com"):
                return None, None
            parts = [p for p in (parsed.path or "").split("/") if p]
            if len(parts) < 2:
                return None, None
            owner = parts[0]
            repo = parts[1]
            if repo.endswith(".git"):
                repo = repo[:-4]
            return owner, repo
        except Exception:
            return None, None

    def _is_readme_only_change(self, files):
        """True if all changed files are README.md (case-insensitive basename match)."""
        try:
            import os as _os

            if not files:
                return False
            for f in files:
                name = str(f.get("filename", ""))
                base = _os.path.basename(name)
                if base.lower() != "readme.md":
                    return False
            return True
        except Exception:
            return False

    def _github_latest_non_readme_date(self, owner: str, repo: str) -> str:
        """Return YYYY-MM-DD of latest commit that isn't README.md-only; empty on failure."""
        if not owner or not repo:
            return ""
        base = f"https://api.github.com/repos/{owner}/{repo}"
        timeout = 15
        try:
            # Get default branch
            r = self.github_session.get(base, timeout=timeout)
            if r.status_code == 404:
                return ""
            r.raise_for_status()
            repo_info = r.json()
            default_branch = repo_info.get("default_branch") or "main"

            # Iterate commits by pages
            for page in range(1, 4):  # up to ~150 commits scanned max
                commits_resp = self.github_session.get(
                    f"{base}/commits",
                    params={"sha": default_branch, "per_page": 50, "page": page},
                    timeout=timeout,
                )
                # Handle rate limits gracefully
                if commits_resp.status_code == 403:
                    return ""
                commits_resp.raise_for_status()
                commits = commits_resp.json() or []
                if not commits:
                    break
                for c in commits:
                    sha = c.get("sha")
                    if not sha:
                        continue
                    detail_resp = self.github_session.get(
                        f"{base}/commits/{sha}", timeout=timeout
                    )
                    if detail_resp.status_code == 403:
                        return ""
                    if detail_resp.status_code == 404:
                        continue
                    detail_resp.raise_for_status()
                    detail = detail_resp.json()
                    files = detail.get("files", [])
                    if self._is_readme_only_change(files):
                        # skip README.md-only commits
                        continue
                    # Use committer date if available, else author date
                    commit_info = detail.get("commit", {})
                    committer = commit_info.get("committer") or {}
                    author = commit_info.get("author") or {}
                    date_str = committer.get("date") or author.get("date") or ""
                    if date_str:
                        try:
                            # normalize to YYYY-MM-DD
                            return str(date_str)[:10]
                        except Exception:
                            return ""
                # be gentle between pages
                time.sleep(0.2)
        except Exception as e:
            logger.debug(f"GitHub latest non-README date error for {owner}/{repo}: {e}")
            return ""
        return ""

    def _apply_last_updated(self, rows):
        """Enrich merged row dicts with 'Last Updated' using GitHub API, based on 'Project URL'."""
        if not rows:
            return rows
        # Build mapping repo -> row indices
        repo_to_indices = {}
        for idx, row in enumerate(rows):
            project_url = row.get("Project URL", "")
            owner, repo = self._parse_github_repo(project_url)
            if owner and repo:
                key = (owner, repo)
                repo_to_indices.setdefault(key, []).append(idx)

        if not repo_to_indices:
            return rows

        # Fetch in parallel
        results = {}

        def _task(owner_repo):
            owner, repo = owner_repo
            return self._github_latest_non_readme_date(owner, repo)

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_key = {
                executor.submit(_task, key): key for key in repo_to_indices.keys()
            }
            for future in as_completed(future_to_key):
                key = future_to_key[future]
                try:
                    results[key] = future.result() or ""
                except Exception:
                    results[key] = ""

        # Apply back to rows
        for key, indices in repo_to_indices.items():
            value = results.get(key, "")
            for idx in indices:
                # Always set/overwrite with the computed value (if any)
                rows[idx]["Last Updated"] = value
        return rows

    def _scrape_by_community_url(self, community_url: str):
        """Fetch single script page directly by community URL and extract comparable fields."""
        # Derive script_name from path for logging only
        try:
            script_name = community_url.replace("https://norns.community/", "").strip(
                "/"
            )
        except Exception:
            script_name = community_url
        return self.scrape_script_details(
            community_url, script_name, existing_data=None, discover_demo=False
        )

    def sync_check_only(self, excel_path: str) -> int:
        """Compute 'Out of Sync' for all rows without skipping logic; save back to Excel.

        Returns number of rows whose 'Out of Sync' value changed.
        """
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found: {excel_path}")
                return 0
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel for sync-check: {e}")
            return 0

        # Ensure columns exist
        if "Community URL" not in df.columns:
            logger.error("Excel missing 'Community URL' column")
            return 0
        if "Out of Sync" not in df.columns:
            df["Out of Sync"] = ""
        # Ensure dtype supports string assignment to avoid FutureWarning
        try:
            df["Out of Sync"] = (
                df["Out of Sync"]
                .astype("object")
                .where(pd.notna(df["Out of Sync"]), "")
            )
        except Exception:
            pass

        # Prepare tasks: only rows with a non-empty Community URL
        tasks = []
        for idx, row in df.iterrows():
            community_url = row.get("Community URL", "")
            if isinstance(community_url, str) and community_url.strip():
                tasks.append((idx, community_url.strip()))

        logger.info(f"Sync-check: processing {len(tasks)} row(s) from Excel")

        # Scrape in parallel
        scraped_by_idx = {}
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_idx = {
                executor.submit(self._scrape_by_community_url, url): idx
                for idx, url in tasks
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    scraped = future.result() or {}
                except Exception:
                    scraped = {}
                scraped_by_idx[idx] = scraped

        # Normalizers reused from merge
        def _norm_text(v: str) -> str:
            try:
                s = "" if pd.isna(v) else str(v)
            except Exception:
                s = str(v)
            s = s.strip().lower()
            s = re.sub(r"\s+", " ", s)
            return s

        def _norm_tags(v):
            try:
                if isinstance(v, (list, tuple, set)):
                    items = list(v)
                else:
                    s = "" if pd.isna(v) else str(v)
                    s = s.strip()
                    if s.startswith("[") and s.endswith("]"):
                        s = s[1:-1]
                    items = s.split(",") if s else []
                tokens = []
                for item in items:
                    t = ("" if item is None else str(item)).strip().strip("'\"").lower()
                    if t:
                        tokens.append(t)
                return tuple(sorted(set(tokens)))
            except Exception:
                return tuple()

        def _norm_authors(v):
            try:
                if isinstance(v, (list, tuple, set)):
                    items = list(v)
                else:
                    s = "" if pd.isna(v) else str(v)
                    items = s.split(",") if s else []
                tokens = []
                for item in items:
                    t = ("" if item is None else str(item)).strip().strip("'\"").lower()
                    if t:
                        tokens.append(t)
                return tuple(sorted(set(tokens)))
            except Exception:
                return tuple()

        def _norm_url(v: str) -> str:
            try:
                raw = "" if pd.isna(v) else str(v)
            except Exception:
                raw = str(v)
            raw = raw.strip()
            if not raw:
                return ""
            try:
                from urllib.parse import urlparse, urlunparse

                p = urlparse(raw)
                scheme = "https"
                netloc = (p.netloc or "").lower().replace("www.", "")
                path = (p.path or "").rstrip("/")
                if path.endswith(".git"):
                    path = path[:-4]
                return urlunparse((scheme, netloc, path, "", "", ""))
            except Exception:
                return raw.lower().rstrip("/")

        # Compute and apply Out of Sync
        changed = 0
        for idx, _ in tasks:
            existing = df.iloc[idx]
            scraped = scraped_by_idx.get(idx) or {}
            diffs = []

            # Name (flag on any difference, including one-sided presence)
            if _norm_text(existing.get("Name", "")) != _norm_text(
                scraped.get("project_name", "")
            ):
                diffs.append("Name")

            # Author (compare as sets; detect one-sided presence)
            authors_existing = _norm_authors(existing.get("Author", ""))
            authors_new = _norm_authors(scraped.get("author", ""))
            if authors_existing != authors_new:
                diffs.append("Author")

            # Tags (set compare, flag when either side differs including one-sided presence)
            tags_existing = _norm_tags(existing.get("Tags", ""))
            tags_scraped = _norm_tags(scraped.get("tags", ""))
            if tags_existing != tags_scraped:
                diffs.append("Tags")

            # Description (flag on any difference)
            if _norm_text(existing.get("Description", "")) != _norm_text(
                scraped.get("description", "")
            ):
                diffs.append("Description")

            # Project URL (flag on any difference)
            if _norm_url(existing.get("Project URL", "")) != _norm_url(
                scraped.get("project_url", "")
            ):
                diffs.append("Project URL")

            # Discussion URL (flag on any difference)
            if _norm_url(existing.get("Discussion URL", "")) != _norm_url(
                scraped.get("discussion_url", "")
            ):
                diffs.append("Discussion URL")

            new_value = ", ".join(diffs)
            old_value = existing.get("Out of Sync", "")
            if str(old_value) != str(new_value):
                df.at[idx, "Out of Sync"] = new_value
                changed += 1

        # Save back using the same formatting as main save path
        try:
            self._write_formatted_excel_from_df(df, excel_path)
        except Exception as e:
            logger.error(f"Failed to save formatted Excel after sync-check: {e}")
            try:
                df.to_excel(excel_path, index=False, engine="openpyxl")
            except Exception:
                pass
        return changed

    def discover_demos_unified(self, existing_df):
        """Single function to handle all demo discovery with both standard and Playwright methods"""
        if existing_df is None:
            logger.info("No existing data - skipping demo discovery")
            return

        # Create lookup for existing scripts
        existing_scripts = {}
        for _, row in existing_df.iterrows():
            script_name = row["Name"]
            existing_scripts[script_name] = {
                "has_demo": pd.notna(row["Demo"]) and row["Demo"] != "",
                "existing_demo_url": (
                    row["Demo"] if pd.notna(row["Demo"]) and row["Demo"] != "" else ""
                ),
                "playwright_status": (
                    row["Playwright Status"]
                    if "Playwright Status" in existing_df.columns
                    and pd.notna(row["Playwright Status"])
                    else ""
                ),
            }

        # Filter scraped scripts to those needing demo discovery
        scripts_needing_demos = []
        for script_data in self.script_data:
            project_name = script_data["project_name"]

            if project_name in existing_scripts:
                existing = existing_scripts[project_name]
                # Run demo discovery if:
                # 1. No demo URL exists, OR
                # 2. Playwright Status is blank (needs to be updated)
                needs_demo = (
                    (
                        not existing["has_demo"]
                        or existing["playwright_status"] == ""
                        or pd.isna(existing["playwright_status"])
                    )
                    and script_data["discussion_url"]
                    and pd.notna(script_data["discussion_url"])
                    and script_data["discussion_url"] != ""
                )
            else:
                # New script - check if it needs demo discovery
                needs_demo = (
                    script_data["discussion_url"]
                    and pd.notna(script_data["discussion_url"])
                    and script_data["discussion_url"] != ""
                )

            if needs_demo:
                scripts_needing_demos.append(
                    {
                        "name": project_name,
                        "discussion_url": script_data["discussion_url"],
                    }
                )

        if not scripts_needing_demos:
            logger.info("No scripts need demo discovery")
            return

        logger.info(
            f"Processing {len(scripts_needing_demos)} scripts for demo discovery with {self.max_workers} workers"
        )

        # Use ThreadPoolExecutor for parallel demo discovery
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks - both regular and playwright
            future_to_script = {}
            for script in scripts_needing_demos:
                discussion_url = script["discussion_url"]

                # Skip if discussion URL is invalid
                if (
                    not discussion_url
                    or pd.isna(discussion_url)
                    or discussion_url == ""
                ):
                    logger.debug(
                        f"Skipping {script['name']} - invalid discussion URL: {discussion_url}"
                    )
                    continue

                # Always submit regular discovery
                future_to_script[
                    executor.submit(self.discover_demo_video, discussion_url)
                ] = script

                # Also submit playwright discovery
                # Add delay between playwright requests to avoid rate limiting
                import random
                import time

                time.sleep(random.uniform(2.0, 5.0))

                future_to_script[
                    executor.submit(
                        self.discover_demo_video_playwright,
                        discussion_url,
                    )
                ] = {"script": script, "type": "playwright"}

            demos_found = 0
            completed = 0

            # Process completed tasks
            regular_results = {}  # Store regular discovery results
            playwright_results = {}  # Store playwright discovery results

            for future in as_completed(future_to_script):
                task_info = future_to_script[future]
                completed += 1

                try:
                    result_url = future.result()

                    if (
                        isinstance(task_info, dict)
                        and task_info.get("type") == "playwright"
                    ):
                        # This is a playwright result
                        script_name = task_info["script"]["name"]
                        playwright_results[script_name] = result_url
                    else:
                        # This is a regular result
                        script_name = task_info["name"]
                        regular_results[script_name] = result_url

                except Exception as e:
                    logger.warning(f"Error in demo discovery: {e}")

                # Progress update every 10 scripts
                if completed % 10 == 0 or completed == len(future_to_script):
                    logger.info(
                        f"Demo discovery progress: {completed}/{len(future_to_script)} tasks completed"
                    )

            # Now process results and detect conflicts
            for script in scripts_needing_demos:
                script_name = script["name"]
                discussion_url = script["discussion_url"]

                # Skip if discussion URL is invalid
                if (
                    not discussion_url
                    or pd.isna(discussion_url)
                    or discussion_url == ""
                ):
                    logger.debug(
                        f"Skipping {script_name} - invalid discussion URL: {discussion_url}"
                    )
                    continue

                regular_url = regular_results.get(script_name, "")
                playwright_url = playwright_results.get(script_name, "")

                # Get existing demo URL if this script exists in Excel
                existing_demo_url = ""
                if script_name in existing_scripts:
                    existing_demo_url = existing_scripts[script_name][
                        "existing_demo_url"
                    ]

                if regular_url or playwright_url:
                    # Check if existing demo URL matches either discovered URL
                    existing_matches_regular = (
                        existing_demo_url and existing_demo_url == regular_url
                    )
                    existing_matches_playwright = (
                        existing_demo_url and existing_demo_url == playwright_url
                    )

                    # If existing demo doesn't match either discovered URL, mark as Manual Override
                    if (
                        existing_demo_url
                        and not existing_matches_regular
                        and not existing_matches_playwright
                    ):
                        logger.info(
                            f"Existing demo URL for {script_name} doesn't match discovery methods - marking as Manual Override"
                        )
                        # Create a script data entry to update the Playwright Status
                        script_data = {
                            "project_name": script_name,
                            "project_url": "",  # We don't have this from the efficient list
                            "author": "",  # We don't have this from the efficient list
                            "description": "",  # We don't have this from the efficient list
                            "discussion_url": script["discussion_url"],
                            "tags": "",  # We don't have this from the efficient list
                            "demo": existing_demo_url,  # Keep existing demo URL
                            "community_url": "",  # We don't have this from the efficient list
                            "playwright_status": "Manual Override",
                        }
                        self.script_data.append(script_data)
                        continue

                    # Check for conflicts between regular and playwright discovery
                    if regular_url and playwright_url and regular_url != playwright_url:
                        logger.info(
                            f"Demo URL conflict for {script_name}: {regular_url} vs {playwright_url}"
                        )
                        self.playwright_conflicts.append(
                            {
                                "script_name": script_name,
                                "original_url": regular_url,
                                "playwright_url": playwright_url,
                            }
                        )
                        # Use regular URL for now, will be resolved later
                        final_url = regular_url
                        status = "Extract Preferred"  # Temporary until user resolves
                    else:
                        # Use whichever URL is available
                        final_url = playwright_url if playwright_url else regular_url

                        # Determine status based on what was found
                        if (
                            regular_url
                            and playwright_url
                            and regular_url == playwright_url
                        ):
                            status = "No Conflict"
                        elif playwright_url and not regular_url:
                            status = "Playwright Preferred"
                        elif regular_url and not playwright_url:
                            status = "Extract Preferred"
                        else:
                            status = "No Conflict"  # Fallback

                    # Check if this script is already in script_data (from scraping)
                    existing_script_data = None
                    for i, existing_data in enumerate(self.script_data):
                        if existing_data["project_name"] == script_name:
                            existing_script_data = existing_data
                            break

                    if existing_script_data:
                        # Update existing entry with demo and status
                        existing_script_data["demo"] = final_url
                        existing_script_data["playwright_status"] = status
                        logger.info(
                            f"Updated existing script data for {script_name} with demo and status"
                        )
                    else:
                        # Create a script data entry for this existing script
                        script_data = {
                            "project_name": script_name,
                            "project_url": "",  # We don't have this from the efficient list
                            "author": "",  # We don't have this from the efficient list
                            "description": "",  # We don't have this from the efficient list
                            "discussion_url": script["discussion_url"],
                            "tags": "",  # We don't have this from the efficient list
                            "demo": final_url,
                            "community_url": "",  # We don't have this from the efficient list
                            "playwright_status": status,
                        }
                        self.script_data.append(script_data)
                    demos_found += 1
                    logger.info(
                        f"Discovered demo for script {script_name}: {final_url}"
                    )
                else:
                    # No URLs discovered, but check if there's an existing demo URL
                    if existing_demo_url:
                        logger.info(
                            f"No demo URLs discovered for {script_name}, but existing demo URL found - marking as Manual Override"
                        )
                        # Check if this script is already in script_data (from scraping)
                        existing_script_data = None
                        for i, existing_data in enumerate(self.script_data):
                            if existing_data["project_name"] == script_name:
                                existing_script_data = existing_data
                                break

                        if existing_script_data:
                            # Update existing entry with status
                            existing_script_data["playwright_status"] = (
                                "Manual Override"
                            )
                            logger.info(
                                f"Updated existing script data for {script_name} with Manual Override status"
                            )
                        else:
                            # Create a script data entry to update the Playwright Status
                            script_data = {
                                "project_name": script_name,
                                "project_url": "",  # We don't have this from the efficient list
                                "author": "",  # We don't have this from the efficient list
                                "description": "",  # We don't have this from the efficient list
                                "discussion_url": script["discussion_url"],
                                "tags": "",  # We don't have this from the efficient list
                                "demo": existing_demo_url,  # Keep existing demo URL
                                "community_url": "",  # We don't have this from the efficient list
                                "playwright_status": "Manual Override",
                            }
                            self.script_data.append(script_data)
                    else:
                        logger.info(
                            f"No demo URLs discovered for {script_name} - marking as Missing Demo"
                        )
                        # Check if this script is already in script_data (from scraping)
                        existing_script_data = None
                        for i, existing_data in enumerate(self.script_data):
                            if existing_data["project_name"] == script_name:
                                existing_script_data = existing_data
                                break

                        if existing_script_data:
                            # Update existing entry with status
                            existing_script_data["playwright_status"] = "Missing Demo"
                            logger.info(
                                f"Updated existing script data for {script_name} with Missing Demo status"
                            )
                        else:
                            # Create a script data entry to update the Playwright Status
                            script_data = {
                                "project_name": script_name,
                                "project_url": "",  # We don't have this from the efficient list
                                "author": "",  # We don't have this from the efficient list
                                "description": "",  # We don't have this from the efficient list
                                "discussion_url": script["discussion_url"],
                                "tags": "",  # We don't have this from the efficient list
                                "demo": "",  # No demo URL found
                                "community_url": "",  # We don't have this from the efficient list
                                "playwright_status": "Missing Demo",
                            }
                            self.script_data.append(script_data)

        logger.info(f"Found {demos_found} demos for scripts needing discovery")

    def test_single_script(self, community_url):
        """Test mode: process only a single script by Community URL - just filter the main flow"""
        # Extract script name from URL
        if not community_url.startswith("https://norns.community/"):
            logger.error(
                f"Invalid Community URL: {community_url}. Must start with https://norns.community/"
            )
            return

        script_name = community_url.replace("https://norns.community/", "").strip("/")
        if not script_name:
            logger.error(f"Could not extract script name from URL: {community_url}")
            return

        logger.info(
            f"Test mode: Processing script '{script_name}' from {community_url}"
        )

        # Use the main scraping flow but filter to only process this one script
        self.scrape_all_scripts(test_filter=script_name)

    def discover_demo_video_playwright(self, discussion_url):
        """Discover demo video using Playwright to access actual page content"""
        if not discussion_url:
            return ""

        try:
            logger.debug(f"Checking discussion URL with Playwright: {discussion_url}")

            with sync_playwright() as p:
                # Launch browser with more realistic settings
                browser = p.chromium.launch(
                    headless=True,
                    args=[
                        "--no-sandbox",
                        "--disable-blink-features=AutomationControlled",
                        "--disable-dev-shm-usage",
                        "--disable-web-security",
                        "--disable-features=VizDisplayCompositor",
                        "--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    ],
                )

                # Create context with realistic settings
                context = browser.new_context(
                    viewport={"width": 1920, "height": 1080},
                    user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    locale="en-US",
                    timezone_id="America/New_York",
                    extra_http_headers={
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
                        "Accept-Language": "en-US,en;q=0.9",
                        "Accept-Encoding": "gzip, deflate, br",
                        "DNT": "1",
                        "Connection": "keep-alive",
                        "Upgrade-Insecure-Requests": "1",
                        "Sec-Fetch-Dest": "document",
                        "Sec-Fetch-Mode": "navigate",
                        "Sec-Fetch-Site": "none",
                        "Cache-Control": "max-age=0",
                    },
                )

                page = context.new_page()

                # Add realistic browser behavior
                page.add_init_script(
                    """
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined,
                    });
                    
                    Object.defineProperty(navigator, 'plugins', {
                        get: () => [1, 2, 3, 4, 5],
                    });
                    
                    Object.defineProperty(navigator, 'languages', {
                        get: () => ['en-US', 'en'],
                    });
                    
                    window.chrome = {
                        runtime: {},
                    };
                """
                )

                # Add random delay to mimic human behavior
                import random

                delay = random.uniform(1.0, 3.0)
                page.wait_for_timeout(int(delay * 1000))

                # Navigate with realistic timing
                try:
                    response = page.goto(
                        discussion_url, timeout=30000, wait_until="domcontentloaded"
                    )

                    if response and response.status >= 400:
                        logger.warning(
                            f"HTTP {response.status} error for {discussion_url}"
                        )
                        return ""

                    # Wait for content to load with realistic timing
                    page.wait_for_timeout(random.randint(2000, 5000))

                    # Try to wait for network to be idle, but with shorter timeout
                    try:
                        page.wait_for_load_state("networkidle", timeout=5000)
                    except:
                        # If networkidle times out, continue anyway
                        pass

                except Exception as e:
                    logger.warning(f"Navigation error for {discussion_url}: {e}")
                    return ""

                # Look for video/audio links in the page content
                demo_urls = []

                # Check for direct links
                links = page.query_selector_all("a[href]")
                for link in links:
                    href = link.get_attribute("href")
                    if href:
                        href_lower = href.lower()
                        if any(
                            pattern in href_lower
                            for pattern in [
                                "youtube.com/watch",
                                "youtu.be/",
                                "vimeo.com/",
                                "soundcloud.com/",
                                "instagram.com/",
                            ]
                        ):
                            demo_urls.append(href)

                # Check for iframe embeds
                iframes = page.query_selector_all("iframe[src]")
                for iframe in iframes:
                    src = iframe.get_attribute("src")
                    if src:
                        src_lower = src.lower()
                        if any(
                            pattern in src_lower
                            for pattern in [
                                "youtube.com/embed/",
                                "vimeo.com/video/",
                                "soundcloud.com/player",
                                "w.soundcloud.com",
                            ]
                        ):
                            demo_urls.append(src)

                context.close()
                browser.close()

                # Return the first valid demo URL found
                if demo_urls:
                    return demo_urls[0]

                logger.debug(f"No demo video found with Playwright: {discussion_url}")
                return ""

        except Exception as e:
            logger.warning(
                f"Error checking discussion URL with Playwright {discussion_url}: {e}"
            )
            return ""

    def resolve_demo_conflicts(self):
        """Present conflicts to user and resolve them using questionary"""
        if not self.playwright_conflicts:
            return

        logger.info(
            f"Found {len(self.playwright_conflicts)} demo URL conflicts to resolve"
        )

        for conflict in self.playwright_conflicts:
            script_name = conflict["script_name"]
            original_url = conflict["original_url"]
            playwright_url = conflict["playwright_url"]

            print(f"\n--- Demo URL Conflict for '{script_name}' ---")
            print(f"Original Discovery: {original_url}")
            print(f"Playwright Discovery: {playwright_url}")

            choice = questionary.select(
                f"Which demo URL would you like to use for '{script_name}'?",
                choices=[
                    questionary.Choice(
                        f"Original Discovery: {original_url}", "original"
                    ),
                    questionary.Choice(
                        f"Playwright Discovery: {playwright_url}", "playwright"
                    ),
                    questionary.Choice("Enter custom URL", "custom"),
                    questionary.Choice("Skip this script", "skip"),
                ],
            ).ask()

            if choice == "original":
                conflict["resolved_url"] = original_url
                conflict["status"] = "Extract Preferred"
            elif choice == "playwright":
                conflict["resolved_url"] = playwright_url
                conflict["status"] = "Playwright Preferred"
            elif choice == "custom":
                custom_url = questionary.text("Enter custom demo URL:").ask()
                if custom_url:
                    conflict["resolved_url"] = custom_url
                    conflict["status"] = "Manual Override"
                else:
                    conflict["resolved_url"] = original_url
                    conflict["status"] = "Extract Preferred"
            else:  # skip
                conflict["resolved_url"] = original_url
                conflict["status"] = "Extract Preferred"

        # Apply resolved conflicts to script data
        for conflict in self.playwright_conflicts:
            script_name = conflict["script_name"]
            resolved_url = conflict["resolved_url"]
            status = conflict["status"]

            # Find and update the script data
            for script in self.script_data:
                if script.get("project_name") == script_name:
                    script["demo"] = resolved_url
                    script["playwright_status"] = status
                    break

    def load_existing_data(self, filename="norns_scripts.xlsx"):
        """Load existing Excel data if it exists"""
        try:
            if os.path.exists(filename):
                existing_df = pd.read_excel(filename)
                logger.info(f"Loaded existing data with {len(existing_df)} scripts")
                return existing_df
            else:
                logger.info("No existing Excel file found, will create new one")
                return None
        except Exception as e:
            logger.warning(f"Could not load existing Excel file: {e}")
            return None

    def merge_data(self, new_data, existing_df=None):
        """Merge new scraped data with existing data, preserving manual corrections"""
        if existing_df is None:
            logger.info("No existing data to merge with")
            # Convert new data to Excel format
            excel_data = []
            for script in new_data:
                excel_script = {}
                for excel_col, internal_key in self.FIELD_MAP.items():
                    excel_script[excel_col] = script.get(internal_key, "")
                excel_data.append(excel_script)

            # Store summary stats for first run
            added_details = []
            for script in excel_data:
                populated_fields = []
                for excel_col in self.FIELD_MAP.keys():
                    if (
                        excel_col != "Playwright Status" and script[excel_col]
                    ):  # Skip Playwright Status for summary
                        populated_fields.append(excel_col)

                added_details.append(
                    {"name": script["Name"], "fields": populated_fields}
                )

            self.summary_stats = {
                "scripts_added": len(excel_data),
                "scripts_updated": 0,
                "scripts_preserved": 0,
                "total_scripts": len(excel_data),
                "added_details": added_details,
                "updated_details": [],
            }

            return excel_data

        logger.info(
            f"Merging {len(new_data)} new scripts with {len(existing_df)} existing scripts"
        )

        # Convert existing DataFrame to list of dicts for easier processing
        existing_scripts = existing_df.to_dict("records")

        # Create lookup dictionaries for existing scripts
        existing_by_name = {
            script.get("Name", ""): script for script in existing_scripts
        }
        existing_by_url = {}
        for script in existing_scripts:
            community_url = script.get("Community URL", "")
            if community_url and isinstance(community_url, str):
                url_path = community_url.replace("https://norns.community/", "").strip(
                    "/"
                )
                if url_path:
                    existing_by_url[url_path] = script

        merged_data = []
        added_count = 0
        updated_count = 0
        preserved_count = 0
        added_details = []
        updated_details = []

        # Helper to merge two row dicts (existing + new) with rules above
        def _merge_rows(existing_script, new_script):
            nonlocal preserved_count, updated_count
            merged_script = {}
            updated_fields = []
            for excel_col, internal_key in self.FIELD_MAP.items():
                existing_value = (
                    existing_script.get(excel_col, "") if existing_script else ""
                )
                new_value = new_script.get(internal_key, "")

                if excel_col == "Playwright Status":
                    if new_value is not None and str(new_value).strip() != "":
                        if str(existing_value) != str(new_value):
                            updated_fields.append(excel_col)
                        merged_script[excel_col] = new_value
                    else:
                        merged_script[excel_col] = existing_value
                    continue

                # Always prefer newly computed Last Updated when available
                if excel_col == "Last Updated":
                    if str(new_value).strip():
                        if str(existing_value) != str(new_value):
                            updated_fields.append(excel_col)
                        merged_script[excel_col] = new_value
                    else:
                        merged_script[excel_col] = existing_value
                    continue

                if (
                    pd.isna(existing_value)
                    or existing_value == ""
                    or str(existing_value).strip() == ""
                    or str(existing_value) == "nan"
                ):
                    merged_script[excel_col] = new_value
                    if new_value != "":
                        updated_fields.append(excel_col)
                else:
                    merged_script[excel_col] = existing_value
                    preserved_count += 1
            return merged_script, updated_fields

        # Helpers to determine out-of-sync for selected fields
        def _norm_text(v: str) -> str:
            try:
                s = "" if pd.isna(v) else str(v)
            except Exception:
                s = str(v)
            s = s.strip().lower()
            s = re.sub(r"\s+", " ", s)
            return s

        def _norm_tags(v: str):
            try:
                s = "" if pd.isna(v) else str(v)
            except Exception:
                s = str(v)
            # split by comma
            tokens = [t.strip().lower() for t in s.split(",")]
            tokens = [t for t in tokens if t]
            return tuple(sorted(set(tokens)))

        def _norm_url(v: str) -> str:
            try:
                raw = "" if pd.isna(v) else str(v)
            except Exception:
                raw = str(v)
            raw = raw.strip()
            if not raw:
                return ""
            try:
                from urllib.parse import urlparse, urlunparse

                p = urlparse(raw)
                scheme = "https"
                netloc = (p.netloc or "").lower().replace("www.", "")
                path = (p.path or "").rstrip("/")
                if path.endswith(".git"):
                    path = path[:-4]
                return urlunparse((scheme, netloc, path, "", "", ""))
            except Exception:
                return raw.lower().rstrip("/")

        def _compute_out_of_sync(existing_script, new_script):
            if not existing_script:
                return ""
            diffs = []
            # Name (flag on any difference, including one-sided presence)
            if _norm_text(existing_script.get("Name", "")) != _norm_text(
                new_script.get("project_name", "")
            ):
                diffs.append("Name")
            # Author (compare as sets; detect one-sided presence)
            authors_existing = _norm_authors(existing_script.get("Author", ""))
            authors_new = _norm_authors(new_script.get("author", ""))
            if authors_existing != authors_new:
                diffs.append("Author")
            # Tags (set compare, flag when either side differs including one-sided presence)
            tags_existing = _norm_tags(existing_script.get("Tags", ""))
            tags_new = _norm_tags(new_script.get("tags", ""))
            if tags_existing != tags_new:
                diffs.append("Tags")
            # Description (flag on any difference)
            if _norm_text(existing_script.get("Description", "")) != _norm_text(
                new_script.get("description", "")
            ):
                diffs.append("Description")
            # Project URL (flag on any difference)
            if _norm_url(existing_script.get("Project URL", "")) != _norm_url(
                new_script.get("project_url", "")
            ):
                diffs.append("Project URL")

            # Discussion URL (flag on any difference)
            if _norm_url(existing_script.get("Discussion URL", "")) != _norm_url(
                new_script.get("discussion_url", "")
            ):
                diffs.append("Discussion URL")
            return ", ".join(diffs)

        # Process new scraped data (prefer URL-path keyed matching)
        processed_keys = set()  # Prefer URL path as key; fallback to Name
        for new_script in new_data:
            script_name = new_script.get("project_name", "")
            community_url = new_script.get("community_url", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )

            existing_script = None
            key_used = None
            if url_key and url_key in existing_by_url:
                existing_script = existing_by_url[url_key]
                key_used = ("url", url_key)
            elif script_name in existing_by_name:
                existing_script = existing_by_name[script_name]
                key_used = ("name", script_name)

            if existing_script is not None:
                merged_script, updated_fields = _merge_rows(existing_script, new_script)
                # Compute Out of Sync vs scraped values
                merged_script["Out of Sync"] = _compute_out_of_sync(
                    existing_script, new_script
                )
                if updated_fields:
                    updated_count += 1
                    updated_details.append(
                        {"name": script_name, "fields": updated_fields}
                    )
                merged_data.append(merged_script)
                logger.debug(
                    f"Merged existing script by {'Community URL' if key_used and key_used[0]=='url' else 'Name'}: {script_name}"
                )
            else:
                # New script, add it
                merged_script = {}
                for excel_col, internal_key in self.FIELD_MAP.items():
                    merged_script[excel_col] = new_script.get(internal_key, "")
                # No comparison available for brand new items
                merged_script["Out of Sync"] = ""
                populated_fields = []
                for excel_col in self.FIELD_MAP.keys():
                    if excel_col != "Playwright Status" and merged_script[excel_col]:
                        populated_fields.append(excel_col)
                added_details.append({"name": script_name, "fields": populated_fields})
                merged_data.append(merged_script)
                added_count += 1
                logger.debug(f"Added new script: {script_name}")

            # Track processed key
            if url_key:
                processed_keys.add(("url", url_key))
            else:
                processed_keys.add(("name", script_name))

        # Add any existing scripts that weren't in the new data (key by URL path when present)
        for existing_script in existing_scripts:
            community_url = existing_script.get("Community URL", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )
            key = (
                ("url", url_key)
                if url_key
                else ("name", existing_script.get("Name", ""))
            )
            if key not in processed_keys:
                merged_data.append(existing_script)
                processed_keys.add(key)
                logger.debug(
                    f"Preserved existing script not in new data: {existing_script.get('Name','')}"
                )

        # Final deduplication by Community URL path to ensure no duplicates slip through
        deduped = {}
        order = []
        for row in merged_data:
            community_url = row.get("Community URL", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )
            if not url_key:
                # Keep rows without URL as-is keyed by name to avoid accidental drop
                url_key = f"__no_url__::{row.get('Name','')}"
            if url_key not in deduped:
                deduped[url_key] = row
                order.append(url_key)
            else:
                # Merge into existing using non-empty preference; Playwright Status pref new non-empty
                base = deduped[url_key]
                for excel_col in self.FIELD_MAP.keys():
                    if excel_col == "Playwright Status":
                        if str(row.get(excel_col, "")).strip():
                            base[excel_col] = row[excel_col]
                        continue
                    if (
                        not str(base.get(excel_col, "")).strip()
                        and str(row.get(excel_col, "")).strip()
                    ):
                        base[excel_col] = row[excel_col]
        merged_data = [deduped[k] for k in order]

        logger.info(
            f"Merge complete: {added_count} added, {updated_count} updated, {preserved_count} preserved"
        )

        # Store summary stats for final report
        self.summary_stats = {
            "scripts_added": added_count,
            "scripts_updated": updated_count,
            "scripts_preserved": preserved_count,
            "total_scripts": len(merged_data),
            "added_details": added_details,
            "updated_details": updated_details,
        }

        return merged_data

    def parse_statuses_from_log(self, log_path):
        """Parse Playwright Status assignments from a run log.

        Returns a dict: script name -> status.
        Heuristics:
          - Explicit lines for Missing Demo and Manual Override are authoritative.
          - "Demo URL conflict for <name>: ... vs ..." implies Extract Preferred.
          - For remaining "Discovered demo for script <name>: <url>" without a preceding conflict,
            mark as Playwright Preferred when the URL looks like an embed-only provider
            (player.vimeo.com, w.soundcloud.com, youtube.com/embed), otherwise No Conflict.
        """
        status_by_name = {}
        try:
            with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()
        except Exception as e:
            logger.error(f"Failed to read log file for status parsing: {e}")
            return status_by_name

        import re as _re

        # Precompile regexes
        rx_missing = _re.compile(
            r"Updated existing script data for (.+?) with Missing Demo status"
        )
        rx_manual = _re.compile(
            r"Updated existing script data for (.+?) with Manual Override status"
        )
        rx_conflict = _re.compile(r"Demo URL conflict for (.+?): ")
        rx_discovered = _re.compile(r"Discovered demo for script (.+?):\s*(\S+)")

        # First pass: direct status lines
        for line in lines:
            m = rx_missing.search(line)
            if m:
                status_by_name[m.group(1)] = "Missing Demo"
                continue
            m = rx_manual.search(line)
            if m:
                status_by_name[m.group(1)] = "Manual Override"
                continue
            m = rx_conflict.search(line)
            if m:
                # Conflict defaults to Extract Preferred in code until user resolution
                name = m.group(1)
                # Don't overwrite explicit statuses already parsed later
                status_by_name.setdefault(name, "Extract Preferred")

        # Second pass: discovered demos without explicit conflict/manual/missing
        # Build a set of names that already have explicit statuses
        already_set = set(status_by_name.keys())
        for line in lines:
            m = rx_discovered.search(line)
            if not m:
                continue
            name = m.group(1)
            url = m.group(2)
            if name in already_set:
                continue

            lower_url = url.lower()
            if (
                "player.vimeo.com/" in lower_url
                or "w.soundcloud.com/" in lower_url
                or "youtube.com/embed/" in lower_url
            ):
                status_by_name[name] = "Playwright Preferred"
            else:
                # Could be No Conflict or Extract Preferred; favor No Conflict when not identifiable
                status_by_name[name] = "No Conflict"

        logger.info(
            f"Parsed {len(status_by_name)} Playwright Status assignments from log"
        )
        return status_by_name

    def apply_status_updates_from_log(self, log_path, excel_path="norns_scripts.xlsx"):
        """Apply Playwright Status updates parsed from a log file directly to the Excel sheet."""
        status_map = self.parse_statuses_from_log(log_path)
        if not status_map:
            logger.warning("No statuses parsed from log; nothing to apply")
            return 0

        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found: {excel_path}")
                return 0
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel file for applying statuses: {e}")
            return 0

        if "Name" not in df.columns:
            logger.error("Excel file missing required 'Name' column")
            return 0

        # Ensure Playwright Status column exists
        if "Playwright Status" not in df.columns:
            df["Playwright Status"] = ""

        # Ensure compatible dtype to avoid pandas warning when assigning strings
        try:
            df["Playwright Status"] = (
                df["Playwright Status"]
                .astype("object")
                .where(pd.notna(df["Playwright Status"]), "")
            )
        except Exception:
            # Fallback: leave as-is if conversion fails
            pass

        # Build a normalization helper for better matching tolerance
        import re as _re2
        import unicodedata as _unicodedata

        def _normalize_name(name: str) -> str:
            if not isinstance(name, str):
                name = "" if pd.isna(name) else str(name)
            # Normalize unicode (convert curly quotes to ASCII equivalents)
            name = (
                name.replace("’", "'")
                .replace("‘", "'")
                .replace("“", '"')
                .replace("”", '"')
            )
            name = _unicodedata.normalize("NFKD", name)
            # Lower, remove diacritics by encoding to ASCII
            try:
                name = name.encode("ascii", "ignore").decode("ascii")
            except Exception:
                pass
            name = name.lower()
            # Collapse punctuation and whitespace to single hyphens/spaces for robust matching
            name = _re2.sub(r"[^a-z0-9\s_\-]", "", name)
            name = _re2.sub(r"[\s\-]+", "-", name).strip("-")
            return name

        # Build fast lookup maps for exact and normalized names
        exact_to_index = {}
        normalized_to_index = {}
        for idx, row in df.iterrows():
            nm = row.get("Name", "")
            exact_to_index[str(nm)] = idx
            normalized_to_index[_normalize_name(str(nm))] = idx

        # Apply updates
        updated_rows = 0
        for raw_name, new_status in status_map.items():
            target_idx = None
            # Prefer exact match first
            if raw_name in exact_to_index:
                target_idx = exact_to_index[raw_name]
            else:
                # Try normalized name match
                norm = _normalize_name(raw_name)
                if norm in normalized_to_index:
                    target_idx = normalized_to_index[norm]

            if target_idx is None:
                continue

            current_value = df.at[target_idx, "Playwright Status"]
            if str(current_value) != str(new_status):
                df.at[target_idx, "Playwright Status"] = new_status
                updated_rows += 1

        if updated_rows == 0:
            logger.info("No Playwright Status cells changed based on log parsing")
            return 0

        # Save back using pandas to minimize side-effects
        try:
            df.to_excel(excel_path, index=False, engine="openpyxl")
            logger.info(
                f"Applied {updated_rows} Playwright Status update(s) from log to {excel_path}"
            )
        except Exception as e:
            logger.error(f"Failed saving Excel after applying statuses: {e}")
            return 0

        return updated_rows

    def _write_formatted_excel_from_df(self, df: pd.DataFrame, filename: str):
        """Write the given DataFrame to Excel with the project's formatting and table styling."""
        # Ensure all required columns exist
        required_columns = list(self.FIELD_MAP.keys())
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""

        # Sort by Name
        df = df.sort_values(by="Name", na_position="last").reset_index(drop=True)

        # Create formatted workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Norns Scripts"

        # Headers
        headers = list(self.FIELD_MAP.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=14)

        # Rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            # Name
            cell = ws.cell(row=row_idx, column=1, value=row["Name"])
            cell.font = Font(size=14)

            # Author
            cell = ws.cell(row=row_idx, column=2, value=row["Author"])
            cell.font = Font(size=14)

            # Tags
            cell = ws.cell(row=row_idx, column=3, value=row["Tags"])
            cell.font = Font(size=14)

            # Description
            cell = ws.cell(row=row_idx, column=4, value=row["Description"])
            cell.font = Font(size=14)

            # Demo
            demo_value = row["Demo"] if pd.notna(row["Demo"]) else ""
            if demo_value and str(demo_value).strip() and str(demo_value) != "nan":
                cell = ws.cell(row=row_idx, column=5, value=demo_value)
                cell.hyperlink = str(demo_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=5, value=demo_value)
                cell.font = Font(size=14)

            # Discussion URL
            discussion_value = (
                row["Discussion URL"] if pd.notna(row["Discussion URL"]) else ""
            )
            if discussion_value and str(discussion_value).strip():
                cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                cell.hyperlink = str(discussion_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                cell.font = Font(size=14)

            # Project URL
            project_value = row["Project URL"] if pd.notna(row["Project URL"]) else ""
            if project_value and str(project_value).strip():
                cell = ws.cell(row=row_idx, column=7, value=project_value)
                cell.hyperlink = str(project_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=7, value=project_value)
                cell.font = Font(size=14)

            # Community URL
            community_value = (
                row["Community URL"] if pd.notna(row["Community URL"]) else ""
            )
            if community_value and str(community_value).strip():
                cell = ws.cell(row=row_idx, column=8, value=community_value)
                cell.hyperlink = str(community_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=8, value=community_value)
                cell.font = Font(size=14)

            # Playwright Status
            playwright_status = (
                row["Playwright Status"] if pd.notna(row["Playwright Status"]) else ""
            )
            cell = ws.cell(row=row_idx, column=9, value=playwright_status)
            cell.font = Font(size=14)

            # Last Updated
            last_updated_value = (
                row["Last Updated"] if pd.notna(row.get("Last Updated")) else ""
            )
            cell = ws.cell(row=row_idx, column=10, value=last_updated_value)
            cell.font = Font(size=14)

            # Out of Sync
            out_of_sync_value = (
                row["Out of Sync"] if pd.notna(row.get("Out of Sync")) else ""
            )
            cell = ws.cell(row=row_idx, column=11, value=out_of_sync_value)
            cell.font = Font(size=14)

        # Auto sizing
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        content_length = len(str(cell.value))
                        adjusted_length = int(content_length * 1.2)
                        if adjusted_length > max_length:
                            max_length = adjusted_length
                except:
                    pass
            adjusted_width = max(min(max_length + 2, 80), 10)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Table
        last_row = len(df) + 1
        table_range = f"A1:K{last_row}"
        table = Table(displayName="NornsScripts", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleDark11",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Save
        wb.save(filename)

    def save_to_excel(self, filename="norns_scripts.xlsx"):
        """Save scraped data to Excel file with clickable hyperlinks, merging with existing data"""
        if not self.script_data:
            logger.error("No data to save")
            return

        logger.info(f"Processing {len(self.script_data)} scraped scripts")

        # Convert tags list to comma-separated string for Excel
        for script in self.script_data:
            script["tags"] = ", ".join(script["tags"])

        # Load existing data and merge
        existing_df = self.load_existing_data(filename)
        merged_data = self.merge_data(self.script_data, existing_df)

        # Compute GitHub-based Last Updated for all rows (based on Project URL)
        try:
            merged_data = self._apply_last_updated(merged_data)
        except Exception as e:
            logger.warning(f"Failed applying Last Updated enrichment: {e}")

        if not merged_data:
            logger.error("No data to save after merging")
            return

        # Create DataFrame from merged data
        df = pd.DataFrame(merged_data)

        # Ensure all required columns exist
        required_columns = list(self.FIELD_MAP.keys())

        for col in required_columns:
            if col not in df.columns:
                df[col] = ""

        # Sort DataFrame by Name column alphabetically
        df = df.sort_values(by="Name", na_position="last").reset_index(drop=True)

        # Create Excel workbook with hyperlinks
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Norns Scripts"

            # Add headers
            headers = list(self.FIELD_MAP.keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=14)

            # Add data with hyperlinks
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                # Name
                cell = ws.cell(row=row_idx, column=1, value=row["Name"])
                cell.font = Font(size=14)

                # Author
                cell = ws.cell(row=row_idx, column=2, value=row["Author"])
                cell.font = Font(size=14)

                # Tags
                cell = ws.cell(row=row_idx, column=3, value=row["Tags"])
                cell.font = Font(size=14)

                # Description
                cell = ws.cell(row=row_idx, column=4, value=row["Description"])
                cell.font = Font(size=14)

                # Demo (as hyperlink if URL found)
                demo_value = row["Demo"] if pd.notna(row["Demo"]) else ""
                if demo_value and str(demo_value).strip() and str(demo_value) != "nan":
                    cell = ws.cell(row=row_idx, column=5, value=demo_value)
                    cell.hyperlink = str(demo_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=5, value=demo_value)
                    cell.font = Font(size=14)

                # Discussion URL (as hyperlink)
                discussion_value = (
                    row["Discussion URL"] if pd.notna(row["Discussion URL"]) else ""
                )
                if discussion_value and str(discussion_value).strip():
                    cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                    cell.hyperlink = str(discussion_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                    cell.font = Font(size=14)

                # Project URL (as hyperlink)
                project_value = (
                    row["Project URL"] if pd.notna(row["Project URL"]) else ""
                )
                if project_value and str(project_value).strip():
                    cell = ws.cell(row=row_idx, column=7, value=project_value)
                    cell.hyperlink = str(project_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=7, value=project_value)
                    cell.font = Font(size=14)

                # Community URL (as hyperlink)
                community_value = (
                    row["Community URL"] if pd.notna(row["Community URL"]) else ""
                )
                if community_value and str(community_value).strip():
                    cell = ws.cell(row=row_idx, column=8, value=community_value)
                    cell.hyperlink = str(community_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=8, value=community_value)
                    cell.font = Font(size=14)

                # Playwright Status (text only)
                playwright_status = (
                    row["Playwright Status"]
                    if pd.notna(row["Playwright Status"])
                    else ""
                )
                cell = ws.cell(row=row_idx, column=9, value=playwright_status)
                cell.font = Font(size=14)

                # Last Updated (text only, YYYY-MM-DD)
                last_updated_value = (
                    row["Last Updated"] if pd.notna(row.get("Last Updated")) else ""
                )
                cell = ws.cell(row=row_idx, column=10, value=last_updated_value)
                cell.font = Font(size=14)

                # Out of Sync (text only)
                out_of_sync_value = (
                    row["Out of Sync"] if pd.notna(row.get("Out of Sync")) else ""
                )
                cell = ws.cell(row=row_idx, column=11, value=out_of_sync_value)
                cell.font = Font(size=14)

            # Auto-adjust column widths to fit all content
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            # Calculate length considering font size and content
                            content_length = len(str(cell.value))
                            # Adjust for font size 14 (roughly 1.2x multiplier)
                            adjusted_length = int(content_length * 1.2)
                            if adjusted_length > max_length:
                                max_length = adjusted_length
                    except:
                        pass
                # Set width with minimum and maximum bounds
                adjusted_width = max(min(max_length + 2, 80), 10)  # Min 10, Max 80
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create table with Orange Table Style Dark 11
            last_row = len(df) + 1
            table_range = f"A1:K{last_row}"  # Updated to include Out of Sync column
            table = Table(displayName="NornsScripts", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleDark11",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # Freeze the top row
            ws.freeze_panes = "A2"

            # Save the workbook
            wb.save(filename)
            logger.info(f"Successfully saved to {filename} with clickable hyperlinks")

        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
            # Fallback to pandas method if openpyxl fails
            try:
                df.to_excel(filename, index=False, engine="openpyxl")
                logger.info(f"Fallback save successful: {filename}")
            except Exception as e2:
                logger.error(f"Fallback save also failed: {e2}")

    def print_summary(self):
        """Print a summary of what was accomplished during scraping"""
        print("\n" + "=" * 60)
        print("SCRAPING SUMMARY")
        print("=" * 60)
        print(f"Scripts scraped from norns.community: {len(self.script_data)}")

        if hasattr(self, "summary_stats") and self.summary_stats:
            print(f"Scripts added to Excel: {self.summary_stats['scripts_added']}")
            print(f"Scripts updated in Excel: {self.summary_stats['scripts_updated']}")
            print(
                f"Scripts preserved (manual edits): {self.summary_stats['scripts_preserved']}"
            )
            print(f"Total scripts in Excel file: {self.summary_stats['total_scripts']}")

            # Show detailed information about added scripts
            if self.summary_stats.get("added_details"):
                print("\nAdded scripts:")
                for detail in self.summary_stats["added_details"]:
                    fields_str = (
                        ", ".join(detail["fields"]) if detail["fields"] else "no fields"
                    )
                    print(
                        f"  Added script '{detail['name']}' with field(s): {fields_str}"
                    )

            # Show detailed information about updated scripts
            if self.summary_stats.get("updated_details"):
                print("\nUpdated scripts:")
                for detail in self.summary_stats["updated_details"]:
                    fields_str = (
                        ", ".join(detail["fields"]) if detail["fields"] else "no fields"
                    )
                    print(
                        f"  Updated script '{detail['name']}' with field(s): {fields_str}"
                    )

        # Count discovered demos
        demos_found = sum(
            1 for script in self.script_data if script.get("demo", "").strip()
        )
        print(f"\nDemo discovery: Found {demos_found} demo videos")

        print("=" * 60)


def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="Scrape norns.community scripts")
    parser.add_argument(
        "--workers",
        type=int,
        default=10,
        help="Number of parallel workers (default: 10)",
    )
    parser.add_argument(
        "--demo-delay",
        type=float,
        default=0.5,
        help="Delay in seconds between demo discovery requests (default: 0.5)",
    )
    parser.add_argument(
        "--test",
        type=str,
        help="Test mode: process only the specified Community URL (e.g., https://norns.community/scriptname)",
    )
    parser.add_argument(
        "--status-log",
        type=str,
        help="Apply Playwright Status updates by parsing the given run log file and updating the Excel sheet",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="norns_scripts.xlsx",
        help="Target Excel file to read/write (default: norns_scripts.xlsx)",
    )
    parser.add_argument(
        "--dedupe",
        action="store_true",
        help="Only deduplicate the target Excel file by Community URL path and exit",
    )
    parser.add_argument(
        "--sync-check",
        action="store_true",
        help="Only compute and update 'Out of Sync' for all rows without skipping logic",
    )

    args = parser.parse_args()

    # You can adjust max_workers based on your system and network
    # More workers = faster scraping, but be respectful to the server
    scraper = NornsScraper(
        max_workers=args.workers,
        demo_delay=args.demo_delay,
    )

    # Optional fast-path: apply statuses from a log and exit
    if args.status_log:
        updated = scraper.apply_status_updates_from_log(args.status_log, args.excel)
        if updated:
            logger.info(
                f"Status application complete. {updated} row(s) updated in {args.excel}"
            )
        else:
            logger.info("No rows updated from status log application")
        return

    # Optional fast-path: compute Out of Sync only and exit
    if args.sync_check:
        try:
            updated = scraper.sync_check_only(args.excel)
            if updated:
                logger.info(
                    f"Sync-check complete. Updated 'Out of Sync' for {updated} row(s) in {args.excel}"
                )
            else:
                logger.info("Sync-check completed; no changes were necessary")
        except Exception as e:
            logger.error(f"Sync-check failed: {e}")
        return

    # Optional fast-path: deduplicate an existing Excel file and exit
    if args.dedupe:
        try:
            if not os.path.exists(args.excel):
                logger.error(f"Excel file not found: {args.excel}")
                return
            df = pd.read_excel(args.excel)
        except Exception as e:
            logger.error(f"Failed to load Excel for dedupe: {e}")
            return

        before = len(df)

        # Derive URL path key
        def url_key(url: str) -> str:
            if not isinstance(url, str):
                return ""
            return url.replace("https://norns.community/", "").strip("/")

        df["__key__"] = df.get("Community URL", "").apply(url_key)
        # Keep first by URL key; for empties, de-dupe by Name
        non_empty = df[df["__key__"] != ""].copy()
        empty = df[df["__key__"] == ""].copy()
        non_empty = non_empty.drop_duplicates(subset=["__key__"], keep="first")
        empty = empty.drop_duplicates(subset=["Name"], keep="first")
        out = pd.concat([non_empty, empty], ignore_index=True).drop(
            columns=["__key__"], errors="ignore"
        )
        after = len(out)
        removed = before - after
        try:
            out.to_excel(args.excel, index=False, engine="openpyxl")
            logger.info(
                f"Deduplicated {removed} row(s) by Community URL path; saved to {args.excel}"
            )
        except Exception as e:
            logger.error(f"Failed saving Excel after dedupe: {e}")
        return

    if args.test:
        logger.info(f"Test mode: Processing single script from {args.test}")
        scraper.test_single_script(args.test)
    else:
        logger.info(
            f"Starting norns.community scraper with parallel processing, demo discovery, and Playwright..."
        )
        scraper.scrape_all_scripts()

    if scraper.script_data:
        scraper.save_to_excel(args.excel)
        logger.info(f"Scraping complete! Found {len(scraper.script_data)} scripts.")
        scraper.print_summary()
    else:
        logger.error("No scripts were scraped successfully")


if __name__ == "__main__":
    main()
