#!/usr/bin/env python3
"""
Norns Community Script Scraper — Discourse-API variant (no Playwright).

This is the experimental B-track of the scraper. It drops the Playwright
fallback entirely and instead queries Discourse's JSON API for demo discovery
on llllllll.co. Two consequences:

  1. No more browser dependency / no more `playwright install`. No preflight.
  2. We can scan ALL posts in a thread (post_stream pagination), not just the
     first ~20 the HTML page renders before its scroll-to-load kicks in.

Run side-by-side with the original norns_scraper.py for A/B comparison; this
file's default excel_path is `norns_scripts_discourse.xlsx` so the outputs
don't collide.
"""

import argparse
import logging
import os
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

import pandas as pd
import questionary
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Set up logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class NornsScraper:
    # Field mapping: Excel column -> internal key
    FIELD_MAP = {
        "Name": "project_name",
        "Author": "author",
        "Tags": "tags",
        "Description": "description",
        "Demo": "demo",
        "Discussion URL": "discussion_url",
        "Project URL": "project_url",
        "Documentation URL": "documentation_url",
        "Community URL": "community_url",
        "Playwright Status": "playwright_status",
        "Last Updated": "last_updated",
        "Out of Sync": "out_of_sync",
    }

    # Playwright status values
    RESOLVED_STATUSES = [
        "No Conflict",
        "Playwright Preferred",
        "Extract Preferred",
        "Manual Override",
        "Missing Demo",
    ]

    # Source of truth for the script catalog: monome/norns-community/community.json
    # The norns.community website is regenerated from this file, so reading it
    # directly is faster, more reliable, and immune to website redesigns.
    COMMUNITY_JSON_URL = (
        "https://raw.githubusercontent.com/monome/norns-community/main/community.json"
    )

    # Fields we expect on each community.json entry. Anything else is reported
    # as a schema deviation at the end of a run so it can be PR'd upstream.
    EXPECTED_JSON_FIELDS = {
        "project_name",
        "project_url",
        "author",
        "description",
        "discussion_url",
        "documentation_url",
        "tags",
    }

    # Known typos in community.json mapped to their intended field name.
    # Used both for defensive reads (so data isn't lost) and for the deviation
    # report (so the user can PR a fix upstream).
    KNOWN_JSON_TYPOS = {
        "tag": "tags",
        "documen0tation_url": "documentation_url",
    }

    # Fields tracked for snapshot-based drift detection. Each tuple is
    # (Excel column, internal/JSON key, name of the static normalizer method).
    # Demo, Last Updated, Playwright Status, Out of Sync, and Community URL are
    # not tracked — they're either workflow state or computed values, not facts
    # we're trying to detect upstream changes in.
    TRACKED_DRIFT_FIELDS = (
        ("Name", "project_name", "_norm_text"),
        ("Author", "author", "_norm_authors"),
        ("Tags", "tags", "_norm_tags"),
        ("Description", "description", "_norm_text"),
        ("Project URL", "project_url", "_norm_url"),
        ("Discussion URL", "discussion_url", "_norm_url"),
        ("Documentation URL", "documentation_url", "_norm_url"),
    )

    # Locally-curated tags we inject into the xlsx Tags column that do NOT exist
    # in community.json (e.g. "norns only"). _norm_tags drops these before drift
    # comparison so their presence can never trigger a false "Out of Sync" — on
    # either side of the diff. They still display in the Tags column verbatim and
    # are preserved by the merge's non-empty-cell rule.
    LOCAL_CURATION_TAGS = frozenset({"norns only"})

    def __init__(
        self,
        base_url="https://norns.community",
        max_workers=10,
        demo_delay=0.5,
        excel_path="norns_scripts_discourse.xlsx",
    ):
        self.base_url = base_url
        self.max_workers = max_workers
        # Stored so sidecar files (snapshots, discourse_resolutions) can be
        # found without threading the path through every code path.
        self.excel_path = excel_path
        # Lazily-probed flag: is Vimeo oembed usable from this environment?
        # (None = not yet probed.) See _vimeo_oembed_usable.
        self._vimeo_oembed_ok = None
        self.session = requests.Session()

        # Configure connection pool to handle high concurrency
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=self.max_workers,
            pool_maxsize=self.max_workers * 2,
            max_retries=3,
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
        )
        self.script_data = []
        self.demo_delay = demo_delay
        self.failed_demo_requests = []  # Track failed requests for retry
        # Note: playwright_conflicts list removed — Discourse variant has only
        # one discovery method, so there are no conflicts to reconcile. The
        # "Playwright Status" Excel column name is kept for A/B parity with the
        # original scraper; this variant only emits "Missing Demo" / "Manual
        # Override" / "" into it.
        self._community_json = None  # Lazy cache for community.json entries
        self._slug_map = None  # Lazy cache for project_name -> URL slug
        self.schema_deviations = []  # Typos/unknown JSON fields, reported at end-of-run
        # Discourse URL resolution cache: stale_url -> resolved_url (or "" if
        # resolution failed). Loaded at scrape_all_scripts start, saved at end.
        self._discourse_resolutions = {}
        self._discourse_resolutions_lock = threading.Lock()
        self._discourse_resolution_log = []  # New resolutions made this run, for report

        # Global rate limiter for Discourse requests. With 8 worker threads
        # we'd otherwise burst-fire and hit llllllll.co's anonymous-user rate
        # limit (HTTP 429). Discourse's default is ~60 req/min (1 RPS); we cap
        # at 2 RPS sustained to leave headroom and avoid bursts that trigger
        # short-term throttling. This run is intended to be daily-CI accuracy,
        # so we trade wall time for reliability.
        self._discourse_rps = 2.0
        self._discourse_min_interval = 1.0 / self._discourse_rps
        self._discourse_last_request = 0.0
        self._discourse_throttle_lock = threading.Lock()
        self.summary_stats = {
            "scripts_added": 0,
            "scripts_updated": 0,
            "scripts_preserved": 0,
            "total_scripts": 0,
            "added_details": [],
            "updated_details": [],
        }

        # GitHub integration
        self.github_token = self._load_github_token()
        self.github_session = self._init_github_session()

        # feed.json generation (ingenue B7). Defaults so save_to_excel emits the
        # feed on a full run even when driven outside main() (e.g. tests); main()
        # overrides these from --no-feed / --feed-output.
        self.feed_enabled = True
        self.feed_output = None
        # catalog.json (canonical full-rows catalog; build_data.py prefers it).
        self.catalog_enabled = True
        self.catalog_output = None
        # GitHub discovery (roadmap #3): off by default — opt in with --discover.
        self.discover_enabled = False
        self.discover_aggressive = True
        self.discover_max_authors = None

        # --- Efficiency: per-run memoization (avoids redundant GitHub/index calls) ---
        # Cached HTML of the norns.community index page. It was fetched twice per
        # run (fetch_slug_map + scrape_all_scripts); it is static within a run.
        self._main_page_html = None
        # One `GET /repos/{owner}/{repo}` per repo per run, shared between the
        # Last-Updated pass and feed enrichment (both need default_branch; the
        # Last-Updated cache also needs pushed_at). Keyed by (owner, repo).
        self._repo_meta_cache = {}
        self._repo_meta_lock = threading.Lock()
        # GitHub search API is ~30 req/min authenticated; the discovery sweep
        # fires many searches, so throttle every search request to stay under it.
        self._search_last = 0.0
        self._search_lock = threading.Lock()
        self._search_min_interval = 2.1

    def discover_demo_video(
        self, discussion_url, _no_resolve=False, author="", script_name=""
    ):
        """Discover demo video from a discussion URL.

        Dispatches based on URL shape:
          - Discourse forum URLs (matches /t/<slug>[/<id>]) -> JSON API path,
            which uses 4-phase priority: early-post video → external search →
            rest-of-thread video → audio fallback.
          - Anything else -> original HTML-page scrape (fallback for non-Discourse).

        author/script_name are required for Phase 2 external search; they're
        passed through from scrape_script_details which has them from
        community.json. Optional for backward-compat; external search is
        skipped if not provided.

        Cache pre-check still applies to both branches.
        """
        if not discussion_url:
            return ""

        # Pre-check the resolution sidecar before doing any HTTP. If we already
        # know this URL is stale, jump straight to the cached replacement (or
        # bail if we previously tried and failed) — saves the wasted 404 call.
        if not _no_resolve:
            cached = self._discourse_resolutions.get(discussion_url)
            if cached is not None:
                if cached == "":
                    logger.debug(
                        f"Skipping known-unresolvable Discourse URL: {discussion_url}"
                    )
                    return ""
                if cached != discussion_url:
                    logger.debug(
                        f"Using cached Discourse resolution: {discussion_url} -> {cached}"
                    )
                    return self.discover_demo_video(
                        cached,
                        _no_resolve=True,
                        author=author,
                        script_name=script_name,
                    )

        # Discourse URL -> use JSON API for full-thread pagination
        m = self._DISCOURSE_TOPIC_RE.match(discussion_url)
        if m:
            return self._discover_demo_via_discourse_api(
                discussion_url,
                m.group(1),
                m.group(2),
                m.group(3),
                m.group(4),
                author=author,
                script_name=script_name,
            )

        # Non-Discourse URL (e.g. short-link redirects like l.llllllll.co/...).
        # In practice, every entry in community.json today resolves through
        # the Discourse path above, so this branch is rarely hit. When it is,
        # external video search by script_name is more reliable than scraping
        # an arbitrary redirect target — we don't know the page structure,
        # and the in-thread context (which would justify trusting an
        # untitled URL) doesn't exist for non-forum links.
        if script_name and not _no_resolve:
            ext = self._external_video_search(script_name, author=author)
            if ext:
                return ext

        logger.debug(f"No demo found for non-Discourse URL: {discussion_url}")
        return ""

    # ---------------------------
    # Discourse JSON API demo discovery (Discourse-only variant)
    # ---------------------------
    # Demo URL patterns split by media type. Discovery prioritizes VIDEO over
    # AUDIO (script demos are more compelling as video). The full priority
    # order is implemented in _discover_demo_via_discourse_api as a 4-phase
    # search: early-post video -> external search -> rest-of-thread video ->
    # audio fallback anywhere.
    _VIDEO_LINK_PATTERNS = (
        "youtube.com/watch",
        "youtu.be/",
        "m.youtube.com/watch",
        "music.youtube.com/watch",
        "vimeo.com/",
        "instagram.com/",
    )
    _VIDEO_IFRAME_PATTERNS = (
        "youtube.com/embed/",
        "vimeo.com/video/",
        "player.vimeo.com",
    )
    _AUDIO_LINK_PATTERNS = (
        "soundcloud.com/",
        # Direct audio-file links (Discourse uploads, zbs.fm, etc.) — norns
        # authors sometimes attach raw .mp3 demos to their threads.
        ".mp3",
        ".wav",
        ".m4a",
        ".ogg",
        ".flac",
    )
    _AUDIO_IFRAME_PATTERNS = (
        "soundcloud.com/player",
        "w.soundcloud.com",
    )

    # Backward-compat aliases (used by the legacy HTML scrape path in
    # discover_demo_video for non-Discourse URLs).
    _DEMO_LINK_PATTERNS = _VIDEO_LINK_PATTERNS + _AUDIO_LINK_PATTERNS
    _DEMO_IFRAME_PATTERNS = _VIDEO_IFRAME_PATTERNS + _AUDIO_IFRAME_PATTERNS

    # Max number of unique demo URLs to collect per script. Multiple demos in
    # a single Excel cell are stored newline-separated; the cell hyperlink
    # points to the first URL (Excel doesn't natively support multi-hyperlink
    # cells), but all are visible/copyable in the cell text.
    _DEMO_LIMIT = 3

    # Signature for the external-search cache file. Bump this whenever
    # `_video_matches_script` semantics change (or any other accept/reject
    # rule for external search results) so the on-disk cache rebuilds itself
    # rather than replaying stale false-positives through the new rules.
    # Append a short reason to the version string for greppable history.
    _MATCHER_SIGNATURE = "v1-strict-norns-required"

    # --- Demo URL normalization + oembed liveness (ported from the 2026-06 manual audit) ---
    # oembed endpoints reliably report whether media still exists, unlike a bare
    # HEAD/GET which returns a 200 "soft shell" for removed YouTube/Vimeo/SoundCloud
    # media. This is the root-cause fix for dead demo links accumulating in the xlsx.
    _OEMBED = {
        "youtube": "https://www.youtube.com/oembed",
        "vimeo": "https://vimeo.com/api/oembed.json",
        "soundcloud": "https://soundcloud.com/oembed",
    }
    # Long-lived public Vimeo videos used to probe whether oembed works here
    # (usable if ANY returns 200). monome's Vimeos are embed-domain-whitelisted,
    # so some networks/CI runners get 404 from oembed for every video — in that
    # case oembed can't tell dead from restricted and must NOT be used to drop.
    _VIMEO_OEMBED_CANARIES = ("22439234", "1084537")  # "The Mountain", "Big Buck Bunny"

    # Extracts a numeric track id from SoundCloud embed-artifact URL forms:
    #   soundcloud.com/track/<id>            (not a real clickable path)
    #   api.soundcloud.com/tracks/<id>       (API endpoint, needs auth in browser)
    #   w.soundcloud.com/player/?url=...tracks/<id>  (iframe embed)
    # Clean permalinks (soundcloud.com/<user>/<slug>) do NOT match.
    _SC_TRACK_ID_RE = re.compile(
        r"(?:api\.soundcloud\.com/tracks/|soundcloud\.com/track/|tracks(?:%2F|/))(\d+)"
    )

    # Preference order when a single post yields multiple demo candidates. Lower =
    # preferred: resilient video hosts first, then social/image, then audio, then
    # fragile/ephemeral hosts (Twitch auto-deletes VODs for non-partners) last.
    # This is a *within-post* tiebreak only — the 4-phase discovery still decides
    # official-vs-external-vs-audio ordering across posts.
    _DEMO_PLATFORM_RANK = (
        ("youtube.com", 0), ("youtu.be", 0),
        ("vimeo.com", 1),
        ("instagram.com", 2),
        ("soundcloud.com", 3), ("bandcamp.com", 3),
        (".mp3", 4), (".wav", 4), (".m4a", 4), (".ogg", 4), (".flac", 4),
        ("twitch.tv", 6),
    )

    @classmethod
    def _demo_rank(cls, url):
        u = (url or "").lower()
        for token, rank in cls._DEMO_PLATFORM_RANK:
            if token in u:
                return rank
        return 5  # unknown hosts sit just above fragile/ephemeral (twitch)

    # Known-dead demo URLs the scraper must NEVER store again. These are removed
    # links the discovery path would otherwise keep re-grabbing from stale thread
    # embeds. Stored as normalized keys (see _denylist_key). This is the belt to
    # oembed-validation's suspenders — and the ONLY guard in environments where
    # Vimeo oembed is unavailable (CI whitelist), since a dead Vimeo there passes
    # both HEAD and the skipped oembed check. Populated by the 2026-06 audit.
    _DEMO_URL_DENYLIST = frozenset({
        # Vimeo videos confirmed 404 (removed/private):
        "vimeo.com/312196152",   # ash
        "vimeo.com/327848801",   # awake
        "vimeo.com/913013027",   # cc-canvas
        "vimeo.com/480411843",   # cheat_codes_2 (old)
        "vimeo.com/146731772",   # meadowphysics (old)
        "vimeo.com/266741634",   # mlr (old)
        "vimeo.com/416730766",   # nc02-rs
        "vimeo.com/484176216",   # norns.online
        # SoundCloud confirmed gone / malformed embed-artifact paths:
        "soundcloud.com/track/675154727",                  # compass (not a real path)
        "soundcloud.com/yobink/021920003-modular-animator",  # animator (404)
        "soundcloud.com/sound-and-process/scarlet",          # nc03-ds (old, 404)
    })

    @staticmethod
    def _denylist_key(url):
        """Normalize a URL for denylist comparison: drop scheme, www., query,
        fragment, and trailing slash; lowercase."""
        u = (url or "").strip().lower()
        u = re.sub(r"^https?://", "", u)
        u = u.split("?", 1)[0].split("#", 1)[0]
        if u.startswith("www."):
            u = u[4:]
        return u.rstrip("/")

    @classmethod
    def _is_denylisted_demo(cls, url):
        return cls._denylist_key(url) in cls._DEMO_URL_DENYLIST

    @staticmethod
    def _write_demo_cell(ws, row_idx, value):
        """Write the Demo cell at column 5, handling both single and multi-URL
        (newline-separated) values. Hyperlinks the first URL; wraps text so all
        URLs in a multi-cell are visible without clipping.
        """
        try:
            v = "" if pd.isna(value) else str(value)
        except Exception:
            v = "" if value is None else str(value)
        if v == "nan":
            v = ""
        v = v.strip()
        cell = ws.cell(row=row_idx, column=5, value=v)
        if not v:
            cell.font = Font(size=14)
            return
        # Take the first non-empty line as the click target.
        first_url = next((line.strip() for line in v.splitlines() if line.strip()), "")
        if first_url:
            cell.hyperlink = first_url
            cell.font = Font(size=14, color="0000FF", underline="single")
        else:
            cell.font = Font(size=14)
        # Wrap text so multi-line cells display all URLs.
        if "\n" in v:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _throttle_discourse(self):
        """Block until enough time has elapsed since the last Discourse request.

        Called by all threads before any Discourse fetch. Single shared timestamp
        means all workers cooperatively pace themselves below ``_discourse_rps``.
        """
        with self._discourse_throttle_lock:
            now = time.monotonic()
            since = now - self._discourse_last_request
            if since < self._discourse_min_interval:
                time.sleep(self._discourse_min_interval - since)
            self._discourse_last_request = time.monotonic()

    # ---------------------------
    # External-cache persistence + ban-list + URL validation
    # ---------------------------
    def _external_search_path(self, excel_path):
        base, _ = os.path.splitext(excel_path)
        return f"{base}.external_searches.json"

    def _load_external_searches(self, excel_path):
        """Load persisted external-search results so daily CI doesn't re-search.

        On-disk format (current):
            {"signature": "<_MATCHER_SIGNATURE>", "entries": {key: url-or-empty}}

        On-disk format (legacy, pre-signature): {key: url-or-empty}. Treated
        as stale and discarded — the matcher rules that produced those entries
        are unknown, so replaying them through the current matcher would
        defeat the point of bumping the signature in the first place.

        Cache key: "<author>|<script_name>". Empty values are intentional
        cached negatives (we tried, found nothing). To force a re-search on a
        row, delete its entry from the sidecar file.
        """
        import json

        path = self._external_search_path(excel_path)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            logger.warning(f"Failed to load external search cache from {path}: {e}")
            return {}

        # Current format: {"signature": "...", "entries": {...}}
        if (
            isinstance(data, dict)
            and "signature" in data
            and "entries" in data
            and isinstance(data["entries"], dict)
        ):
            on_disk_sig = data.get("signature")
            if on_disk_sig != self._MATCHER_SIGNATURE:
                logger.info(
                    f"Matcher signature changed ({on_disk_sig!r} → "
                    f"{self._MATCHER_SIGNATURE!r}); rebuilding external "
                    f"search cache (discarded {len(data['entries'])} entries)"
                )
                return {}
            entries = data["entries"]
            logger.info(f"Loaded {len(entries)} cached external searches from {path}")

            # One-shot legacy migration: bare-string entries (pre-smart-recheck
            # schema) get promoted to dict shape so they participate in the
            # smart-recheck logic instead of being treated as cache-misses.
            # We trust the URL as written (the matcher rules haven't changed
            # since the legacy entry's signature, which we just validated).
            # Stamp `cached_at` with `now` to start the 6-month TTL clock from
            # migration time; leave upstream signals null (we don't know what
            # they were at original cache time, so any future upstream movement
            # will trigger a recheck via the normal lookup logic).
            entries, migrated_count = self._migrate_legacy_entries(entries)
            if migrated_count:
                logger.info(
                    f"Migrated {migrated_count} legacy bare-string cache "
                    f"entries to dict schema (will persist on save)"
                )

            cleaned = self._drop_dead_cache_urls(entries)
            if migrated_count or len(cleaned) < len(entries):
                # Persist migration and/or dead-link cleanup so subsequent
                # runs don't redo this work. Cached negatives (empty url
                # values) are preserved by _drop_dead_cache_urls.
                self._save_external_searches(excel_path, cleaned)
            return cleaned

        # Legacy format: bare {key: url} dict, no signature. Discard.
        if isinstance(data, dict):
            logger.info(
                f"External search cache at {path} is in legacy (unsigned) "
                f"format; rebuilding (discarded {len(data)} entries)"
            )
            return {}

        logger.warning(
            f"External search cache at {path} has unexpected shape; ignoring."
        )
        return {}

    def _save_external_searches(self, excel_path, cache):
        import json

        if not cache:
            return
        path = self._external_search_path(excel_path)
        payload = {
            "signature": self._MATCHER_SIGNATURE,
            "entries": cache,
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            logger.info(
                f"Saved {len(cache)} external searches to {path} "
                f"(sig={self._MATCHER_SIGNATURE})"
            )
        except Exception as e:
            logger.warning(f"Failed to save external search cache to {path}: {e}")

    # Hosts where HEAD validation is unreliable (we share an IP with a service
    # that rate-limits us, or HEAD/GET semantics for the host are funky enough
    # that a failed probe doesn't mean the URL is dead). For URLs on these
    # hosts we keep them without HEAD-checking on the assumption that they
    # reached us via Discourse in-thread context, which is itself validation.
    _HEAD_VALIDATION_SKIP_HOSTS = ("llllllll.co",)

    def _should_skip_head_validation(self, url):
        if not url:
            return False
        u = url.lower()
        for h in self._HEAD_VALIDATION_SKIP_HOSTS:
            # Match host segment, not arbitrary substring (avoid false matches
            # like "evil.com/llllllll.co/anything").
            if f"://{h}/" in u or u.startswith(f"{h}/") or f"://{h}" == u.rstrip("/"):
                return True
        return False

    # ---------------------------
    # Demo URL normalization (SoundCloud track-IDs -> clean permalinks)
    # ---------------------------
    def _sc_track_id(self, url):
        m = self._SC_TRACK_ID_RE.search(url or "")
        return m.group(1) if m else None

    def _normalize_soundcloud_url(self, url, timeout=8):
        """Resolve an embed-artifact SoundCloud URL to a clean, clickable permalink.

        `soundcloud.com/track/<id>` is not a real path; `api.soundcloud.com/tracks/<id>`
        and the `w.soundcloud.com/player/?url=...tracks/<id>` iframe form aren't
        click-through links either. The public oembed endpoint accepts the API track
        URL and returns author_url + title, from which we rebuild the permalink
        (SoundCloud truncates long slugs, so we trim trailing words until it 200s).
        Returns the original URL unchanged on any failure.
        """
        tid = self._sc_track_id(url)
        if not tid:
            return url
        try:
            r = self.session.get(
                self._OEMBED["soundcloud"],
                params={"url": f"https://api.soundcloud.com/tracks/{tid}", "format": "json"},
                timeout=timeout,
            )
            if r.status_code != 200:
                return url
            j = r.json()
            author_url = (j.get("author_url") or "").rstrip("/")
            title = j.get("title") or ""
            if not author_url or not title:
                return url
            slug = re.sub(r"[^a-z0-9]+", "-", title.split(" by ")[0].lower()).strip("-")
            words = [w for w in slug.split("-") if w]
            # Try the full slug first, then progressively shorter (SoundCloud
            # truncates long permalink slugs). Bounded by word count.
            for n in range(len(words), 0, -1):
                cand = f"{author_url}/{'-'.join(words[:n])}"
                vr = self.session.get(
                    self._OEMBED["soundcloud"],
                    params={"url": cand, "format": "json"},
                    timeout=timeout,
                )
                if vr.status_code == 200:
                    logger.debug(f"Normalized SoundCloud {url[:50]} -> {cand}")
                    return cand
        except Exception as e:
            logger.debug(f"SoundCloud normalize failed for {url[:60]}: {e}")
        return url

    def _normalize_demo_url(self, url):
        """Dispatch demo-URL normalization. Currently only SoundCloud needs it;
        no-op (and no network call) for everything else."""
        if not url:
            return url
        if self._sc_track_id(url):
            return self._normalize_soundcloud_url(url)
        return url

    # ---------------------------
    # oembed liveness (catches removed media that HEAD passes as a 200 soft-shell)
    # ---------------------------
    @staticmethod
    def _oembed_host(url):
        u = (url or "").lower()
        if "youtube.com" in u or "youtu.be" in u:
            return "youtube"
        if "vimeo.com" in u:
            return "vimeo"
        if "soundcloud.com" in u:
            return "soundcloud"
        return None

    def _vimeo_oembed_usable(self, timeout=8):
        """Probe once whether Vimeo oembed is meaningful from this environment.

        If every Vimeo oembed request 404s here (embed-domain whitelist on some
        networks/CI), oembed can't distinguish a dead video from a restricted one,
        so we must not use it to drop Vimeo links. Fails safe to False (= don't use
        oembed for Vimeo, fall through to lenient HEAD/keep).
        """
        if self._vimeo_oembed_ok is None:
            ok = False
            for vid in self._VIMEO_OEMBED_CANARIES:
                try:
                    r = self.session.get(
                        self._OEMBED["vimeo"],
                        params={"url": f"https://vimeo.com/{vid}"},
                        timeout=timeout,
                    )
                    if r.status_code == 200:
                        ok = True
                        break
                except Exception:
                    continue
            self._vimeo_oembed_ok = ok
            if not ok:
                logger.info(
                    "Vimeo oembed unavailable in this environment; "
                    "skipping oembed liveness for Vimeo (HEAD fallback)."
                )
        return self._vimeo_oembed_ok

    def _oembed_alive(self, url, timeout=8):
        """Return True (alive) / False (dead) / None (inconclusive) via oembed.

        Catches removed/private YouTube/Vimeo/SoundCloud media that a bare HEAD
        would wrongly pass (those hosts serve a 200 'soft shell' for dead media).
        None means "can't tell from oembed" — caller should fall back to HEAD.
        """
        host = self._oembed_host(url)
        if not host:
            return None
        if host == "vimeo" and not self._vimeo_oembed_usable():
            return None
        try:
            params = {"url": url}
            if host != "vimeo":
                params["format"] = "json"
            r = self.session.get(self._OEMBED[host], params=params, timeout=timeout)
            if r.status_code == 200:
                return True
            if r.status_code in (401, 403, 404):
                return False
            return None
        except Exception:
            return None

    def _validate_demo_url(self, url, timeout=8):
        """Check that a demo URL still points at live media.

        For video/audio platforms (YouTube/Vimeo/SoundCloud) a bare HEAD/GET
        returns a 200 "soft shell" even for removed videos, so HEAD alone lets
        dead links through — the cause of the dead-link backlog cleaned up in the
        2026-06 audit. We oembed-check those hosts first (definitive True/False),
        and fall back to HEAD (with GET-on-405) for everything else or when oembed
        is inconclusive. Failure is a soft-drop: callers keep going on error.
        """
        if not url:
            return False
        if self._is_denylisted_demo(url):
            logger.info(f"Demo URL on denylist, refusing to store: {url[:80]}")
            return False
        verdict = self._oembed_alive(url, timeout=timeout)
        if verdict is True:
            return True
        if verdict is False:
            return False
        try:
            r = self.session.head(url, timeout=timeout, allow_redirects=True)
            if r.status_code == 405:
                # Some hosts (e.g. soundcloud) reject HEAD; fall back to GET stream
                r = self.session.get(url, timeout=timeout, stream=True)
                r.close()
            return 200 <= r.status_code < 400
        except Exception as e:
            logger.debug(f"URL validation failed for {url[:80]}: {e}")
            return False

    def _finalize_demos(self, collected):
        """Final pre-return pass: validate URLs are reachable, then join.

        Used at every return point of the discovery API path so dead links
        never reach the xlsx. Empty input returns empty string.
        """
        if not collected:
            return ""
        live = self._validate_demo_urls(collected)
        return "\n".join(live)

    def _validate_demo_urls(self, urls):
        """Filter URLs by reachability with host-aware policy.

        URLs on `_HEAD_VALIDATION_SKIP_HOSTS` (currently just llllllll.co)
        are passed through without HEAD-checking. They reached us via
        Discourse in-thread context — the post itself is validation —
        and the same host rate-limits our IP, so HEAD failures are
        diagnostically ambiguous rather than evidence of a dead link.
        Concrete case the old behavior dropped: thebangs's
        /uploads/.../e31b2c25 attachment, which appeared in-thread but
        got HEAD-failed during a 429 storm and was wrongly removed.

        For all other hosts, run the parallel HEAD-check as before.
        Drop URLs that fail; preserve original order.
        """
        if not urls:
            return urls
        skip = [u for u in urls if self._should_skip_head_validation(u)]
        check = [u for u in urls if not self._should_skip_head_validation(u)]

        if not check:
            return list(urls)

        # Cap workers low; these requests aren't throttled but we don't need to hammer.
        workers = min(8, len(check))
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = {ex.submit(self._validate_demo_url, u): u for u in check}
            ok = set()
            for f in as_completed(futures):
                u = futures[f]
                try:
                    if f.result():
                        ok.add(u)
                except Exception:
                    pass
        ok.update(skip)  # skipped URLs are kept by policy
        kept = [u for u in urls if u in ok]
        if len(kept) < len(urls):
            dropped = [u for u in urls if u not in ok]
            logger.info(
                f"Dropped {len(dropped)} unreachable demo URL(s): "
                + ", ".join(d[:60] for d in dropped[:3])
            )
        if skip:
            logger.debug(
                f"Skipped HEAD validation for {len(skip)} URL(s) on trusted-context hosts"
            )
        return kept

    @staticmethod
    def _entry_url(entry):
        """Extract the URL from a cache entry, tolerating both new dict schema
        and legacy bare-string entries.

        Legacy entries (bare strings from pre-smart-recheck cache files) are
        treated as "URL only, no timestamps." Lookup logic invalidates them
        anyway, so the URL is read here only for HEAD-validation purposes.
        """
        if isinstance(entry, str):
            return entry
        if isinstance(entry, dict):
            return entry.get("url", "") or ""
        return ""

    def _drop_dead_cache_urls(self, entries):
        """Filter cached external-search entries to drop dead URLs.

        Cached negatives (empty values) are preserved — those represent
        "we searched and found nothing", which is still valid. Only
        non-empty URLs that fail HEAD validation are dropped.

        Tolerates both new dict-shape entries and legacy bare-string entries.
        Returns a new dict; does not mutate the input.
        """
        if not entries:
            return entries
        urls = [self._entry_url(v) for v in entries.values()]
        urls = [u for u in urls if u]
        if not urls:
            return dict(entries)
        live = set(self._validate_demo_urls(urls))
        cleaned = {
            k: v
            for k, v in entries.items()
            if (not self._entry_url(v)) or (self._entry_url(v) in live)
        }
        dropped = len(entries) - len(cleaned)
        if dropped:
            logger.info(
                f"Dropped {dropped} dead URL(s) from external search cache "
                f"({len(cleaned)} entries remaining)"
            )
        return cleaned

    # ---------------------------
    # Smart cache invalidation: per-entry timestamps + upstream signals
    # ---------------------------
    # The cache file stores `{signature, entries: {key: <entry>}}`. An entry
    # is either:
    #   - a bare string (legacy, pre-smart-recheck) — invalidated on lookup
    #   - a dict with shape:
    #       {
    #         "url": str,                       # may be "" for cached negative
    #         "repo_updated_at": str | None,    # GitHub Last Updated when cached (YYYY-MM-DD)
    #         "thread_last_post_at": str | None,# Discourse last_posted_at when cached (ISO 8601)
    #         "cached_at": str,                 # write timestamp (ISO 8601 UTC)
    #         "transient_failure": bool,        # True if API was unavailable; always recheck
    #       }
    #
    # Invalidation rules (any one triggers a recheck):
    #   1. Entry is legacy (bare string) — drop, force recheck.
    #   2. transient_failure=True — caching was a fallback; we never got a real signal.
    #   3. cached_at older than _CACHE_TTL_DAYS — long-tail safety net for
    #      cases where neither repo nor thread moved but a demo got uploaded
    #      to YouTube anyway (rare but real).
    #   4. current(repo_updated_at) > entry.repo_updated_at — author committed.
    #   5. current(thread_last_post_at) > entry.thread_last_post_at — community posted.
    #
    # Missing signals on either side default to "no invalidation from this signal."
    # That keeps transient upstream flakes (GitHub/Discourse 429) from forcing
    # spurious cache invalidation.

    _CACHE_TTL_DAYS = 180  # 6 months

    @staticmethod
    def _now_iso():
        from datetime import datetime, timezone
        return datetime.now(timezone.utc).isoformat()

    def _make_cache_entry(self, url, repo_updated_at, thread_last_post_at, transient=False):
        return {
            "url": url or "",
            "repo_updated_at": (repo_updated_at or None),
            "thread_last_post_at": (thread_last_post_at or None),
            "cached_at": self._now_iso(),
            "transient_failure": bool(transient),
        }

    def _migrate_legacy_entries(self, entries):
        """One-shot upgrade: promote bare-string cache entries to dict shape.

        Bare strings come from cache files written before smart-recheck. We
        trust them on two grounds:
          - The matcher signature was validated by the caller before this
            runs, so the verdict (positive or negative) was produced by the
            same matcher rules in effect today.
          - Forcing a recheck just because the schema changed would burn
            quota replicating verdicts we already have.

        Stamping `cached_at` with `now` starts the 6-month TTL clock from
        the migration moment, so legacy entries get re-checked at the same
        cadence as freshly-written ones. Upstream signals are left as None
        — we don't know what they were at original cache time, so the only
        invalidation triggers for migrated entries are TTL or future
        upstream movement (which still triggers recheck via the normal
        lookup-time comparison once a real signal exists to compare).

        Returns (migrated_dict, count_migrated). Does not mutate the input.
        """
        if not entries:
            return entries, 0
        out = {}
        migrated = 0
        for k, v in entries.items():
            if isinstance(v, str):
                out[k] = {
                    "url": v,
                    "repo_updated_at": None,
                    "thread_last_post_at": None,
                    "cached_at": self._now_iso(),
                    "transient_failure": False,
                }
                migrated += 1
            else:
                out[k] = v
        return out, migrated

    def _cache_lookup(self, key, current_repo_updated, current_thread_last_post):
        """Return cached URL string if entry is fresh; otherwise None.

        None means "treat as cache miss — re-run external search."
        Empty string is a valid return: a cached negative we still trust.
        """
        entry = self._external_search_cache.get(key)
        if entry is None:
            return None
        if not isinstance(entry, dict):
            # Legacy bare-string entry — schema migration. Force recheck once
            # so the entry recaches in the new dict shape.
            return None
        if entry.get("transient_failure"):
            return None
        # 6-month TTL safety net for cases neither upstream signal catches.
        cached_at = entry.get("cached_at") or ""
        if cached_at:
            try:
                from datetime import datetime, timezone
                age = (datetime.now(timezone.utc) - datetime.fromisoformat(cached_at)).days
                if age > self._CACHE_TTL_DAYS:
                    return None
            except (ValueError, TypeError):
                # Malformed timestamp — be conservative, force recheck.
                return None
        # Compare upstream signals where both sides have a value.
        cached_repo = entry.get("repo_updated_at") or ""
        if current_repo_updated and cached_repo and current_repo_updated > cached_repo:
            return None
        cached_thread = entry.get("thread_last_post_at") or ""
        if (
            current_thread_last_post
            and cached_thread
            and current_thread_last_post > cached_thread
        ):
            return None
        return entry.get("url", "") or ""

    def _load_vimeo_api_token(self):
        """Load Vimeo API token from env var VIMEO_API_TOKEN or vimeo.api file.

        Mirrors gh.api / yt.api pattern. Get one at https://developer.vimeo.com.
        Empty return = no token, callers skip Vimeo (it's anonymous-search-broken).
        """
        token = os.getenv("VIMEO_API_TOKEN", "").strip()
        if token:
            return token
        for path in ("vimeo.api", os.path.join(os.path.dirname(__file__), "vimeo.api")):
            t = self._read_token_file(path)
            if t:
                return t
        return ""

    def _vimeo_api_search(self, query, max_results=10):
        """Search Vimeo via the API.

        Returns (results, available) — same contract as _youtube_api_search.
        See that method's docstring for why caller cares about availability.
        """
        if not hasattr(self, "_vimeo_api_token"):
            self._vimeo_api_token = self._load_vimeo_api_token()
        if not self._vimeo_api_token:
            return [], False
        try:
            r = self.session.get(
                "https://api.vimeo.com/videos",
                params={"query": query, "per_page": max_results},
                headers={
                    "Authorization": f"Bearer {self._vimeo_api_token}",
                    "Accept": "application/vnd.vimeo.*+json;version=3.4",
                },
                timeout=15,
            )
            if r.status_code != 200:
                logger.debug(
                    f"Vimeo API returned {r.status_code} for {query!r}: {r.text[:200]}"
                )
                return [], False
            data = r.json()
            results = []
            for item in data.get("data") or []:
                # link looks like https://vimeo.com/<id> (sometimes with /<hash>)
                link = item.get("link", "")
                m = re.search(r"vimeo\.com/(\d+)", link)
                if not m:
                    continue
                vid = m.group(1)
                title = item.get("name", "") or ""
                desc = item.get("description", "") or ""
                results.append((vid, title, desc))
            return results, True
        except Exception as e:
            logger.debug(f"Vimeo API request failed for {query!r}: {e}")
            return [], False

    def _load_youtube_api_key(self):
        """Load YouTube Data API key from env var YOUTUBE_API_KEY or local yt.api file.

        Mirrors the gh.api / GITHUB_TOKEN pattern. Empty return = no key,
        which means HTML scraping fallback is used.
        """
        key = os.getenv("YOUTUBE_API_KEY", "").strip()
        if key:
            return key
        for path in ("yt.api", os.path.join(os.path.dirname(__file__), "yt.api")):
            k = self._read_token_file(path)
            if k:
                return k
        return ""

    def _youtube_api_search(self, query, max_results=10):
        """Search YouTube via the Data API.

        Returns (results, available):
          - results: list of (videoId, title, description)
          - available: True if the API responded normally (whether or not it
            returned hits), False if the API was unavailable for any reason
            (no key, HTTP error including 403 quota, network failure)

        Callers MUST check `available` before caching a negative — caching
        a "no match" verdict from an unavailable API would poison the cache
        with false negatives that survive until the matcher signature
        bumps. With smart-recheck, transient_failure entries get retried
        on the next run instead.

        API results are higher-signal than HTML scraping AND include the
        description field, which is critical for disambiguating short-name
        scripts (e.g. 'ufo' demo's title is generic but description mentions
        'Monome Norns').
        """
        if not hasattr(self, "_youtube_api_key"):
            self._youtube_api_key = self._load_youtube_api_key()
        if not self._youtube_api_key:
            return [], False
        try:
            r = self.session.get(
                "https://www.googleapis.com/youtube/v3/search",
                params={
                    "part": "snippet",
                    "q": query,
                    "type": "video",
                    "maxResults": max_results,
                    "key": self._youtube_api_key,
                },
                timeout=15,
            )
            if r.status_code != 200:
                logger.debug(
                    f"YouTube API returned {r.status_code} for {query!r}: {r.text[:200]}"
                )
                # All non-200s (403 quota, 401 bad key, 5xx) signal API
                # unavailability — caller should treat resulting empty as
                # transient, not a final negative.
                return [], False
            data = r.json()
            results = []
            for item in data.get("items") or []:
                vid = (item.get("id") or {}).get("videoId")
                snippet = item.get("snippet") or {}
                title = snippet.get("title", "")
                desc = snippet.get("description", "")
                if vid:
                    results.append((vid, title, desc))
            return results, True
        except Exception as e:
            logger.debug(f"YouTube API request failed for {query!r}: {e}")
            return [], False

    def _external_video_search(
        self,
        script_name,
        author="",
        repo_updated_at="",
        thread_last_post_at="",
    ):
        """Search YouTube → Vimeo for a video demo of this norns script.

        Layered strategy:
          1. If YouTube Data API key is configured, use it FIRST — richer
             metadata (description) lets us disambiguate short script names.
          2. Otherwise (or as fallback), HTML-scrape YouTube search.
          3. Vimeo as final fallback (no API; first-result with name length gate).

        All YouTube hits are filtered through _video_matches_script: name
        substring in title required, plus norns/monome signal in
        title-or-description for short names.

        Cache: backed by `_cache_lookup` for smart invalidation. Cached
        entries get reused only if upstream signals (repo Last Updated,
        Discourse last_posted_at) and 6-month TTL all agree they're still
        fresh. When ALL search paths return empty AND at least one was
        unavailable (API quota, network failure), the negative is recorded
        as transient_failure=True so the next run unconditionally retries.
        """
        if not script_name:
            return ""
        if not hasattr(self, "_external_search_cache"):
            self._external_search_cache = {}
        cache_key = f"{author}|{script_name}"

        cached = self._cache_lookup(cache_key, repo_updated_at, thread_last_post_at)
        if cached is not None:
            return cached

        # Build search queries. Per explicit guidance: when author is known,
        # ALWAYS include it. Never fall back to author-less search — that
        # generates false positives by ranking generic "norns SCRIPT" hits.
        # Only use script-only as a last resort when author is missing entirely.
        if author:
            queries = [f"norns {author} {script_name}"]
        else:
            queries = [f"norns {script_name}"]

        # Track whether any search path was unavailable. If we end up with
        # zero matches AND any path failed, the resulting negative is
        # transient — we don't actually know it's a true "no match."
        api_available_any = False
        api_unavailable_any = False

        # Path 1: YouTube Data API (when key available — best signal).
        if not hasattr(self, "_youtube_api_key"):
            self._youtube_api_key = self._load_youtube_api_key()
        if self._youtube_api_key:
            for q in queries:
                results, available = self._youtube_api_search(q)
                if available:
                    api_available_any = True
                else:
                    api_unavailable_any = True
                for vid, title, desc in results:
                    if self._video_matches_script(script_name, title, desc, author=author):
                        url = f"https://www.youtube.com/watch?v={vid}"
                        logger.info(
                            f"External search hit (YT API, title='{title[:50]}'): "
                            f"{q!r} -> {url}"
                        )
                        self._external_search_cache[cache_key] = self._make_cache_entry(
                            url, repo_updated_at, thread_last_post_at
                        )
                        return url

        # Path 2: YouTube HTML scrape (no API key required, but title-only
        # match — short script names without norns context will be filtered out).
        for q in queries:
            for vid, title in self._youtube_search_titled(q):
                if self._video_matches_script(script_name, title, author=author):
                    url = f"https://www.youtube.com/watch?v={vid}"
                    logger.info(
                        f"External search hit (YT scrape, title='{title[:50]}'): "
                        f"{q!r} -> {url}"
                    )
                    self._external_search_cache[cache_key] = self._make_cache_entry(
                        url, repo_updated_at, thread_last_post_at
                    )
                    return url

        # Path 3: Vimeo via API (when VIMEO_API_TOKEN / vimeo.api is configured).
        # Returns rich metadata (title + description) for accurate short-name
        # disambiguation, same as YouTube API. Falls through silently if no
        # token — anonymous Vimeo search is JS-rendered and not viable.
        if not hasattr(self, "_vimeo_api_token"):
            self._vimeo_api_token = self._load_vimeo_api_token()
        if self._vimeo_api_token:
            for q in queries:
                results, available = self._vimeo_api_search(q)
                if available:
                    api_available_any = True
                else:
                    api_unavailable_any = True
                for vid, title, desc in results:
                    if self._video_matches_script(
                        script_name, title, desc, author=author
                    ):
                        url = f"https://vimeo.com/{vid}"
                        logger.info(
                            f"External search hit (Vimeo API, title='{title[:50]}'): "
                            f"{q!r} -> {url}"
                        )
                        self._external_search_cache[cache_key] = self._make_cache_entry(
                            url, repo_updated_at, thread_last_post_at
                        )
                        return url

        # No match found. Decide whether this is a true negative (every
        # search path responded normally but found nothing) or transient
        # (at least one path was unavailable, so absence of a hit is
        # diagnostically ambiguous).
        transient = api_unavailable_any and not api_available_any
        # If both happened (e.g., YT API quota'd but Vimeo API responded),
        # we DID get a real signal from at least one source — treat as a
        # legit negative and cache normally.
        self._external_search_cache[cache_key] = self._make_cache_entry(
            "", repo_updated_at, thread_last_post_at, transient=transient
        )
        return ""

    @staticmethod
    def _video_matches_script(script_name, title, description="", author=""):
        """True if `title` (with optional description) plausibly identifies the script.

        REQUIREMENTS (all must hold):
          1. script_name (normalized to alphanumeric only) appears in title
          2. "norns" or "monome" appears somewhere in title OR description

        Both are mandatory regardless of script-name length. Long common-word
        names like "parrot" used to bypass the norns-signal check, which let
        through completely unrelated videos. We always require the norns/monome
        signal as a non-negotiable accuracy floor.

        Author is intentionally NOT used in matching: it's used in the search
        query (to bias results) but not as a title-validation signal, since
        Lines/GitHub usernames don't always match author names in YouTube titles.

        Returns False on any failure — better to leave a row blank than
        attribute a wrong video to a script.
        """
        if not script_name or not title:
            return False
        name_compact = re.sub(r"[\W_]+", "", script_name.lower())
        if len(name_compact) < 3:
            return False
        title_compact = re.sub(r"[\W_]+", "", (title or "").lower())
        desc_compact = re.sub(r"[\W_]+", "", (description or "").lower())
        # Requirement 1: script name must appear in title
        if name_compact not in title_compact:
            return False
        # Requirement 2: norns or monome must appear in title OR description
        # (no author fallback — author signals search query, not title match)
        haystack = title_compact + desc_compact
        if "norns" in haystack or "monome" in haystack:
            return True
        return False

    # Backward-compat alias.
    @classmethod
    def _title_matches_script(cls, script_name, title):
        return cls._video_matches_script(script_name, title)

    def _youtube_search_titled(self, query):
        """Search YouTube; return list of (videoId, title) tuples in result order."""
        import urllib.parse as _ul

        encoded = _ul.quote_plus(query)
        try:
            r = self.session.get(
                f"https://www.youtube.com/results?search_query={encoded}",
                timeout=20,
            )
            if r.status_code != 200:
                return []
            results = []
            seen = set()
            # YouTube embeds video metadata as inline JSON; videoId and title
            # appear in the same videoRenderer block. Non-greedy {0,500} bridges
            # the variable middle without crossing into the next entry.
            for m in re.finditer(
                r'"videoId":"([\w\-]{11})".{0,500}?"title":\{"runs":\[\{"text":"([^"]+)"',
                r.text,
            ):
                vid, title = m.group(1), m.group(2)
                if vid in seen:
                    continue
                seen.add(vid)
                # Decode JSON unicode escapes (& etc.)
                try:
                    title = bytes(title, "utf-8").decode("unicode_escape")
                except Exception:
                    pass
                results.append((vid, title))
                if len(results) >= 10:
                    break
            return results
        except Exception as e:
            logger.debug(f"YouTube search failed for {query!r}: {e}")
            return []

    def _discourse_get_with_retry(self, url, params=None, timeout=20, max_retries=10):
        """GET with global throttle + automatic backoff on HTTP 429.

        The throttle prevents bursts that trigger rate limiting in the first
        place. The retry handles whatever still leaks through. Without both,
        parallel runs of 4+ workers regularly drop demos to empty.
        """
        import random as _random

        for attempt in range(max_retries + 1):
            self._throttle_discourse()
            try:
                r = self.session.get(url, params=params, timeout=timeout)
            except Exception:
                raise
            if r.status_code != 429:
                return r
            if attempt >= max_retries:
                logger.warning(
                    f"Discourse rate limit not lifted after {max_retries} retries: {url}"
                )
                return r
            retry_after = r.headers.get("Retry-After", "")
            try:
                wait = float(retry_after) if retry_after else (2 ** attempt)
            except ValueError:
                wait = 2 ** attempt
            wait = min(wait, 30) + _random.uniform(0, 0.5)
            logger.debug(
                f"429 on {url}; sleeping {wait:.1f}s (attempt {attempt + 1}/{max_retries})"
            )
            time.sleep(wait)
        return r

    def _discover_demo_via_discourse_api(
        self,
        discussion_url,
        base,
        slug,
        topic_id_hint=None,
        post_number_hint=None,
        author="",
        script_name="",
    ):
        """Discover demos via Discourse JSON API with 4-phase priority.

        Phase 1: scan EARLY posts (anchor slice + first ~5 of thread) for VIDEO.
        Phase 2: if no video yet, EXTERNAL search (YouTube → Vimeo).
        Phase 3: scan REST of the thread for VIDEO (paginate if needed).
        Phase 4: AUDIO fallback (SoundCloud, .mp3 etc.) anywhere in thread,
                 only used to top up to the cap or as last resort.

        Author/script_name are needed for Phase 2 external search and are
        threaded through from scrape_script_details. The cap is _DEMO_LIMIT
        (default 3); results are newline-joined.

        URL validation policy:
        - URLs found inside Discourse posts (Phases 1, 3, 4) are trusted on
          context: the post is in the script's own thread, so the posting
          context itself is the validation. We do NOT title-match these URLs
          and do NOT require the poster to match script_author — any user
          contributing a demo in-thread is fine.
        - URLs found via external search (Phase 2) are NOT in-context, so
          they go through `_video_matches_script` (script_name + norns/monome
          must appear in title/description) before being accepted.
        - Both paths share a final HEAD-check in `_finalize_demos`; dead
          links never reach the xlsx. That is the only validation an
          in-thread URL gets, and it is sufficient by design.
        """
        collected = []
        seen = set()

        def absorb(urls):
            for u in urls:
                if not u or u in seen:
                    continue
                seen.add(u)
                collected.append(u)
                if len(collected) >= self._DEMO_LIMIT:
                    return True
            return False

        try:
            # ====== PHASE 1A: anchor slice video scan (for /t/.../<post#> URLs) ======
            anchor_posts = []
            if topic_id_hint and post_number_hint:
                anchor_url = f"{base}/t/{topic_id_hint}/{post_number_hint}.json"
                try:
                    ra = self._discourse_get_with_retry(anchor_url, timeout=20)
                    if ra.status_code == 200:
                        anchor_data = ra.json()
                        anchor_posts = (
                            anchor_data.get("post_stream") or {}
                        ).get("posts") or []
                        for post in anchor_posts:
                            if absorb(
                                self._extract_demo_urls_from_post_html(
                                    post.get("cooked") or "", media="video"
                                )
                            ):
                                return self._finalize_demos(collected)
                except Exception as e:
                    logger.debug(f"Anchor-slice fetch failed for {discussion_url}: {e}")

            # Fetch the canonical topic JSON for the rest of the discovery work.
            #
            # Prefer the topic-id form when available — it's the always-
            # resolvable canonical URL on Discourse. Bare-slug fetches 404
            # for short/common slugs (concrete case: `sam` 404s as
            # /t/sam.json but /t/23943.json works), and the previous
            # 404→topic_id fallback only triggered on a clean 404 — during
            # a 429 storm the slug-form returns 429, the fallback never
            # fires, and the entire row is lost. Going topic-id-first
            # eliminates a wasted round-trip on every short-slug script
            # AND avoids the 429-pathological case.
            if topic_id_hint:
                r = self._discourse_get_with_retry(
                    f"{base}/t/{topic_id_hint}.json", timeout=20
                )
            else:
                r = self._discourse_get_with_retry(
                    f"{base}/t/{slug}.json", timeout=20
                )
            if r.status_code == 404:
                if collected:
                    return self._finalize_demos(collected)
                return self._discover_demo_via_discourse_search(
                    discussion_url, base, slug
                )
            if r.status_code != 200:
                logger.warning(
                    f"Discourse API returned {r.status_code} for {discussion_url}"
                )
                return self._finalize_demos(collected)

            data = r.json()
            topic_id = data.get("id")
            if not topic_id:
                logger.warning(
                    f"Discourse API response missing topic id for {discussion_url}"
                )
                return ""

            canonical_slug = data.get("slug") or slug
            canonical_url = f"{base}/t/{canonical_slug}/{topic_id}"
            if canonical_url != discussion_url:
                with self._discourse_resolutions_lock:
                    if discussion_url not in self._discourse_resolutions:
                        self._discourse_resolutions[discussion_url] = canonical_url
                        # Skip the report entry for trivial diffs — i.e. when
                        # the only difference is appending /<topic_id> to a
                        # bare-slug URL (e.g. /t/bitebeet/ -> /t/bitebeet/61957).
                        # Both URLs resolve correctly upstream; not PR-worthy.
                        if not self._is_trivial_resolution(
                            discussion_url, canonical_url
                        ):
                            self._discourse_resolution_log.append(
                                {"stale": discussion_url, "resolved": canonical_url}
                            )

            post_stream = data.get("post_stream") or {}
            initial_posts = list(post_stream.get("posts") or [])
            all_ids = post_stream.get("stream") or []
            loaded_ids = {p.get("id") for p in initial_posts if p.get("id")}
            anchor_post_ids = {p.get("id") for p in anchor_posts if p.get("id")}
            remaining_ids = [
                pid for pid in all_ids
                if pid not in loaded_ids and pid not in anchor_post_ids
            ]
            full_thread_posts = [
                p for p in initial_posts if p.get("id") not in anchor_post_ids
            ]

            # ====== PHASE 1B: early-thread video scan (OP + first ~5 replies) ======
            EARLY_N = 5
            early_posts = full_thread_posts[:EARLY_N]
            for post in early_posts:
                if absorb(
                    self._extract_demo_urls_from_post_html(
                        post.get("cooked") or "", media="video"
                    )
                ):
                    return self._finalize_demos(collected)

            # ====== PHASE 2: external video search (only if NOTHING found yet) ======
            # Most legitimate norns demos are linked in the OP or first few replies;
            # if we got past Phase 1 with nothing, the author probably didn't link
            # a video in the thread, so YouTube/Vimeo search is our best shot.
            #
            # Pass upstream signals into the search so the cache can do
            # smart-recheck: thread last_posted_at comes from the topic JSON
            # we just loaded, repo_updated_at comes from the loaded xlsx state
            # stashed on `self._upstream_repo_updated` at scrape_all_scripts
            # start. Either signal absent (e.g. row not yet in xlsx, or
            # last_posted_at missing from response) silently degrades to
            # "no upstream signal," which means the cache only invalidates
            # via 6-month TTL or transient_failure markers.
            if not collected and script_name:
                thread_last_post = data.get("last_posted_at") or ""
                repo_updated = ""
                if hasattr(self, "_upstream_repo_updated"):
                    repo_updated = self._upstream_repo_updated.get(script_name, "") or ""
                ext = self._external_video_search(
                    script_name,
                    author=author,
                    repo_updated_at=repo_updated,
                    thread_last_post_at=thread_last_post,
                )
                if ext:
                    absorb([ext])

            # ====== PHASE 3: rest-of-thread video scan (paginated) ======
            BATCH = 50
            late_posts = list(full_thread_posts[EARLY_N:])
            # Scan late posts already in initial chunk first
            for post in late_posts:
                if absorb(
                    self._extract_demo_urls_from_post_html(
                        post.get("cooked") or "", media="video"
                    )
                ):
                    return self._finalize_demos(collected)
            # Then paginate
            while remaining_ids and len(collected) < self._DEMO_LIMIT:
                batch, remaining_ids = remaining_ids[:BATCH], remaining_ids[BATCH:]
                params = [("post_ids[]", str(pid)) for pid in batch]
                try:
                    rb = self._discourse_get_with_retry(
                        f"{base}/t/{topic_id}/posts.json",
                        params=params,
                        timeout=20,
                    )
                    if rb.status_code != 200:
                        logger.debug(
                            f"posts.json returned {rb.status_code} "
                            f"for topic {topic_id}"
                        )
                        continue
                    bdata = rb.json()
                    bposts = (bdata.get("post_stream") or {}).get("posts") or []
                    full_thread_posts.extend(bposts)
                    for post in bposts:
                        if absorb(
                            self._extract_demo_urls_from_post_html(
                                post.get("cooked") or "", media="video"
                            )
                        ):
                            return self._finalize_demos(collected)
                except Exception as e:
                    logger.debug(
                        f"Failed batch fetch for topic {topic_id}: {e}"
                    )

            # ====== PHASE 4: AUDIO fallback (anywhere in thread) ======
            # Only kicks in when video discovery didn't fill the cap. Walks all
            # posts (anchor slice → full thread) for SoundCloud / .mp3 / etc.
            for post in anchor_posts + full_thread_posts:
                if absorb(
                    self._extract_demo_urls_from_post_html(
                        post.get("cooked") or "", media="audio"
                    )
                ):
                    return self._finalize_demos(collected)

            return self._finalize_demos(collected)
        except Exception as e:
            logger.warning(
                f"Error in Discourse API discovery for {discussion_url}: {e}"
            )
            return ""

    def _discover_demo_via_discourse_search(self, discussion_url, base, slug):
        """Last-resort recovery when bare-slug 404s — search by title."""
        try:
            r = self._discourse_get_with_retry(
                f"{base}/search.json",
                params={"q": f"{slug} in:title"},
                timeout=15,
            )
            if r.status_code != 200:
                return ""
            topics = (r.json() or {}).get("topics") or []
            if not topics:
                return ""
            best = next(
                (t for t in topics if t.get("slug") == slug),
                topics[0],
            )
            topic_id = best.get("id")
            if not topic_id:
                return ""
            best_slug = best.get("slug") or slug
            # Recurse into the API path with the resolved slug.
            return self._discover_demo_via_discourse_api(
                discussion_url, base, best_slug
            )
        except Exception as e:
            logger.debug(f"Discourse search fallback failed for {slug}: {e}")
            return ""

    def _extract_demo_urls_from_post_html(
        self, html, anchors=True, iframes=True, media="any"
    ):
        """Scan a single post's `cooked` HTML and return matching demo URLs.

        media: "video" → only YouTube/Vimeo/Instagram; "audio" → only
        SoundCloud and audio-file uploads; "any" → both. The 4-phase
        discovery in _discover_demo_via_discourse_api uses media-typed
        passes: video first (early posts → external → rest of thread),
        audio only as last resort.
        """
        out = []
        if not html:
            return out
        if media == "video":
            link_pats = self._VIDEO_LINK_PATTERNS
            iframe_pats = self._VIDEO_IFRAME_PATTERNS
        elif media == "audio":
            link_pats = self._AUDIO_LINK_PATTERNS
            iframe_pats = self._AUDIO_IFRAME_PATTERNS
        else:
            link_pats = self._DEMO_LINK_PATTERNS
            iframe_pats = self._DEMO_IFRAME_PATTERNS
        try:
            from html import unescape

            soup = BeautifulSoup(html, "html.parser")

            if anchors:
                for link in soup.find_all("a", href=True):
                    href = unescape(link.get("href", ""))
                    if not href:
                        continue
                    if any(p in href.lower() for p in link_pats):
                        out.append(href)

            if iframes:
                for iframe in soup.find_all("iframe", src=True):
                    src = iframe.get("src", "")
                    if not src:
                        continue
                    if any(p in src.lower() for p in iframe_pats):
                        out.append(src)

            # Normalize embed-artifact URLs (SoundCloud track-IDs/player iframes)
            # to clean permalinks, then stable-sort by platform preference so the
            # most resilient candidate from this post is first (becomes the cell's
            # click target). Stable sort preserves in-post discovery order per rank.
            out = [self._normalize_demo_url(u) for u in out]
            out.sort(key=self._demo_rank)
            return out
        except Exception as e:
            logger.debug(f"Error parsing post HTML for demo URLs: {e}")
            return out

    def retry_failed_demo_requests(self):
        """Retry demo discovery for URLs that failed due to rate limiting"""
        if not self.failed_demo_requests:
            return

        logger.info(
            f"Retrying {len(self.failed_demo_requests)} failed demo requests..."
        )

        # Increase delay for retries to be more respectful
        original_delay = self.demo_delay
        self.demo_delay = max(2.0, original_delay * 2)  # At least 2 seconds for retries

        retry_successes = 0
        for discussion_url in self.failed_demo_requests[
            :
        ]:  # Copy list to modify during iteration
            try:
                logger.info(f"Retrying demo discovery for: {discussion_url}")
                demo_url = self.discover_demo_video(discussion_url)

                if demo_url:
                    # Find the script in our data and update it
                    for script in self.script_data:
                        if script.get("discussion_url") == discussion_url:
                            script["demo"] = demo_url
                            retry_successes += 1
                            logger.info(f"Successfully found demo on retry: {demo_url}")
                            break

                    # Remove from failed list
                    self.failed_demo_requests.remove(discussion_url)

            except Exception as e:
                logger.warning(f"Retry failed for {discussion_url}: {e}")

        # Restore original delay
        self.demo_delay = original_delay

        logger.info(
            f"Retry completed: {retry_successes} demos found, {len(self.failed_demo_requests)} still failed"
        )

    def get_main_page(self):
        """Fetch the main norns.community page to get list of scripts.

        Memoized for the run: the index page is requested by both fetch_slug_map
        and scrape_all_scripts, but it's static within a single run, so we fetch
        it once. Failures are not cached (so a later call can retry)."""
        if self._main_page_html is not None:
            return self._main_page_html
        try:
            logger.info("Fetching main page...")
            response = self.session.get(self.base_url)
            response.raise_for_status()
            self._main_page_html = response.text
            return self._main_page_html
        except requests.RequestException as e:
            logger.error(f"Error fetching main page: {e}")
            return None

    def fetch_community_json(self):
        """Fetch the canonical script catalog from monome/norns-community."""
        if self._community_json is not None:
            return self._community_json
        try:
            logger.info(f"Fetching community.json from {self.COMMUNITY_JSON_URL}")
            response = self.session.get(self.COMMUNITY_JSON_URL, timeout=30)
            response.raise_for_status()
            data = response.json()
            entries = data.get("entries") or []
            logger.info(f"community.json provided {len(entries)} entries")
            self._community_json = entries
            return entries
        except (requests.RequestException, ValueError) as e:
            logger.error(f"Failed to fetch community.json: {e}")
            self._community_json = []
            return []

    def fetch_slug_map(self):
        """Harvest {project_name -> URL slug} from norns.community's index table.

        Most entries have project_name == slug, but ~13 entries with non-ASCII
        or special characters (apostrophes, dots, exclamation marks) get
        normalized differently by the website (e.g. 'høst' -> 'hst',
        "carter's delay" -> 'cartersdelay'). Harvesting from the live page
        is more robust than reimplementing their slugifier.
        """
        if self._slug_map is not None:
            return self._slug_map
        html = self.get_main_page()
        if not html:
            logger.warning(
                "Could not fetch index page; falling back to project_name as slug"
            )
            self._slug_map = {}
            return self._slug_map
        soup = BeautifulSoup(html, "html.parser")
        table = soup.find("table", id="index-table")
        if not table:
            logger.warning(
                "Index table not found on norns.community; falling back to project_name as slug"
            )
            self._slug_map = {}
            return self._slug_map
        body = table.find("tbody") or table
        slug_map = {}
        for row in body.find_all("tr"):
            first_td = row.find("td")
            if not first_td:
                continue
            link = first_td.find("a", href=True)
            if not link:
                continue
            display_name = link.get_text().strip()
            slug = link["href"].strip("/")
            if display_name and slug:
                slug_map[display_name] = slug
        logger.info(f"Harvested {len(slug_map)} name->slug mappings from index page")
        self._slug_map = slug_map
        return slug_map

    def _read_json_field(self, entry, field):
        """Read a field from a community.json entry, falling back to known typo aliases.

        Returns the value (or None). Records any deviation found onto self.schema_deviations.
        """
        if field in entry:
            return entry[field]
        # Look for typo aliases that map to this field
        for typo, intended in self.KNOWN_JSON_TYPOS.items():
            if intended == field and typo in entry:
                return entry[typo]
        return None

    def _record_schema_deviations(self, entry):
        """Inspect a single community.json entry and append any deviations to the report."""
        slug_or_name = entry.get("project_name") or "<unnamed>"
        for key in entry.keys():
            if key in self.EXPECTED_JSON_FIELDS:
                continue
            if key in self.KNOWN_JSON_TYPOS:
                self.schema_deviations.append(
                    {
                        "entry": slug_or_name,
                        "kind": "typo",
                        "field": key,
                        "suggestion": self.KNOWN_JSON_TYPOS[key],
                    }
                )
            else:
                self.schema_deviations.append(
                    {
                        "entry": slug_or_name,
                        "kind": "unknown_field",
                        "field": key,
                        "suggestion": None,
                    }
                )

    def extract_script_links(self, html_content=None):
        """Build the script link list from community.json + the index page slug map.

        Returns a list of dicts: {"name": <slug>, "url": <community_url>, "json_entry": <dict>}
        The `name` is the URL slug (not the display name) for backward compatibility
        with downstream code that keys existing data by URL path.

        The html_content parameter is accepted for backward compatibility but is
        unused; slug harvesting is handled internally via fetch_slug_map().
        """
        del html_content  # unused; preserved for caller signature compatibility
        entries = self.fetch_community_json()
        if not entries:
            return []
        slug_map = self.fetch_slug_map()

        script_links = []
        seen_slugs = set()
        for entry in entries:
            self._record_schema_deviations(entry)
            project_name = entry.get("project_name") or ""
            if not project_name:
                continue
            # Prefer the harvested slug; fall back to project_name when the index page
            # isn't reachable or this entry isn't in the map (most match 1:1).
            slug = slug_map.get(project_name) or project_name
            if slug in seen_slugs:
                continue
            seen_slugs.add(slug)
            script_links.append(
                {
                    "name": slug,
                    "url": f"{self.base_url}/{slug}/",
                    "json_entry": entry,
                }
            )

        logger.info(f"Found {len(script_links)} script links")
        return script_links

    def scrape_script_details(
        self,
        script_url,
        script_name,
        existing_data=None,
        discover_demo=True,
        json_entry=None,
    ):
        """Build script details from the community.json entry.

        community.json is the canonical source for all metadata fields (project_name,
        author, tags, description, discussion_url, project_url, documentation_url) —
        the norns.community website is regenerated from it. So we read directly
        from JSON instead of scraping per-script HTML pages.

        The Demo field is the only thing that still needs discovery; it's harvested
        from the discussion thread on llllllll.co (unchanged from the old flow).

        If no json_entry is provided (e.g. a caller didn't supply one), we look it
        up in the cached catalog by slug. If still not found, we return None — the
        catalog has authority, and a script not in it likely no longer exists.
        """
        try:
            entry = json_entry
            if entry is None:
                entry = self._lookup_json_entry_by_slug(script_name)
            if entry is None:
                logger.warning(
                    f"No community.json entry for {script_name}; cannot build details"
                )
                return None

            tags_value = self._read_json_field(entry, "tags") or []
            if not isinstance(tags_value, list):
                tags_value = list(tags_value) if tags_value else []

            script_data = {
                "project_name": entry.get("project_name") or script_name,
                "author": entry.get("author") or "",
                "tags": list(tags_value),
                "description": entry.get("description") or "",
                "demo": "",
                "discussion_url": entry.get("discussion_url") or "",
                "project_url": entry.get("project_url") or "",
                "documentation_url": self._read_json_field(entry, "documentation_url")
                or "",
                "community_url": script_url,
            }

            # Carry an existing Demo value forward so we don't re-hit llllllll.co
            # for scripts whose demo we already know.
            existing_demo = ""
            if existing_data is not None:
                ev = existing_data.get("Demo", "")
                try:
                    if pd.notna(ev) and str(ev).strip() and str(ev) != "nan":
                        existing_demo = str(ev).strip()
                except Exception:
                    existing_demo = ""
            if existing_demo:
                script_data["demo"] = existing_demo

            if (
                discover_demo
                and script_data["discussion_url"]
                and not script_data["demo"]
            ):
                demo_url = self.discover_demo_video(
                    script_data["discussion_url"],
                    author=script_data.get("author", "") or "",
                    script_name=script_data.get("project_name", "") or script_name,
                )
                if demo_url:
                    script_data["demo"] = demo_url
                    logger.info(f"Discovered demo for {script_name}: {demo_url}")

            logger.debug(f"Final data for {script_name}: {script_data}")
            return script_data

        except Exception as e:
            logger.error(f"Unexpected error building {script_name} from JSON: {e}")
            return None

    def _lookup_json_entry_by_slug(self, slug):
        """Find the community.json entry whose URL slug matches the given slug.

        Builds and caches a reverse map (slug -> entry) the first time it's called.
        Returns None if no match.
        """
        if not hasattr(self, "_slug_to_entry") or self._slug_to_entry is None:
            entries = self.fetch_community_json()
            slug_map = self.fetch_slug_map()
            # name_to_slug -> slug_to_entry
            name_to_entry = {e.get("project_name"): e for e in entries if e.get("project_name")}
            self._slug_to_entry = {}
            for name, entry in name_to_entry.items():
                resolved_slug = slug_map.get(name) or name
                self._slug_to_entry[resolved_slug] = entry
        return self._slug_to_entry.get(slug)

    # ---------------------------
    # Field normalizers (shared by drift detection in merge_data, sync_check_only,
    # and the snapshot machinery). Lifted from inline closures so all call sites
    # use one definition.
    # ---------------------------
    @staticmethod
    def _norm_text(v) -> str:
        try:
            s = "" if pd.isna(v) else str(v)
        except Exception:
            s = str(v) if v is not None else ""
        s = s.strip().lower()
        s = re.sub(r"\s+", " ", s)
        return s

    @staticmethod
    def _norm_tags(v):
        """Normalize tags (list/tuple/set or comma-string) into a sorted unique tuple.

        Locally-curated tags in LOCAL_CURATION_TAGS (e.g. "norns only") are
        dropped here so they're invisible to drift detection — their presence in
        the xlsx never causes (or masks) an "Out of Sync" flag.
        """
        try:
            if isinstance(v, (list, tuple, set)):
                items = list(v)
            else:
                s = "" if pd.isna(v) else str(v)
                s = s.strip()
                if s.startswith("[") and s.endswith("]"):
                    s = s[1:-1]
                items = s.split(",") if s else []
            tokens = []
            for item in items:
                t = ("" if item is None else str(item)).strip().strip("'\"").lower()
                if t and t not in NornsScraper.LOCAL_CURATION_TAGS:
                    tokens.append(t)
            return tuple(sorted(set(tokens)))
        except Exception:
            return tuple()

    @staticmethod
    def _norm_authors(v):
        """Normalize authors (list or comma-string) into a sorted unique tuple."""
        try:
            if isinstance(v, (list, tuple, set)):
                items = list(v)
            else:
                s = "" if pd.isna(v) else str(v)
                items = s.split(",") if s else []
            tokens = []
            for item in items:
                t = ("" if item is None else str(item)).strip().strip("'\"").lower()
                if t:
                    tokens.append(t)
            return tuple(sorted(set(tokens)))
        except Exception:
            return tuple()

    @staticmethod
    def _norm_url(v) -> str:
        try:
            raw = "" if pd.isna(v) else str(v)
        except Exception:
            raw = str(v) if v is not None else ""
        raw = raw.strip()
        if not raw:
            return ""
        try:
            from urllib.parse import urlparse, urlunparse

            p = urlparse(raw)
            scheme = "https"
            netloc = (p.netloc or "").lower().replace("www.", "")
            path = (p.path or "").rstrip("/")
            if path.endswith(".git"):
                path = path[:-4]
            return urlunparse((scheme, netloc, path, "", "", ""))
        except Exception:
            return raw.lower().rstrip("/")

    # ---------------------------
    # Snapshot-based drift detection
    # ---------------------------
    def _snapshot_path(self, excel_path: str) -> str:
        """Sidecar JSON file alongside the xlsx, e.g. norns_scripts.snapshots.json."""
        base, _ = os.path.splitext(excel_path)
        return f"{base}.snapshots.json"

    def _load_snapshots(self, excel_path: str) -> dict:
        """Load drift snapshots from the sidecar file, or return {} if absent.

        Snapshots are keyed by URL slug -> {field_internal_key: last_known_json_value}.
        First-run absence is treated as "initialize this run" — see _compute_drift.
        """
        import json

        path = self._snapshot_path(excel_path)
        if not os.path.exists(path):
            logger.info(
                f"No snapshot file at {path} — drift will be initialized this run"
            )
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                logger.warning(f"Snapshot file {path} is not a dict; ignoring")
                return {}
            logger.info(f"Loaded {len(data)} drift snapshot entries from {path}")
            return data
        except Exception as e:
            logger.warning(f"Failed to load snapshots from {path}: {e}")
            return {}

    def _save_snapshots(self, excel_path: str, snapshots: dict) -> None:
        """Persist drift snapshots to the sidecar JSON file."""
        import json

        path = self._snapshot_path(excel_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(snapshots, f, indent=2, sort_keys=True, ensure_ascii=False)
            logger.info(f"Saved {len(snapshots)} drift snapshot entries to {path}")
        except Exception as e:
            logger.warning(f"Failed to save snapshots to {path}: {e}")

    def _slug_for_row(self, row_or_url) -> str:
        """Derive a URL slug from either a row dict (Community URL) or a raw URL string."""
        if isinstance(row_or_url, dict):
            url = row_or_url.get("Community URL", "") or row_or_url.get(
                "community_url", ""
            )
        else:
            url = row_or_url or ""
        try:
            url = str(url)
        except Exception:
            return ""
        return url.replace("https://norns.community/", "").strip("/")

    def _compute_drift(
        self,
        existing_script,
        new_script,
        post_merge_script,
        snapshot_for_slug: dict,
    ) -> str:
        """Three-way drift detection with in-place snapshot updates.

        Snapshot semantics: snapshot[field] holds the JSON value of the field
        at the last run when xlsx and JSON agreed for that field. So:

          - snapshot missing for a field -> first sighting; initialize to current
            JSON, no drift reported. (Existing in-xlsx values are not analyzed
            retroactively; this run becomes the baseline.)
          - snapshot == current JSON -> JSON hasn't moved. No drift, regardless
            of whether the user has overridden the value in xlsx.
          - snapshot != current JSON, post-merge xlsx == current JSON -> the
            user (or merge auto-fill) accepted the upstream change. Update
            snapshot to current JSON, no drift.
          - snapshot != current JSON, post-merge xlsx != current JSON -> real
            drift: upstream moved while xlsx didn't follow. Snapshot stays put
            so the drift signal persists across runs until resolved.

        Returns the comma-joined Excel column names that have drifted.
        Mutates `snapshot_for_slug` in place.
        """
        if snapshot_for_slug is None:
            return ""
        diffs = []
        for excel_col, internal_key, norm_method_name in self.TRACKED_DRIFT_FIELDS:
            norm_fn = getattr(self, norm_method_name)
            new_v = new_script.get(internal_key, "") if new_script else ""
            post_v = (
                post_merge_script.get(excel_col, "") if post_merge_script else ""
            )
            snap_v = snapshot_for_slug.get(internal_key)  # may be missing/None

            if snap_v is None:
                # First sighting — initialize snapshot, do not report drift.
                snapshot_for_slug[internal_key] = new_v
                continue

            new_norm = norm_fn(new_v)
            if norm_fn(snap_v) == new_norm:
                # JSON unchanged from snapshot. User overrides in xlsx are silent.
                continue

            # JSON has moved since snapshot.
            if norm_fn(post_v) == new_norm:
                # Post-merge xlsx tracks the new JSON value -> accepted.
                snapshot_for_slug[internal_key] = new_v
                continue

            # Real drift: JSON moved, xlsx didn't follow.
            diffs.append(excel_col)
        return ", ".join(diffs)

    # ---------------------------
    # Discourse stale-URL recovery
    # ---------------------------
    # llllllll.co is a Discourse forum. Topic IDs sometimes change (merge,
    # rebuild, etc.) while slugs stay stable. The bare /t/<slug> URL
    # 302-redirects to the canonical topic, which lets us recover from 404s
    # transparently. Resolutions cache to a sidecar JSON next to the xlsx so
    # repeat runs don't re-resolve.

    # base, slug, optional topic_id, optional post_number anchor.
    # Discourse URL convention: /t/<slug>/<topic_id>/<post_number> jumps the user
    # to a specific post within a long thread. community.json links use this for
    # scripts announced mid-thread (e.g. nb_* plugins inside the n-b-et-al thread).
    _DISCOURSE_TOPIC_RE = re.compile(
        r"^(https?://[^/]+)/t/([^/?#]+)(?:/(\d+))?(?:/(\d+))?"
    )

    def _discourse_resolutions_path(self, excel_path: str) -> str:
        base, _ = os.path.splitext(excel_path)
        return f"{base}.discourse_resolutions.json"

    def _load_discourse_resolutions(self, excel_path: str) -> dict:
        import json

        path = self._discourse_resolutions_path(excel_path)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                logger.warning(
                    f"Discourse resolutions file {path} is not a dict; ignoring"
                )
                return {}
            logger.info(
                f"Loaded {len(data)} cached Discourse URL resolution(s) from {path}"
            )
            return data
        except Exception as e:
            logger.warning(f"Failed to load Discourse resolutions from {path}: {e}")
            return {}

    def _save_discourse_resolutions(self, excel_path: str, resolutions: dict) -> None:
        import json

        path = self._discourse_resolutions_path(excel_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(
                    resolutions, f, indent=2, sort_keys=True, ensure_ascii=False
                )
            logger.info(
                f"Saved {len(resolutions)} Discourse URL resolution(s) to {path}"
            )
        except Exception as e:
            logger.warning(f"Failed to save Discourse resolutions to {path}: {e}")

    def _resolve_discourse_url(self, stale_url: str):
        """Lazily resolve a stale Discourse topic URL. Returns resolved URL or "".

        Three-step strategy in order of cheapness:
          1. Hit /t/<slug> (no ID) and follow redirects — Discourse redirects
             to the canonical topic. This handles the common ID-change case.
          2. GET /t/<slug>.json for canonical metadata if redirect didn't help.
          3. Fall back to /search.json?q=<slug> in:title for slug renames.

        Cached results (including failed resolutions, stored as "") are persisted
        in a sidecar JSON. Thread-safe via self._discourse_resolutions_lock.
        """
        if not stale_url:
            return ""

        with self._discourse_resolutions_lock:
            if stale_url in self._discourse_resolutions:
                return self._discourse_resolutions[stale_url]

        resolved = self._do_resolve_discourse_url(stale_url) or ""

        with self._discourse_resolutions_lock:
            self._discourse_resolutions[stale_url] = resolved
            if (
                resolved
                and resolved != stale_url
                and not self._is_trivial_resolution(stale_url, resolved)
            ):
                self._discourse_resolution_log.append(
                    {"stale": stale_url, "resolved": resolved}
                )
        return resolved

    @staticmethod
    def _is_trivial_resolution(stale, resolved):
        """True when stale and resolved differ only in clerical ways.

        Suppresses noise from the deviation report for diffs that aren't
        worth a PR upstream:
          - trailing /<topic_id> (Discourse accepts /t/<slug> and
            /t/<slug>/<id> as the same topic)
          - http vs https
          - leading www.
          - trailing slash / case differences
        Real slug renames or topic moves are not trivial and still report.
        """
        if not stale or not resolved:
            return False

        def core(u):
            n = re.sub(r"/\d+/?$", "", u or "").rstrip("/").lower()
            n = n.replace("https://", "").replace("http://", "")
            if n.startswith("www."):
                n = n[4:]
            return n

        return core(stale) == core(resolved)

    def _do_resolve_discourse_url(self, stale_url: str):
        """Run the three-step resolution. Returns URL string or None."""
        m = self._DISCOURSE_TOPIC_RE.match(stale_url)
        if not m:
            return None
        base, slug = m.group(1), m.group(2)

        # Step 1: bare slug, follow redirects. Cheapest path; covers most cases.
        try:
            r = self.session.get(
                f"{base}/t/{slug}",
                timeout=15,
                allow_redirects=True,
                stream=True,
            )
            try:
                if r.status_code == 200 and r.url and r.url != stale_url:
                    logger.debug(f"Discourse step 1 resolved: {stale_url} -> {r.url}")
                    return r.url
            finally:
                r.close()
        except Exception as e:
            logger.debug(f"Discourse step 1 failed for {stale_url}: {e}")

        # Step 2: structured topic JSON.
        try:
            r = self.session.get(f"{base}/t/{slug}.json", timeout=15)
            if r.status_code == 200:
                data = r.json()
                tid = data.get("id")
                tslug = data.get("slug") or slug
                if tid:
                    resolved = f"{base}/t/{tslug}/{tid}"
                    logger.debug(f"Discourse step 2 resolved: {stale_url} -> {resolved}")
                    return resolved
        except Exception as e:
            logger.debug(f"Discourse step 2 failed for {stale_url}: {e}")

        # Step 3: search by slug, prefer exact slug match.
        try:
            r = self.session.get(
                f"{base}/search.json",
                params={"q": f"{slug} in:title"},
                timeout=15,
            )
            if r.status_code == 200:
                topics = (r.json() or {}).get("topics") or []
                # Prefer exact slug match
                for t in topics:
                    if t.get("slug") == slug and t.get("id"):
                        resolved = f"{base}/t/{t['slug']}/{t['id']}"
                        logger.debug(
                            f"Discourse step 3 resolved (exact slug): "
                            f"{stale_url} -> {resolved}"
                        )
                        return resolved
                # Fall back to first result
                if topics and topics[0].get("id"):
                    t = topics[0]
                    resolved = f"{base}/t/{t.get('slug', slug)}/{t['id']}"
                    logger.debug(
                        f"Discourse step 3 resolved (best-match): "
                        f"{stale_url} -> {resolved}"
                    )
                    return resolved
        except Exception as e:
            logger.debug(f"Discourse step 3 failed for {stale_url}: {e}")

        return None

    def scrape_all_scripts(self, test_filter=None):
        """Scrape all scripts from norns.community using parallel processing.

        Discourse-only variant: no Playwright preflight, no browser dependency.
        Demo discovery uses the Discourse JSON API (see _discover_demo_via_discourse_api).
        """
        # Load cached Discourse URL resolutions (sidecar JSON). Pre-existing
        # resolutions are used immediately; new ones discovered this run are
        # added to the same dict and persisted at the end of save_to_excel.
        self._discourse_resolutions = self._load_discourse_resolutions(self.excel_path)
        # External search cache (YouTube/Vimeo API hits) — persisted across
        # daily runs so we don't re-search the same scripts every day.
        self._external_search_cache = self._load_external_searches(self.excel_path)

        # Get main page
        main_html = self.get_main_page()
        if not main_html:
            logger.error("Failed to fetch main page")
            return

        # Extract script links
        script_links = self.extract_script_links(main_html)
        if not script_links:
            logger.error("No script links found")
            return

        # Filter to single script if in test mode
        if test_filter:
            original_count = len(script_links)
            script_links = [
                link for link in script_links if link["name"] == test_filter
            ]
            if not script_links:
                logger.error(
                    f"Test script '{test_filter}' not found in {original_count} available scripts"
                )
                return
            logger.info(
                f"Test mode: Filtered to 1 script out of {original_count} available"
            )

        # Load existing data to check which scripts we already have
        existing_df = self.load_existing_data()

        # Create efficient lookup for existing scripts
        # Map both project names and URL paths to existing scripts
        existing_scripts = {}
        url_to_project_name = {}
        # Snapshot of "Last Updated" per project, for the smart-recheck cache.
        # _external_video_search reads this via self._upstream_repo_updated to
        # decide whether a cached entry is still fresh (cached value == loaded
        # xlsx value means the repo hasn't moved since we cached).
        self._upstream_repo_updated = {}

        if existing_df is not None:
            for _, row in existing_df.iterrows():
                project_name = row["Name"]
                community_url = row.get("Community URL", "")
                if "Last Updated" in existing_df.columns:
                    last_updated = row.get("Last Updated", "")
                    if pd.notna(last_updated) and str(last_updated).strip():
                        self._upstream_repo_updated[project_name] = str(last_updated).strip()

                # Store by project name
                doc_url_value = (
                    row.get("Documentation URL", "")
                    if "Documentation URL" in existing_df.columns
                    else ""
                )
                existing_scripts[project_name] = {
                    "community_url": community_url,
                    "has_author": pd.notna(row["Author"]) and row["Author"] != "",
                    "has_description": pd.notna(row["Description"])
                    and row["Description"] != "",
                    "has_tags": pd.notna(row["Tags"]) and row["Tags"] != "",
                    "has_demo": pd.notna(row["Demo"]) and row["Demo"] != "",
                    "has_discussion_url": pd.notna(row["Discussion URL"])
                    and row["Discussion URL"] != "",
                    "has_project_url": pd.notna(row["Project URL"])
                    and row["Project URL"] != "",
                    # Documentation URL is sparse upstream (~125/349 entries have it),
                    # so we don't gate skip-eligibility on it — but we do track presence
                    # so the sync flag can detect drift.
                    "has_documentation_url": pd.notna(doc_url_value)
                    and str(doc_url_value).strip() != "",
                    "playwright_status": (
                        row["Playwright Status"]
                        if "Playwright Status" in existing_df.columns
                        and pd.notna(row["Playwright Status"])
                        else ""
                    ),
                }

                # Create URL path to project name mapping
                if community_url:
                    url_path = community_url.replace(
                        "https://norns.community/", ""
                    ).strip("/")
                    url_to_project_name[url_path] = project_name

            logger.info(f"Found {len(existing_scripts)} existing scripts in Excel file")

        # Determine which scripts need scraping
        scripts_to_scrape = []
        scripts_skipped = 0

        for script_link in script_links:
            script_name = script_link["name"]  # URL path (e.g., "mxsynths")
            community_url = script_link["url"]

            # Use reverse lookup to find actual project name
            actual_project_name = url_to_project_name.get(script_name)

            if actual_project_name is None:
                # New script - needs full scraping
                scripts_to_scrape.append(script_link)
                logger.debug(f"New script '{script_name}' needs full scraping")
            else:
                existing = existing_scripts[actual_project_name]

                # Check if community URL differs (script moved/renamed)
                if existing["community_url"] != community_url:
                    scripts_to_scrape.append(script_link)
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) has different community URL - needs scraping"
                    )
                    continue

                # Check if all fields have data AND Playwright Status is not blank
                # Note: We need Playwright Status even if demo exists, to track resolution
                all_fields_complete = (
                    existing["has_author"]
                    and existing["has_description"]
                    and existing["has_tags"]
                    and existing["has_discussion_url"]
                    and existing["has_project_url"]
                )

                playwright_status_resolved = existing[
                    "playwright_status"
                ] != "" and not pd.isna(existing["playwright_status"])

                if all_fields_complete and playwright_status_resolved:
                    # Script is complete - skip entirely
                    scripts_skipped += 1
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) is complete - skipping"
                    )
                else:
                    # Script needs scraping for missing fields
                    scripts_to_scrape.append(script_link)
                    missing_fields = []
                    if not existing["has_author"]:
                        missing_fields.append("Author")
                    if not existing["has_description"]:
                        missing_fields.append("Description")
                    if not existing["has_tags"]:
                        missing_fields.append("Tags")
                    if not existing["has_discussion_url"]:
                        missing_fields.append("Discussion URL")
                    if not existing["has_project_url"]:
                        missing_fields.append("Project URL")
                    if not existing["has_demo"] and existing["playwright_status"] == "":
                        missing_fields.append("Demo")
                    logger.debug(
                        f"Script '{actual_project_name}' (URL: {script_name}) needs scraping for: {', '.join(missing_fields)}"
                    )

        # Log efficiency summary
        total_scripts = len(script_links)
        scraping_count = len(scripts_to_scrape)

        logger.info(f"Efficiency analysis:")
        logger.info(f"  Total scripts found: {total_scripts}")
        logger.info(f"  Scripts needing scraping: {scraping_count}")
        logger.info(f"  Scripts skipped (already complete): {scripts_skipped}")

        if not scripts_to_scrape:
            logger.info("No scripts need scraping - all are complete")
            return

        logger.info(
            f"Starting parallel scraping of {scraping_count} scripts with {self.max_workers} workers..."
        )

        # Use ThreadPoolExecutor for parallel scraping
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks with existing data
            future_to_script = {}
            for script_link in scripts_to_scrape:
                script_name = script_link["name"]
                existing_data = None

                # Get existing data for this script if it exists
                if existing_df is not None:
                    # Prefer matching by Community URL path (most reliable)
                    try:
                        community_url_series = existing_df.get("Community URL")
                        if community_url_series is not None:
                            # Normalize to URL path for comparison
                            url_path_series = (
                                community_url_series.fillna("")
                                .astype(str)
                                .str.replace(
                                    "https://norns.community/", "", regex=False
                                )
                                .str.strip("/")
                            )
                            existing_row = existing_df[url_path_series == script_name]
                        else:
                            existing_row = existing_df[
                                existing_df["Name"] == script_name
                            ]
                    except Exception:
                        existing_row = existing_df[existing_df["Name"] == script_name]
                    if not existing_row.empty:
                        existing_data = existing_row.iloc[0].to_dict()

                future_to_script[
                    executor.submit(
                        self.scrape_script_details,
                        script_link["url"],
                        script_name,
                        existing_data,
                        True,
                        script_link.get("json_entry"),
                    )
                ] = script_link

            # Process completed tasks
            completed = 0
            for future in as_completed(future_to_script):
                script_link = future_to_script[future]
                completed += 1

                try:
                    script_data = future.result()
                    if script_data:
                        self.script_data.append(script_data)
                        logger.info(
                            f"Completed {completed}/{scraping_count}: {script_link['name']}"
                        )
                    else:
                        logger.warning(f"Failed to scrape {script_link['name']}")
                except Exception as e:
                    logger.error(f"Error processing {script_link['name']}: {e}")

        logger.info(
            f"Parallel scraping completed. Successfully scraped {len(self.script_data)} scripts."
        )

        # Handle demo discovery for scripts that need it
        self.discover_demos_unified(existing_df)

        # Retry failed demo requests if demo discovery was enabled
        if self.failed_demo_requests:
            self.retry_failed_demo_requests()
        # (Playwright conflict resolution removed in the Discourse-only variant —
        # there is only one discovery method now, so there are no conflicts.)

    # ---------------------------
    # GitHub Last Updated helpers
    # ---------------------------
    @staticmethod
    def _read_token_file(path):
        """Read a single-token file, ignoring lines that start with '#' (comments)
        and blank lines. Returns the first non-comment, non-blank line stripped,
        or "" if the file has no real content. Lets example files carry
        instructions without breaking the loader.
        """
        try:
            if not os.path.exists(path):
                return ""
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        return line
        except Exception:
            pass
        return ""

    def _load_github_token(self):
        """Load GitHub token: GH_PAT (preferred) or GITHUB_TOKEN env, else gh.api file."""
        token = (os.getenv("GH_PAT") or os.getenv("GITHUB_TOKEN") or "").strip()
        if token:
            return token
        for path in ("gh.api", os.path.join(os.path.dirname(__file__), "gh.api")):
            t = self._read_token_file(path)
            if t:
                return t
        return ""

    def _init_github_session(self):
        """Create a requests session for GitHub API with optional auth."""
        s = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=self.max_workers,
            pool_maxsize=self.max_workers * 2,
            max_retries=3,
        )
        s.mount("http://", adapter)
        s.mount("https://", adapter)
        headers = {
            "Accept": "application/vnd.github+json",
            "User-Agent": "NornsScraper-GitHub/1.0",
        }
        if self.github_token:
            headers["Authorization"] = f"Bearer {self.github_token}"
        s.headers.update(headers)
        return s

    def _parse_github_repo(self, url: str):
        """Return (owner, repo) if url points to a GitHub repo; else (None, None)."""
        try:
            from urllib.parse import urlparse

            parsed = urlparse(str(url))
            host = (parsed.netloc or "").lower()
            if not host.endswith("github.com"):
                return None, None
            parts = [p for p in (parsed.path or "").split("/") if p]
            if len(parts) < 2:
                return None, None
            owner = parts[0]
            repo = parts[1]
            if repo.endswith(".git"):
                repo = repo[:-4]
            return owner, repo
        except Exception:
            return None, None

    @staticmethod
    def _extract_github_url(text: str):
        """First github.com/{owner}/{repo} in `text` as (owner, repo), or None.
        Ignores org/user pages (no repo segment) and normalizes a .git suffix."""
        if not text:
            return None
        m = re.search(r"github\.com/([A-Za-z0-9_.-]+)/([A-Za-z0-9_.-]+)", text)
        if not m:
            return None
        owner, repo = m.group(1), re.sub(r"\.git$", "", m.group(2))
        if not owner or not repo or repo.lower() in ("", "blob", "tree"):
            return None
        # github.com system paths aren't repos (e.g. /orgs/x, /marketplace/y) —
        # skip so the forum loop doesn't waste a repo-meta 404 on them.
        if owner.lower() in ("orgs", "marketplace", "features", "apps", "sponsors", "topics", "settings"):
            return None
        return owner.lower(), repo.lower()

    def _is_readme_only_change(self, files):
        """True if all changed files are README.md (case-insensitive basename match)."""
        try:
            import os as _os

            if not files:
                return False
            for f in files:
                name = str(f.get("filename", ""))
                base = _os.path.basename(name)
                if base.lower() != "readme.md":
                    return False
            return True
        except Exception:
            return False

    def _repo_meta(self, owner: str, repo: str) -> dict:
        """Fetch `GET /repos/{owner}/{repo}` at most once per repo per run.

        Both the Last-Updated pass and feed enrichment need `default_branch`,
        and the Last-Updated cache gates on `pushed_at` — so this single call is
        shared between them (was previously fetched 2x/repo). Returns the parsed
        repo JSON, or a sentinel dict on failure: `{"_status": 404}` for a
        missing repo, `{"_status": 403}` for a rate-limited/transient failure.
        The sentinels let callers distinguish "gone" (cache the empty result)
        from "try again later" (don't cache the miss), exactly as the inline
        calls did before."""
        if not owner or not repo:
            return {"_status": 404}
        key = (owner, repo)
        with self._repo_meta_lock:
            cached = self._repo_meta_cache.get(key)
        if cached is not None:
            return cached
        meta = {"_status": 403}  # default to transient so a miss isn't cached as "gone"
        try:
            r = self.github_session.get(
                f"https://api.github.com/repos/{owner}/{repo}", timeout=15
            )
            if r.status_code == 403:
                meta = {"_status": 403}
            elif r.status_code == 404:
                meta = {"_status": 404}
            else:
                r.raise_for_status()
                body = r.json()
                meta = body if isinstance(body, dict) else {"_status": 404}
        except Exception as e:
            logger.debug(f"GitHub repo-meta error for {owner}/{repo}: {e}")
            meta = {"_status": 403}
        with self._repo_meta_lock:
            self._repo_meta_cache[key] = meta
        return meta

    def _github_latest_non_readme_date(self, owner: str, repo: str) -> str:
        """Return YYYY-MM-DD of latest commit that isn't README.md-only; empty on failure."""
        if not owner or not repo:
            return ""
        base = f"https://api.github.com/repos/{owner}/{repo}"
        timeout = 15
        try:
            # Get default branch (shared per-run repo-meta call)
            repo_info = self._repo_meta(owner, repo)
            if repo_info.get("_status") in (403, 404):
                return ""
            default_branch = repo_info.get("default_branch") or "main"

            # Iterate commits by pages
            for page in range(1, 4):  # up to ~150 commits scanned max
                commits_resp = self.github_session.get(
                    f"{base}/commits",
                    params={"sha": default_branch, "per_page": 50, "page": page},
                    timeout=timeout,
                )
                # Handle rate limits gracefully
                if commits_resp.status_code == 403:
                    return ""
                commits_resp.raise_for_status()
                commits = commits_resp.json() or []
                if not commits:
                    break
                for c in commits:
                    sha = c.get("sha")
                    if not sha:
                        continue
                    detail_resp = self.github_session.get(
                        f"{base}/commits/{sha}", timeout=timeout
                    )
                    if detail_resp.status_code == 403:
                        return ""
                    if detail_resp.status_code == 404:
                        continue
                    detail_resp.raise_for_status()
                    detail = detail_resp.json()
                    files = detail.get("files", [])
                    if self._is_readme_only_change(files):
                        # skip README.md-only commits
                        continue
                    # Use committer date if available, else author date
                    commit_info = detail.get("commit", {})
                    committer = commit_info.get("committer") or {}
                    author = commit_info.get("author") or {}
                    date_str = committer.get("date") or author.get("date") or ""
                    if date_str:
                        try:
                            # normalize to YYYY-MM-DD
                            return str(date_str)[:10]
                        except Exception:
                            return ""
                # be gentle between pages
                time.sleep(0.2)
        except Exception as e:
            logger.debug(f"GitHub latest non-README date error for {owner}/{repo}: {e}")
            return ""
        return ""

    def _github_head_sha(self, owner: str, repo: str, branch: str) -> str:
        """Full 40-char HEAD commit SHA of the default branch; '' on any failure.
        One cheap call (per_page=1). ingenue diffs the installed SHA against this
        to detect updates without its own per-script GitHub call."""
        if not owner or not repo:
            return ""
        try:
            r = self.github_session.get(
                f"https://api.github.com/repos/{owner}/{repo}/commits",
                params={"sha": branch or "main", "per_page": 1},
                timeout=15,
            )
            if r.status_code != 200:
                return ""
            commits = r.json() or []
            return str(commits[0].get("sha") or "") if commits else ""
        except Exception as e:
            logger.debug(f"HEAD sha error for {owner}/{repo}: {e}")
            return ""

    # Recompute an unchanged repo's Last-Updated at most this stale even when
    # pushed_at is unchanged — heals any transient state and bounds drift.
    LASTUPD_CACHE_TTL_DAYS = 30

    def _lastupd_cache_path(self, excel_path: str) -> str:
        """Sidecar cache of per-repo Last-Updated, alongside the xlsx. Committed
        by CI (like the feed cache) so nightly runs stay warm."""
        base, _ = os.path.splitext(excel_path)
        return f"{base}.lastupd_cache.json"

    def _load_lastupd_cache(self, excel_path: str) -> dict:
        import json

        path = self._lastupd_cache_path(excel_path)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                logger.info(f"Loaded {len(data)} Last-Updated cache entries from {path}")
                return data
        except Exception as e:
            logger.warning(f"Failed to load Last-Updated cache from {path}: {e}")
        return {}

    def _save_lastupd_cache(self, excel_path: str, cache: dict) -> None:
        import json

        path = self._lastupd_cache_path(excel_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cache, f, indent=2, sort_keys=True, ensure_ascii=False)
            logger.info(f"Saved {len(cache)} Last-Updated cache entries to {path}")
        except Exception as e:
            logger.warning(f"Failed to save Last-Updated cache to {path}: {e}")

    def _lastupd_cache_fresh(self, entry: dict, current_pushed_at: str) -> bool:
        """Reusable only if the repo hasn't been pushed since we computed the
        date (same `pushed_at`), the entry actually holds a date, and it's within
        TTL. `pushed_at` is the precise change signal — it advances exactly when
        commits land — so an unchanged repo costs ZERO commit-paging calls on
        later nightly runs (just the one shared repo-meta call). A README-only
        push bumps pushed_at and triggers a recompute, which correctly returns
        the same (older) non-README date; never the reverse, so a stale date can
        never mask a real change."""
        if not isinstance(entry, dict) or not entry.get("last_updated"):
            return False
        cached_pa = str(entry.get("pushed_at") or "")
        if not cached_pa or cached_pa != str(current_pushed_at or ""):
            return False
        computed_at = str(entry.get("computed_at") or "")[:10]
        try:
            from datetime import datetime

            age_days = (datetime.now() - datetime.strptime(computed_at, "%Y-%m-%d")).days
            return age_days <= self.LASTUPD_CACHE_TTL_DAYS
        except Exception:
            return False

    def _apply_last_updated(self, rows, excel_path: str = None):
        """Enrich merged rows with 'Last Updated' (latest non-README commit date)
        from GitHub, keyed by 'Project URL'.

        Efficiency core: the costly part is paging a repo's commits and pulling
        per-commit details to skip README-only changes — previously done for
        EVERY repo EVERY night (~900+ uncached calls). We now cache the result
        per repo and reuse it whenever the repo's `pushed_at` is unchanged: an
        unchanged repo costs one shared `GET /repos` call and zero commit calls.
        Only repos pushed since the last run pay full price. `excel_path` enables
        the committed sidecar cache; without it (ad-hoc/test callers) we still
        share repo-meta within the run but recompute dates."""
        if not rows:
            return rows
        repo_to_indices = {}
        for idx, row in enumerate(rows):
            owner, repo = self._parse_github_repo(row.get("Project URL", ""))
            if owner and repo:
                repo_to_indices.setdefault((owner, repo), []).append(idx)
        if not repo_to_indices:
            return rows

        cache = self._load_lastupd_cache(excel_path) if excel_path else {}
        cache_lock = threading.Lock()
        results = {}
        # Per-run repo status (archived / missing-404), captured for free from the
        # shared repo-meta — feeds the catalog 'status' field (roadmap #5).
        self._repo_status_map = {}

        def _task(owner_repo):
            owner, repo = owner_repo
            meta = self._repo_meta(owner, repo)  # shared per-run call
            missing = meta.get("_status") == 404
            archived = bool(meta.get("archived"))
            self._repo_status_map[owner_repo] = (
                "missing" if missing else ("archived" if archived else "active")
            )
            pushed_at = (
                "" if meta.get("_status") in (403, 404) else (meta.get("pushed_at") or "")
            )
            ck = f"{owner}/{repo}"
            entry = cache.get(ck)
            if entry and self._lastupd_cache_fresh(entry, pushed_at):
                return owner_repo, entry["last_updated"], True
            value = self._github_latest_non_readme_date(owner, repo) or ""
            # Only persist a real date paired with a known change signal, so
            # transient failures / missing repos self-heal on the next run
            # instead of caching an empty value behind a matching pushed_at.
            # `archived` rides along so --catalog-only can read it without a fetch.
            if value and pushed_at:
                with cache_lock:
                    cache[ck] = {
                        "last_updated": value,
                        "pushed_at": pushed_at,
                        "computed_at": self._today_iso(),
                        "archived": archived,
                    }
            return owner_repo, value, False

        served_from_cache = 0
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_key = {
                executor.submit(_task, key): key for key in repo_to_indices.keys()
            }
            for future in as_completed(future_to_key):
                key = future_to_key[future]
                try:
                    _, value, hit = future.result()
                    results[key] = value
                    if hit:
                        served_from_cache += 1
                except Exception:
                    results[key] = ""

        for key, indices in repo_to_indices.items():
            value = results.get(key, "")
            for idx in indices:
                rows[idx]["Last Updated"] = value

        total = len(repo_to_indices)
        logger.info(
            f"Last-Updated: {served_from_cache}/{total} repo(s) served from cache "
            f"(pushed_at unchanged); {total - served_from_cache} recomputed"
        )
        if excel_path:
            self._save_lastupd_cache(excel_path, cache)
        return rows

    # ---------------------------
    # feed.json generation (ingenue B7 integration)
    #
    # The norns on-device app `ingenue` consumes a nightly static `feed.json`
    # so it doesn't have to compute README text, image galleries, engine names,
    # nb-voice detection, or last-updated dates live on the (slow, offline-ish)
    # device. We precompute all of that here. The consumer contract lives in
    # ingenue/web/index.html: a JSON object whose `scripts` map is keyed by
    # project_name.toLowerCase(), each value carrying independently-optional
    # {engine, nb, readme, images, tags, upd}. Every field is best-effort —
    # ingenue degrades gracefully if the feed is missing, partial, or malformed.
    # ---------------------------
    FEED_CACHE_TTL_DAYS = 30  # re-fetch unchanged repos at most this stale (heals transient misses)
    FEED_README_MAXLEN = 1200  # plaintext README prefix length shipped to the device
    FEED_MAX_IMAGES = 6  # carousel cap per script
    VOICE_CORPUS_MAX_FILES = 16  # bounded blob fetch per repo (API-call ceiling)
    # Bump when engine/nb/readme/image *processing* logic changes. Cached entries
    # store processed output, so a stamped version mismatch invalidates them and
    # forces a one-time rebuild — same idea as the external-search _MATCHER_SIGNATURE.
    FEED_LOGIC_VERSION = 3  # v3: + HEAD sha per repo (ingenue update detection)

    @staticmethod
    def _today_iso() -> str:
        from datetime import datetime

        return datetime.now().strftime("%Y-%m-%d")

    def _feed_cache_path(self, excel_path: str) -> str:
        """Sidecar cache of per-repo enrichment, alongside the xlsx."""
        base, _ = os.path.splitext(excel_path)
        return f"{base}.feed_cache.json"

    def _default_feed_output(self, excel_path: str) -> str:
        """Default feed.json location: a plain `feed.json` next to the xlsx.

        Named exactly `feed.json` (the consumer contract name) so a deploy step
        can copy it verbatim into ingenue/web/. Override with --feed-output.
        """
        return os.path.join(os.path.dirname(os.path.abspath(excel_path)), "feed.json")

    def _load_feed_cache(self, excel_path: str) -> dict:
        import json

        path = self._feed_cache_path(excel_path)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                logger.info(f"Loaded {len(data)} feed-enrichment cache entries from {path}")
                return data
        except Exception as e:
            logger.warning(f"Failed to load feed cache from {path}: {e}")
        return {}

    def _save_feed_cache(self, excel_path: str, cache: dict) -> None:
        import json

        path = self._feed_cache_path(excel_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cache, f, indent=2, sort_keys=True, ensure_ascii=False)
            logger.info(f"Saved {len(cache)} feed-enrichment cache entries to {path}")
        except Exception as e:
            logger.warning(f"Failed to save feed cache to {path}: {e}")

    def _feed_cache_fresh(self, entry: dict, current_upd: str) -> bool:
        """A cache entry is reusable only if the repo hasn't changed since we
        fetched it (same Last Updated), it wasn't a transient failure, and it
        isn't older than the TTL. This is the efficiency core: unchanged repos
        are never re-fetched from GitHub on subsequent nightly runs."""
        if not isinstance(entry, dict) or entry.get("error"):
            return False
        if entry.get("logic_version") != self.FEED_LOGIC_VERSION:
            return False
        if (entry.get("source_upd") or "") != (current_upd or ""):
            return False
        fetched_at = str(entry.get("fetched_at") or "")[:10]
        try:
            from datetime import datetime

            age_days = (datetime.now() - datetime.strptime(fetched_at, "%Y-%m-%d")).days
            return age_days <= self.FEED_CACHE_TTL_DAYS
        except Exception:
            return False

    @staticmethod
    def _raw_github_url(owner: str, repo: str, branch: str, path: str) -> str:
        from urllib.parse import quote

        return (
            f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/"
            + quote(path.lstrip("/"))
        )

    @staticmethod
    def _engine_from_paths(paths) -> str:
        """SuperCollider engine class name the repo *ships*, from `Engine_<Name>.sc`.

        This is the definitive signal for the consumer's engine-deconfliction
        modal ("registers a SuperCollider engine named X") — a shipped engine,
        not one a script merely uses via a shared lib. Repos that ship more than
        one engine report the first alphabetically (deterministic)."""
        names = []
        for p in paths:
            m = re.search(r"(?:^|/)Engine_([A-Za-z0-9]+)\.sc$", str(p))
            if m:
                names.append(m.group(1))
        return sorted(set(names))[0] if names else ""

    @staticmethod
    def _facets_from_paths(paths) -> list:
        """What a repo *is*, structurally — the same facet model ingenue applies to
        live GitHub results, computed here once so cataloged entries carry it too.
        Non-exclusive: a repo can be both a script and a mod (e.g. cyrene).
          - script  : a top-level .lua (norns runs only top-level scripts)
          - mod      : lib/mod.lua (the mod convention — our own signal, not a hand tag)
          - library  : only lib/*.lua, no top-level script and no mod (require'd by others)
          - engine   : ships a SuperCollider engine (.sc / engine/ dir)"""
        ps = [str(p) for p in paths]
        top_lua = [p for p in ps if "/" not in p and p.lower().endswith(".lua")]
        has_mod = "lib/mod.lua" in ps
        lib_lua = [p for p in ps if p.lower().endswith(".lua") and re.search(r"(?:^|/)lib/", p)]
        has_engine = any(p.lower().endswith(".sc") for p in ps) or any(
            re.search(r"(?:^|/)engine/", p) for p in ps
        )
        facets = []
        if top_lua:
            facets.append("script")
        if has_mod:
            facets.append("mod")
        if not top_lua and not has_mod and lib_lua:
            facets.append("library")
        if has_engine:
            facets.append("engine")
        return facets

    @staticmethod
    def _bundled_libs_from_paths(paths) -> set:
        """Names X of bundled support libs the repo ships under lib/<X>/ (dir
        containing at least one .lua/.sc/.sh). Mirrors ingenue analyze_dir: a
        bundled copy self-resolves its own require's, so its source is EXCLUDED
        from the voice corpus — otherwise a vendored lib/nb/lib/nb.lua would make
        the host falsely look like an nb consumer (the dreamsequence false-pos)."""
        dirs = {}
        for p in paths:
            m = re.match(r"lib/([^/]+)/(.+)$", str(p))
            if m:
                dirs.setdefault(m.group(1).lower(), []).append(m.group(2))
        return {x for x, inner in dirs.items()
                if any(f.lower().endswith((".lua", ".sc", ".sh")) for f in inner)}

    @staticmethod
    def _voice_corpus_paths(paths, bundled) -> list:
        """Files whose contents form the voice/deps corpus: top-level *.lua,
        lib/**/*.lua (minus bundled lib/<X>/ copies), and Engine_*.sc. Deterministic
        order (top-level first, then lib, then engines), capped at
        VOICE_CORPUS_MAX_FILES so a pathological repo can't explode the fetch."""
        bundled = {b.lower() for b in (bundled or set())}
        def is_bundled(p):
            m = re.match(r"lib/([^/]+)/", p)
            return bool(m and m.group(1).lower() in bundled)
        top = [p for p in paths if "/" not in p and p.lower().endswith(".lua")]
        lib = [p for p in paths if p.lower().endswith(".lua")
               and re.search(r"(?:^|/)lib/", p) and not is_bundled(p)]
        sc = [p for p in paths if re.search(r"(?:^|/)Engine_[A-Za-z0-9]+\.sc$", p) and not is_bundled(p)]
        ordered, seen = [], set()
        for group in (sorted(top), sorted(lib), sorted(sc)):
            for p in group:
                if p not in seen:
                    seen.add(p); ordered.append(p)
        return ordered[: NornsScraper.VOICE_CORPUS_MAX_FILES]

    # Voice systems whose `require "<X>/lib"` presence means the script USES that
    # system's voices. Dotted names matter (mx.samples, mx.synths). Extend as new
    # voice frameworks appear. Mirrors ingenue analyze_dir's requires extraction.
    VOICE_USE_LIBS = {"mx.samples", "mx.synths"}

    @staticmethod
    def _detect_voices(blob: str, paths, bundled, facets, repo: str) -> dict:
        """Classify a repo's relationship to the norns voice ecosystem from its
        corpus blob + tree. Returns {provides, uses, systems}. provides = voices
        OTHER scripts can load (drives the 'additional voices' umbrella tag); uses
        = voice systems this script consumes. Mirrors ingenue's analyze_dir regex
        vocabulary so precomputed signals match its live /api/deps. Pure/offline."""
        text = blob or ""
        facets = list(facets or [])
        bundled = {b.lower() for b in (bundled or set())}
        provides, uses = [], []

        # --- nb ---
        nb_pack_name = bool(re.match(r"nb[_-]", (repo or "").lower()))
        nb_pack_file = any(re.match(r"nb[_-].+\.lua$", os.path.basename(str(p)).lower())
                           or re.search(r"[_-]nb\.lua$", os.path.basename(str(p)).lower())
                           for p in paths)
        if re.search(r"nb:add_player", text) or nb_pack_name or nb_pack_file:
            provides.append("nb")
        elif re.search(r"require[\s(]+['\"]nb/|/nb/lib|nb_voice|nb:add", text) and "nb" not in bundled:
            uses.append("nb")

        # --- required voice libs (mx.samples, mx.synths, …) -> uses ---
        for lib in sorted(set(re.findall(r"require[\s(]+['\"]([A-Za-z0-9_.\-]+)/lib", text))):
            if lib.lower() in bundled:
                continue
            if lib in NornsScraper.VOICE_USE_LIBS:
                uses.append(lib)

        # --- SuperCollider engines ---
        self_engines = {m.group(1).lower() for p in paths
                        for m in [re.search(r"Engine_([A-Za-z0-9]+)\.sc$", os.path.basename(str(p)))] if m}
        if self_engines and "script" not in facets:
            provides.append("sc-engine")          # engine-only/lib/mod repo: lends its engine
        used_engines = {e.lower() for e in re.findall(r"engine\.name\s*=\s*['\"]([A-Za-z0-9]+)['\"]", text)}
        if any(e not in self_engines for e in used_engines):
            uses.append("sc-engine")

        provides = sorted(set(provides))
        uses = sorted(set(u for u in uses if u not in provides))
        return {"provides": provides, "uses": uses, "systems": sorted(set(provides) | set(uses))}

    @staticmethod
    def _has_init_params(blob: str) -> tuple:
        """(has_init, has_params) from the corpus — proof the repo is a runnable
        script (defines init / adds params) vs a bare fragment. Used by the
        installability classifier downstream."""
        text = blob or ""
        return (bool(re.search(r"function\s+init\s*\(", text)),
                bool(re.search(r"params\s*:\s*add", text)))

    @staticmethod
    def _nb_key_file(repo: str, paths) -> str:
        """The one file most likely to reference nb — mirrors ingenue's classifyRepo
        key-file pick: top-level <repo>.lua, else any top-level .lua, else lib/mod.lua,
        else the first lib/*.lua. One fetch of this catches nb integration even when the
        nb:add_param call lives in lib/parameters.lua (which README/filename heuristics miss)."""
        ps = [str(p) for p in paths]
        top_lua = [p for p in ps if "/" not in p and p.lower().endswith(".lua")]
        return (
            next((p for p in top_lua if p.lower() == f"{repo.lower()}.lua"), None)
            or (top_lua[0] if top_lua else None)
            or ("lib/mod.lua" if "lib/mod.lua" in ps else None)
            or next(
                (p for p in ps if p.lower().endswith(".lua") and re.search(r"(?:^|/)lib/", p)),
                None,
            )
        )

    def _nb_from_keyfile(self, owner: str, repo: str, branch: str, paths) -> tuple:
        """Read the key file and detect nb accurately: add_player => provides a voice
        (a voice pack), add_param / require nb/lib => uses voices (assignable in-script).
        Returns (nb: bool, role: str). Best-effort; one authenticated contents fetch."""
        key = self._nb_key_file(repo, paths)
        if not key:
            return (False, "")
        try:
            from urllib.parse import quote

            seg = "/".join(quote(s) for s in key.split("/"))
            rr = self.github_session.get(
                f"https://api.github.com/repos/{owner}/{repo}/contents/{seg}",
                headers={"Accept": "application/vnd.github.raw"},
                params={"ref": branch},
                timeout=15,
            )
            if rr.status_code == 200:
                blob = rr.text
                if re.search(r"nb:add_player", blob):
                    return (True, "provides")
                if re.search(r"nb:add_param|require[\s(]*['\"][^'\"]*nb/lib", blob):
                    return (True, "uses")
        except Exception:
            pass
        return (False, "")

    @staticmethod
    def _detect_nb(readme_md: str, paths) -> bool:
        """Best-effort detection that a script *provides* an nb (note-bridge)
        voice. Without fetching file contents (too many API calls across 350+
        repos) this is a heuristic — ingenue's live `/api/deps` remains
        authoritative. Conservative on purpose so false positives don't pollute
        the synthetic "additional voice" tag facet."""
        for p in paths:
            b = os.path.basename(str(p)).lower()
            if re.match(r"nb[_-].+\.lua$", b) or re.search(r"[_-]nb\.lua$", b):
                return True
        text = (readme_md or "").lower()
        if "note bridge" in text or "note-bridge" in text:
            return True
        if re.search(r"\bnb\b[^.\n]{0,40}(voice|player)", text) or re.search(
            r"(voice|player)[^.\n]{0,40}\bnb\b", text
        ):
            return True
        return False

    @staticmethod
    def _is_badge_url(url: str) -> bool:
        low = url.lower()
        markers = (
            "shields.io",
            "/badge",
            "badgen.net",
            "travis-ci",
            "circleci.com",
            "codecov.io",
            "coveralls.io",
            "app.netlify.com",
            "ko-fi.com",
            "buymeacoffee",
            "paypal.",
            "/workflows/",  # github actions status badges
            ".svg",  # badges are overwhelmingly svg; screenshots are raster
        )
        return any(m in low for m in markers)

    @staticmethod
    def _looks_like_image(url: str) -> bool:
        low = url.split("?", 1)[0].lower()
        if low.rsplit(".", 1)[-1] in ("png", "jpg", "jpeg", "gif", "webp"):
            return True
        # GitHub-hosted uploaded/proxied images often carry no file extension.
        host_ok = (
            "user-images.githubusercontent.com",
            "raw.githubusercontent.com",
            "camo.githubusercontent.com",
            "github.com/user-attachments",
        )
        return any(h in low for h in host_ok)

    def _resolve_readme_url(self, url: str, owner: str, repo: str, branch: str) -> str:
        url = url.strip().strip("<>")
        if url.startswith(("http://", "https://")):
            return url
        if url.startswith("//"):
            return "https:" + url
        if url.startswith(("#", "data:", "mailto:")):
            return ""
        # Relative repo path -> raw URL on the default branch.
        return self._raw_github_url(owner, repo, branch, url)

    def _extract_readme_images(self, md: str, owner: str, repo: str, branch: str) -> list:
        """Curated gallery from README markdown/HTML images, badges filtered,
        relative paths resolved to absolute raw.githubusercontent URLs."""
        raw = []
        for m in re.finditer(r"!\[[^\]]*\]\(\s*([^)\s]+)", md):
            raw.append(m.group(1))
        for m in re.finditer(r"<img[^>]+src=[\"']([^\"']+)[\"']", md, re.I):
            raw.append(m.group(1))
        out, seen = [], set()
        for u in raw:
            if not u or self._is_badge_url(u):
                continue
            full = self._resolve_readme_url(u, owner, repo, branch)
            if not full or full in seen or not self._looks_like_image(full):
                continue
            seen.add(full)
            out.append(full)
            if len(out) >= self.FEED_MAX_IMAGES:
                break
        return out

    # README demo extraction: embeddable platforms the site renders, video>audio.
    _MEDIA_VIDEO = re.compile(
        r"https?://(?:www\.)?(?:youtube\.com/(?:watch\?[^\s)]*v=|embed/|shorts/)[\w-]{6,}"
        r"|youtu\.be/[\w-]{6,}"
        r"|vimeo\.com/(?:video/)?\d+"
        r"|(?:www\.)?instagram\.com/(?:p|reel)/[\w-]+)",
        re.I,
    )
    _MEDIA_AUDIO = re.compile(
        r"https?://(?:(?:www\.)?soundcloud\.com/[\w-]+/[\w-]+"
        r"|[\w-]+\.bandcamp\.com/(?:track|album)/[\w-]+)",
        re.I,
    )

    @staticmethod
    def _extract_readme_media(md: str) -> str:
        """Best single demo URL from README text: first video link if any, else
        first audio link, else ''. Matches the site's embeddable platforms."""
        if not md:
            return ""
        mv = NornsScraper._MEDIA_VIDEO.search(md)
        if mv:
            return mv.group(0).rstrip(").,")
        ma = NornsScraper._MEDIA_AUDIO.search(md)
        if ma:
            return ma.group(0).rstrip(").,")
        return ""

    def _screenshots_from_paths(self, paths, owner: str, repo: str, branch: str) -> list:
        """Fallback gallery: raster image files committed to the repo (excluding
        .github/ assets). Used only when the README carried no images."""
        out = []
        for p in paths:
            low = str(p).lower()
            if low.startswith(".github/") or "/.github/" in low:
                continue
            if low.rsplit(".", 1)[-1] in ("png", "jpg", "jpeg", "gif", "webp"):
                out.append(self._raw_github_url(owner, repo, branch, str(p)))
            if len(out) >= self.FEED_MAX_IMAGES:
                break
        return out

    @classmethod
    def _readme_to_plaintext(cls, md: str) -> str:
        """Strip markdown/HTML to a plain-text prose prefix for the device.
        ingenue HTML-escapes this on display, so emit plain text only."""
        t = md
        t = re.sub(r"```.*?```", " ", t, flags=re.S)  # fenced code
        t = re.sub(r"`([^`]*)`", r"\1", t)  # inline code
        t = re.sub(r"<!--.*?-->", " ", t, flags=re.S)  # html comments
        t = re.sub(r"!\[[^\]]*\]\([^)]*\)", " ", t)  # images
        t = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", t)  # links -> text
        t = re.sub(r"<[^>]+>", " ", t)  # html tags
        t = re.sub(r"^\s{0,3}#{1,6}\s*", "", t, flags=re.M)  # atx headings
        t = re.sub(r"^\s{0,3}>\s?", "", t, flags=re.M)  # blockquotes
        t = re.sub(r"^\s{0,3}[-*+]\s+", "", t, flags=re.M)  # bullets
        t = re.sub(r"^\s*[-=]{3,}\s*$", "", t, flags=re.M)  # hr / setext underlines
        t = re.sub(r"[*_]{1,3}", "", t)  # emphasis markers
        t = re.sub(r"\|", " ", t)  # table pipes
        t = re.sub(r"[ \t]+", " ", t)
        t = re.sub(r"[ \t]+\n", "\n", t)  # trailing space (from stripped imgs/badges)
        t = re.sub(r"\n[ \t]+", "\n", t)  # leading space -> exposes empty lines
        t = re.sub(r"\n{2,}", "\n\n", t).strip()
        if len(t) > cls.FEED_README_MAXLEN:
            cut = t[: cls.FEED_README_MAXLEN]
            boundary = max(cut.rfind(". "), cut.rfind("\n"))
            if boundary > cls.FEED_README_MAXLEN * 0.5:
                cut = cut[: boundary + 1]
            t = cut.rstrip() + " …"
        return t

    @staticmethod
    def _tags_list(raw) -> list:
        items = raw if isinstance(raw, list) else str(raw or "").split(",")
        out, seen = [], set()
        for tag in items:
            tag = str(tag).strip()
            if tag and tag.lower() not in seen:
                seen.add(tag.lower())
                out.append(tag)
        return out

    def _github_fetch_feed_enrichment(self, owner: str, repo: str) -> dict:
        """Fetch {engine, nb, readme, images} for one repo. Best-effort: returns
        partial/empty data on any failure, and sets 'error' True on transient
        failures (rate limit / network) so the caller won't cache the miss long
        term. Three GitHub calls max: repo meta, README, recursive tree."""
        import base64

        result = {"engine": "", "nb": False, "nb_role": "", "facets": [], "readme": "", "images": [], "sha": "", "demo": ""}
        if not owner or not repo:
            return result
        base = f"https://api.github.com/repos/{owner}/{repo}"
        timeout = 15
        try:
            # default branch via the shared per-run repo-meta call (also used by
            # the Last-Updated pass, so GET /repos is fetched once per repo/run).
            meta = self._repo_meta(owner, repo)
            if meta.get("_status") == 403:
                result["error"] = True
                return result
            if meta.get("_status") == 404:
                return result
            branch = meta.get("default_branch") or "main"
            result["sha"] = self._github_head_sha(owner, repo, branch)

            # README: content -> plaintext + curated images + nb hint
            readme_md = ""
            try:
                rr = self.github_session.get(f"{base}/readme", timeout=timeout)
                if rr.status_code == 403:
                    result["error"] = True
                elif rr.status_code == 200:
                    content = rr.json().get("content") or ""
                    if content:
                        readme_md = base64.b64decode(content).decode("utf-8", "replace")
            except Exception:
                pass
            if readme_md:
                result["readme"] = self._readme_to_plaintext(readme_md)
                result["demo"] = self._extract_readme_media(readme_md)
                result["images"] = self._extract_readme_images(
                    readme_md, owner, repo, branch
                )

            # Recursive tree: engine name, nb file hint, screenshot fallback
            try:
                tr = self.github_session.get(
                    f"{base}/git/trees/{branch}",
                    params={"recursive": "1"},
                    timeout=timeout,
                )
                if tr.status_code == 403:
                    result["error"] = True
                elif tr.status_code == 200:
                    paths = [
                        str(t.get("path", ""))
                        for t in (tr.json().get("tree") or [])
                        if t.get("type") == "blob"
                    ]
                    result["engine"] = self._engine_from_paths(paths)
                    result["facets"] = self._facets_from_paths(paths)
                    # accurate nb: read the key file (catches add_param buried in lib/),
                    # falling back to the filename/README heuristic for providers whose
                    # registration lives in a lib file we don't read.
                    nb_kf, nb_role = self._nb_from_keyfile(owner, repo, branch, paths)
                    heuristic_nb = self._detect_nb(readme_md, paths)
                    result["nb"] = bool(nb_kf or heuristic_nb)
                    result["nb_role"] = nb_role or ("provides" if heuristic_nb else "")
                    if not result["images"]:
                        result["images"] = self._screenshots_from_paths(
                            paths, owner, repo, branch
                        )
            except Exception:
                pass
        except Exception as e:
            logger.debug(f"Feed enrichment error for {owner}/{repo}: {e}")
            result["error"] = True
        return result

    def _gather_feed_enrichment(self, repos: dict, cache: dict) -> dict:
        """repos: {(owner, repo): current_upd}. Returns {(owner, repo): enrichment}.
        Serves unchanged repos from cache; fetches only stale/new ones in
        parallel. Mutates `cache` in place so the caller can persist it."""
        results, stale = {}, []
        for key, upd in repos.items():
            entry = cache.get(f"{key[0]}/{key[1]}")
            if entry and self._feed_cache_fresh(entry, upd):
                results[key] = entry
            else:
                stale.append((key, upd))

        if not stale:
            logger.info(
                f"Feed: all {len(results)} GitHub repo(s) served from cache "
                f"(no enrichment fetch needed)"
            )
            return results

        logger.info(
            f"Feed: fetching enrichment for {len(stale)} changed/new repo(s); "
            f"{len(results)} served from cache"
        )

        def _task(item):
            (owner, repo), upd = item
            enr = self._github_fetch_feed_enrichment(owner, repo)
            enr["source_upd"] = upd or ""
            enr["fetched_at"] = self._today_iso()
            enr["logic_version"] = self.FEED_LOGIC_VERSION
            return (owner, repo), enr

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_item = {executor.submit(_task, it): it for it in stale}
            for future in as_completed(future_to_item):
                try:
                    key, enr = future.result()
                    results[key] = enr
                    cache[f"{key[0]}/{key[1]}"] = enr
                except Exception:
                    pass
        return results

    def _build_feed_scripts(self, rows, enrichment: dict) -> dict:
        """Assemble the consumer-keyed `scripts` map from merged rows + enrichment.
        Keys are project_name.toLowerCase(); every field is emitted only when
        truthy/valid so the consumer's per-field guards stay meaningful."""
        scripts = {}
        for row in rows:
            name = str(row.get("Name") or "").strip()
            if not name:
                continue
            entry = {}
            tags = self._tags_list(row.get("Tags"))
            if tags:
                entry["tags"] = tags
            upd = str(row.get("Last Updated") or "").strip()
            if re.match(r"^\d{4}-\d{2}-\d{2}$", upd):
                entry["upd"] = upd
            owner, repo = self._parse_github_repo(row.get("Project URL", ""))
            enr = enrichment.get((owner, repo)) if owner and repo else None
            if enr:
                if enr.get("engine"):
                    entry["engine"] = enr["engine"]
                if enr.get("facets"):
                    entry["facets"] = list(enr["facets"])
                if enr.get("nb"):
                    entry["nb"] = True
                if enr.get("nb_role"):
                    entry["nb_role"] = enr["nb_role"]
                if enr.get("readme"):
                    entry["readme"] = enr["readme"]
                if enr.get("images"):
                    entry["images"] = list(enr["images"])[: self.FEED_MAX_IMAGES]
                if enr.get("sha"):
                    entry["sha"] = enr["sha"]
            if entry:
                scripts[name.lower()] = entry
        return scripts

    def write_feed_json(self, rows, excel_path: str, output_path: str = None) -> None:
        """Build and write feed.json from merged rows. Best-effort: logs and
        returns on any failure so it can never abort a run whose xlsx already
        saved. `tags` and `upd` come for free from existing columns; engine/nb/
        readme/images come from cached GitHub enrichment."""
        import json

        try:
            repos = {}
            for row in rows:
                owner, repo = self._parse_github_repo(row.get("Project URL", ""))
                if owner and repo:
                    repos[(owner, repo)] = str(row.get("Last Updated") or "").strip()

            cache = self._load_feed_cache(excel_path)
            enrichment = self._gather_feed_enrichment(repos, cache) if repos else {}
            self._save_feed_cache(excel_path, cache)

            scripts = self._build_feed_scripts(rows, enrichment)
            payload = {
                "file_info": {"version": 2, "kind": "script_feed"},
                "date": self._today_iso(),
                "scripts": scripts,
            }
            out = output_path or self._default_feed_output(excel_path)
            with open(out, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, sort_keys=True, ensure_ascii=False)

            n_eng = sum(1 for v in scripts.values() if v.get("engine"))
            n_nb = sum(1 for v in scripts.values() if v.get("nb"))
            n_rd = sum(1 for v in scripts.values() if v.get("readme"))
            n_img = sum(1 for v in scripts.values() if v.get("images"))
            logger.info(
                f"Wrote feed.json: {len(scripts)} scripts "
                f"({n_eng} engine, {n_nb} nb, {n_rd} readme, {n_img} images) -> {out}"
            )
        except Exception as e:
            logger.warning(f"Failed to write feed.json: {e}")

    def regenerate_feed_only(self, excel_path: str, output_path: str = None) -> None:
        """Rebuild feed.json from an existing xlsx + cache, with no re-scrape.
        Cheap iteration path: only repos that changed since the cache was built
        hit GitHub."""
        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found: {excel_path}")
            return
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel for feed regeneration: {e}")
            return
        rows = df.fillna("").to_dict("records")
        self.write_feed_json(rows, excel_path, output_path)

    # ---------------------------
    # catalog.json — canonical, diffable, PR-able full-rows catalog.
    #
    # docs/build_data.py PREFERS catalog.json over the xlsx, so this is the
    # forward-looking source of truth for the public site; the xlsx becomes a
    # derived export (kept for the user's manual curation workflow). It carries
    # every FIELD_MAP column so the xlsx is reconstructable from it. Emitted as
    # a list under `scripts`, sorted by Name with sorted keys, so nightly diffs
    # are minimal and reviewable.
    # ---------------------------
    def _default_catalog_output(self, excel_path: str) -> str:
        """Default catalog.json location: the repo root (next to the xlsx),
        named exactly `catalog.json` — the path build_data.py defaults to."""
        return os.path.join(os.path.dirname(os.path.abspath(excel_path)), "catalog.json")

    @staticmethod
    def _catalog_clean(value):
        """Coerce a merged-row value to a clean JSON scalar: NaN/None -> '',
        everything else stripped to a string. (Lists are returned as-is for Tags.)"""
        if isinstance(value, list):
            return value
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            pass
        return str(value).strip() if value is not None else ""

    def _discovered_to_catalog_entry(self, rec: dict) -> dict:
        """Map a discovery record (GitHub-sourced) into a catalog entry. Carries
        GitHub basics (name/author/desc/proj/upd/tags=topics) + discovery-only
        fields (facets/stars/archived/source), plus any README/forum enrichment the
        record supplies (Demo, Discussion URL, readme, images). Community URL stays
        empty (discovered repos aren't on norns.community)."""
        entry = {c: "" for c in self.FIELD_MAP}
        entry["Name"] = rec.get("name") or ""
        entry["Author"] = rec.get("author") or ""
        entry["Description"] = rec.get("desc") or ""
        entry["Project URL"] = rec.get("proj") or ""
        entry["Last Updated"] = rec.get("upd") or ""
        entry["Demo"] = rec.get("demo") or ""
        entry["Discussion URL"] = rec.get("disc") or ""
        entry["Tags"] = list(rec.get("topics") or [])
        if rec.get("readme"):
            entry["readme"] = rec["readme"]
        if rec.get("images"):
            entry["images"] = list(rec["images"])
        if rec.get("sha"):
            entry["sha"] = rec["sha"]
        entry["source"] = "github"
        entry["facets"] = list(rec.get("facets") or [])
        entry["stars"] = int(rec.get("stars") or 0)
        entry["status"] = "archived" if rec.get("archived") else "active"
        if rec.get("archived"):
            entry["archived"] = True
        return entry

    def _community_status_map(self, rows, excel_path: str) -> dict:
        """name_lower -> status for community rows (roadmap #5). `delisted` when the
        name is gone from the live community.json, `missing` when the repo 404s this
        run, `archived` from the GitHub archived flag (this run's status map or the
        persisted Last-Updated cache), else `active`. Network-light: reuses the
        cached community.json + the committed Last-Updated cache."""
        comm = None
        try:
            entries = self.fetch_community_json()
            if entries:
                comm = {(e.get("project_name") or "").strip().lower() for e in entries}
                comm.discard("")
        except Exception:
            comm = None  # unknown -> never flag delisted (fail safe)
        lu = self._load_lastupd_cache(excel_path) if excel_path else {}
        run_status = getattr(self, "_repo_status_map", {}) or {}
        out = {}
        for row in rows:
            nm = self._catalog_clean(row.get("Name")).lower()
            if not nm:
                continue
            owner, repo = self._parse_github_repo(row.get("Project URL", ""))
            rst = run_status.get((owner, repo)) if owner else None
            archived = owner and (
                rst == "archived" or bool((lu.get(f"{owner}/{repo}") or {}).get("archived"))
            )
            if comm is not None and nm not in comm:
                out[nm] = "delisted"
            elif rst == "missing":
                out[nm] = "missing"
            elif archived:
                out[nm] = "archived"
            else:
                out[nm] = "active"
        return out

    def write_catalog_json(self, rows, excel_path: str, output_path: str = None,
                           discovered: dict = None) -> None:
        """Write catalog.json from merged community rows, plus any GitHub-discovered
        repos (source='github'). Best-effort: logs and returns on any failure so it
        can never abort a run whose xlsx already saved."""
        import json

        try:
            cols = list(self.FIELD_MAP.keys())
            status_map = self._community_status_map(rows, excel_path)
            scripts = []
            for row in rows:
                name = self._catalog_clean(row.get("Name"))
                if not name:
                    continue
                entry = {}
                for c in cols:
                    if c == "Tags":
                        entry[c] = self._tags_list(row.get("Tags"))
                    else:
                        entry[c] = self._catalog_clean(row.get(c, ""))
                entry["Name"] = name
                entry["source"] = "community"
                entry["status"] = status_map.get(name.lower(), "active")
                scripts.append(entry)

            # Append GitHub-discovered repos that aren't already a community script.
            # Dedupe by lowercase name — the site slug / install key — so forks and
            # renamed repos collapse onto the community entry (which always wins).
            # Among same-named GitHub repos, the most-starred wins (stable, not
            # arbitrary classification order).
            if discovered:
                have = {s["Name"].lower() for s in scripts}
                for rec in sorted(discovered.values(),
                                  key=lambda r: (-(r.get("stars") or 0), r.get("name", ""))):
                    nm = (rec.get("name") or "").strip()
                    if nm and nm.lower() not in have:
                        scripts.append(self._discovered_to_catalog_entry(rec))
                        have.add(nm.lower())

            scripts.sort(key=lambda r: r["Name"].lower())
            payload = {
                "file_info": {"version": 1, "kind": "script_catalog"},
                "date": self._today_iso(),
                "scripts": scripts,
            }
            out = output_path or self._default_catalog_output(excel_path)
            with open(out, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            n_gh = sum(1 for s in scripts if s.get("source") == "github")
            logger.info(
                f"Wrote catalog.json: {len(scripts)} scripts "
                f"({len(scripts) - n_gh} community, {n_gh} GitHub) -> {out}"
            )
        except Exception as e:
            logger.warning(f"Failed to write catalog.json: {e}")

    def _run_discovery(self, rows, excel_path: str) -> dict:
        """Extract the community repo set from merged rows and run GitHub +
        forum-driven discovery. Returns {(owner,repo): record} or {} on failure.
        Forum-found repos (repo-linked norns-tag threads not already known) are
        classified + enriched like GitHub-search finds, but additionally carry a
        `lines` tag, the thread's discussion URL, and a demo mined from the thread."""
        try:
            community = set()
            for row in rows:
                owner, repo = self._parse_github_repo(row.get("Project URL", ""))
                if owner and repo:
                    community.add((owner, repo))
            discovered = self.discover_github_repos(
                community, excel_path,
                aggressive=getattr(self, "discover_aggressive", True),
                max_author_searches=getattr(self, "discover_max_authors", None),
            )
            # Forum-driven soft launch: norns-tag threads that link a GitHub repo.
            # Exclude ONLY community repos (they carry their own disc/demo). The
            # "soft launch" signal is a repo that's on GitHub AND has a lines
            # thread but isn't on norns.community yet — so a repo that github
            # search ALSO found must be UPGRADED in place (add disc + `lines` tag),
            # not skipped. Repos seen only on the forum are classified + added new.
            forum = self.discover_forum_repos(set(community),
                                              max_pages=getattr(self, "forum_max_pages", 5))
            cache = self._load_discovery_cache(excel_path)
            lock = threading.Lock()
            new_recs = {}
            for (owner, name), meta in forum.items():
                # Isolate each forum repo: a single bad repo must never discard the
                # already-complete github-search `discovered` set (the outer except
                # would otherwise swallow everything and return {}).
                try:
                    existing = discovered.get((owner, name))
                    if existing is not None:
                        # github search already has it -> promote to soft launch
                        existing["disc"] = meta["disc"]
                        if "lines" not in (existing.get("topics") or []):
                            existing.setdefault("topics", []).append("lines")
                        continue
                    m = self._repo_meta(owner, name)
                    if m.get("_status") in (403, 404):
                        continue
                    branch = m.get("default_branch") or "main"
                    verdict = self._classify_norns_repo(
                        owner, name, branch, m.get("pushed_at"), cache, lock)
                    if not verdict.get("is_norns"):
                        continue
                    new_recs[(owner, name)] = {
                        "owner": owner, "name": name,
                        "author": owner, "desc": m.get("description") or "",
                        "proj": f"https://github.com/{owner}/{name}",
                        "upd": str(m.get("pushed_at") or "")[:10],
                        "topics": ["lines"], "facets": verdict.get("facets") or [],
                        "archived": bool(m.get("archived")),
                        "stars": int(m.get("stargazers_count") or 0),
                        "source": "github", "demo": "", "readme": "", "images": [],
                        "disc": meta["disc"],
                    }
                except Exception as exc:
                    logger.debug(f"Forum classify failed for {owner}/{name}: {exc}")
                    continue
            self._save_discovery_cache(excel_path, cache)
            self._enrich_discovered(new_recs, excel_path)  # cached README/demo/images
            # Thread-demo fallback (upgraded + new) where the README had no demo.
            for (owner, name), meta in forum.items():
                rec = discovered.get((owner, name)) or new_recs.get((owner, name))
                if rec and not rec.get("demo"):
                    try:
                        rec["demo"] = self.discover_demo_video(meta["disc"]) or ""
                    except Exception:
                        pass
            discovered.update(new_recs)
            return discovered
        except Exception as e:
            logger.warning(f"Discovery failed (catalog still written, community-only): {e}")
            return {}

    def regenerate_catalog_only(self, excel_path: str, output_path: str = None,
                                discover: bool = False) -> None:
        """Rebuild catalog.json from an existing xlsx, with no community re-scrape.
        With discover=True, also run GitHub discovery and fold in net-new repos.
        Use this to seed/refresh the canonical catalog from the spreadsheet."""
        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found: {excel_path}")
            return
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel for catalog regeneration: {e}")
            return
        rows = df.fillna("").to_dict("records")
        discovered = self._run_discovery(rows, excel_path) if discover else None
        self.write_catalog_json(rows, excel_path, output_path, discovered)

    # ===========================================================================
    # GitHub discovery (roadmap #3): broaden the catalog beyond community.json.
    #
    # community.json is the curated ground truth (~351 repos) but a search-only
    # crawl tops out at ~72% of it (forks, SuperCollider engines, renames, and a
    # long tail of repos with no "norns" keyword/topic). So we SEED from
    # community.json (100% guaranteed) and UNION several GitHub search strategies
    # to surface net-new norns repos. Every net-new candidate is gated by the
    # runtime-API fingerprint (NORNS_FP) — the same signal ingenue uses, verified
    # 343/343 on the community corpus — so generic-lua cruft stays out. Verdicts
    # cache per repo keyed on pushed_at, so nightly runs only reclassify
    # changed/new repos.
    # ===========================================================================

    # Runtime-API fingerprint, ported verbatim from ingenue's classifier.
    NORNS_FP = re.compile(
        r"engine\.|softcut\.|\bscreen\.|params:add|function init\(|function redraw\(|"
        r"function enc\(|function key\(|controlspec|musicutil|grid\.connect|"
        r"arc\.connect|metro\.|norns\.|_norns|mod\.hook|mod\.menu"
    )
    # norns infra / lists / hardware — never installable scripts.
    GH_BLOCK = {
        "monome/norns", "monome/norns-shield", "okyeron/shieldxl",
        "p3r7/awesome-monome-norns", "monome/norns-image", "monome/dust",
        "monome/norns-community", "figrhed/norns-on-raspberry-pi",
        "jguzak/shieldxl_battery", "seajaysec/ingenue",
    }
    GH_BLOCK_NAMES = {"ingenue"}
    DISCOVERY_CACHE_TTL_DAYS = 30  # reclassify an unchanged repo at most this stale
    # Bump when classifier logic changes — invalidates cached verdicts so they rebuild.
    DISCOVERY_LOGIC_VERSION = 1

    def _discovery_cache_path(self, excel_path: str) -> str:
        base, _ = os.path.splitext(excel_path)
        return f"{base}.discovery_cache.json"

    def _load_discovery_cache(self, excel_path: str) -> dict:
        import json

        path = self._discovery_cache_path(excel_path)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                logger.info(f"Loaded {len(data)} discovery-cache entries from {path}")
                return data
        except Exception as e:
            logger.warning(f"Failed to load discovery cache from {path}: {e}")
        return {}

    def _save_discovery_cache(self, excel_path: str, cache: dict) -> None:
        import json

        path = self._discovery_cache_path(excel_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cache, f, indent=2, sort_keys=True, ensure_ascii=False)
            logger.info(f"Saved {len(cache)} discovery-cache entries to {path}")
        except Exception as e:
            logger.warning(f"Failed to save discovery cache to {path}: {e}")

    def _discovery_fresh(self, entry: dict, pushed_at: str) -> bool:
        """A classification is reusable while the repo hasn't been pushed since we
        classified it (same pushed_at) and it's within TTL."""
        if not isinstance(entry, dict):
            return False
        if entry.get("logic_version") != self.DISCOVERY_LOGIC_VERSION:
            return False
        cached_pa = str(entry.get("pushed_at") or "")
        if not cached_pa or cached_pa != str(pushed_at or ""):
            return False
        classified_at = str(entry.get("classified_at") or "")[:10]
        try:
            from datetime import datetime

            age = (datetime.now() - datetime.strptime(classified_at, "%Y-%m-%d")).days
            return age <= self.DISCOVERY_CACHE_TTL_DAYS
        except Exception:
            return False

    def _search_repos(self, q: str, max_pages: int = 10, sort: str = "updated") -> list:
        """Paginated GitHub repo search -> list of item dicts. Stops at the GitHub
        1000-result cap or `max_pages`. Degrades to a partial/empty list on
        rate-limit or error so discovery never aborts a run."""
        items, seen_ids = [], set()
        for page in range(1, max_pages + 1):
            tries = 0
            while True:
                tries += 1
                self._throttle_search()
                try:
                    r = self.github_session.get(
                        "https://api.github.com/search/repositories",
                        params={"q": q, "per_page": 100, "page": page,
                                "sort": sort, "order": "desc"},
                        timeout=30,
                    )
                except Exception as e:
                    logger.debug(f"Discovery: search error {q!r} p{page}: {e}")
                    return items
                if r.status_code == 403 and tries <= 2:
                    # primary/secondary rate limit — honor Retry-After, then retry once.
                    wait = int(r.headers.get("Retry-After") or 0) or 15
                    logger.warning(f"Discovery: search 403 on {q!r} p{page}; backing off {wait}s")
                    time.sleep(wait)
                    continue
                break
            if r.status_code != 200:
                if r.status_code == 403:
                    logger.warning(f"Discovery: search still rate-limited on {q!r} p{page}; partial")
                break
            try:
                batch = r.json().get("items") or []
            except Exception:
                break
            for it in batch:
                if it.get("id") not in seen_ids:
                    seen_ids.add(it.get("id"))
                    items.append(it)
            if len(batch) < 100:
                break
        return items

    def _throttle_search(self) -> None:
        """Space out GitHub search requests to respect the ~30 req/min limit."""
        with self._search_lock:
            now = time.monotonic()
            wait = self._search_min_interval - (now - self._search_last)
            if wait > 0:
                time.sleep(wait)
            self._search_last = time.monotonic()

    def _classify_norns_repo(self, owner: str, repo: str, branch: str, pushed_at: str,
                             cache: dict, cache_lock) -> dict:
        """Gate a candidate repo on the norns fingerprint. Structural proof
        (lib/mod.lua or a .sc engine / engine/ dir) short-circuits with no content
        read; otherwise read ONE key file and test NORNS_FP. Returns
        {is_norns, facets, ...}. Cached on pushed_at; transient/unknown verdicts
        are NOT cached so they self-heal."""
        ck = f"{owner}/{repo}"
        cached = cache.get(ck)
        if cached and self._discovery_fresh(cached, pushed_at):
            return cached

        result = {"is_norns": None, "facets": []}  # None = couldn't determine (don't cache)
        try:
            tr = self.github_session.get(
                f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch or 'HEAD'}",
                params={"recursive": "1"}, timeout=15,
            )
            if tr.status_code == 403 or tr.status_code != 200:
                return result  # rate-limited / unreadable -> undetermined
            paths = [str(t.get("path", "")) for t in (tr.json().get("tree") or [])
                     if t.get("type") == "blob"]
            top_lua = [p for p in paths if "/" not in p and p.lower().endswith(".lua")]
            has_mod = any(re.match(r"^lib/mod\.lua$", p, re.I) for p in paths)
            lib_lua = [p for p in paths if p.lower().endswith(".lua") and re.search(r"(^|/)lib/", p, re.I)]
            # `.sc` files are the norns SuperCollider-engine convention -> strong proof.
            # An `engine/` *directory* alone is too generic (kernels, generic audio
            # projects) -> only a facet hint; such repos still must pass the content gate.
            has_sc = any(p.lower().endswith(".sc") for p in paths)
            is_norns = bool(has_mod or has_sc)  # structural proof — no content read

            if not is_norns:
                keyfile = (next((p for p in top_lua if p.lower() == f"{repo.lower()}.lua"), None)
                           or (top_lua[0] if top_lua else None)
                           or ("lib/mod.lua" if has_mod else None)
                           or (lib_lua[0] if lib_lua else None))
                if keyfile:
                    from urllib.parse import quote

                    seg = "/".join(quote(s) for s in keyfile.split("/"))
                    fr = self.github_session.get(
                        f"https://api.github.com/repos/{owner}/{repo}/contents/{seg}",
                        headers={"Accept": "application/vnd.github.raw"},
                        params={"ref": branch}, timeout=15,
                    )
                    if fr.status_code == 200 and self.NORNS_FP.search(fr.text):
                        is_norns = True
                elif not paths:
                    return result  # empty/unreadable tree -> undetermined

            result = {
                "is_norns": is_norns,
                "facets": self._facets_from_paths(paths),
                "pushed_at": pushed_at or "",
                "classified_at": self._today_iso(),
                "logic_version": self.DISCOVERY_LOGIC_VERSION,
            }
            if pushed_at:  # only cache a determinate verdict with a change signal
                with cache_lock:
                    cache[ck] = result
        except Exception as e:
            logger.debug(f"Discovery: classify error {owner}/{repo}: {e}")
        return result

    def _enrich_discovered(self, records: dict, excel_path: str) -> None:
        """Fill demo/readme/images on discovered records via the CACHED feed
        enrichment path (keyed by pushed_at in feed_cache.json). Unchanged repos
        are served from cache, so a steady-state nightly run only fetches READMEs
        for repos that actually changed — not all ~1.2k passing candidates every
        night. Mutates `records` in place. Each rec must carry 'upd' (pushed_at).
        Best-effort: on any failure the records keep their (empty) demo/readme."""
        if not records:
            return
        try:
            repos = {k: (rec.get("upd") or "") for k, rec in records.items()}
            cache = self._load_feed_cache(excel_path)
            enr_map = self._gather_feed_enrichment(repos, cache)
            self._save_feed_cache(excel_path, cache)
        except Exception as e:
            logger.warning(f"Discovered-repo enrichment failed (records left bare): {e}")
            return
        for k, rec in records.items():
            enr = enr_map.get(k) or {}
            if not rec.get("demo"):
                rec["demo"] = enr.get("demo") or ""
            if not rec.get("readme"):
                rec["readme"] = enr.get("readme") or ""
            if not rec.get("images"):
                rec["images"] = list(enr.get("images") or [])
            # HEAD sha rides along from the same cached enrichment — lets ingenue
            # detect updates for discovered (gh-exclusive / soft-launch) installs
            # by diffing the local sha against this, with no GitHub call of its own.
            if not rec.get("sha"):
                rec["sha"] = enr.get("sha") or ""

    def discover_github_repos(self, community_repos: set, excel_path: str,
                              aggressive: bool = True,
                              max_author_searches: int = None) -> dict:
        """Union GitHub search strategies, gate net-new candidates on NORNS_FP,
        and return {(owner, repo): record} for repos that pass — each record
        carrying the fields a catalog row needs (name/author/desc/proj/upd/tags/
        facets/archived/source). community repos are excluded here (they're seeded
        separately and known-norns). The per-author sweep seeds from the repo
        OWNERS in community_repos (the actual GitHub logins) — community.json's
        `author` field is a human display name, not a username. Read-only."""
        cache = self._load_discovery_cache(excel_path)
        candidates = {}  # (owner, name) -> richest item dict

        def add(items):
            for it in items:
                owner = (it.get("owner") or {}).get("login", "").lower()
                name = (it.get("name") or "").lower()
                if not owner or not name:
                    continue
                if f"{owner}/{name}" in self.GH_BLOCK or name in self.GH_BLOCK_NAMES:
                    continue
                candidates.setdefault((owner, name), it)

        strategies = ["norns language:lua", "topic:norns", "norns"]
        if aggressive:
            strategies.append("fork:true norns language:lua")
        for q in strategies:
            add(self._search_repos(q))
            logger.info(f"Discovery: after [{q}] -> {len(candidates)} unique candidates")

        if aggressive:
            owners = sorted({o for (o, _) in community_repos if o})
            if max_author_searches is not None:
                owners = owners[:max_author_searches]
            logger.info(f"Discovery: per-author sweep over {len(owners)} community repo owners")
            for i, o in enumerate(owners, 1):
                add(self._search_repos(f"user:{o} language:lua", max_pages=3))
                if i % 25 == 0:
                    logger.info(f"Discovery: author sweep {i}/{len(owners)} -> "
                                f"{len(candidates)} candidates")

        netnew = {k: it for k, it in candidates.items() if k not in community_repos}
        logger.info(f"Discovery: {len(candidates)} total candidates, "
                    f"{len(netnew)} net-new (not on community); classifying...")

        discovered, cache_lock = {}, threading.Lock()

        def _task(item):
            (owner, name), it = item
            verdict = self._classify_norns_repo(
                owner, name, it.get("default_branch"), it.get("pushed_at"), cache, cache_lock)
            if not verdict.get("is_norns"):
                return None
            # Classification only — demo/readme/images are filled in one cached
            # batch (_enrich_discovered) after the gate, so README fetches happen
            # only for repos that changed since the last run.
            return (owner, name), {
                "owner": owner, "name": it.get("name") or name,
                "author": (it.get("owner") or {}).get("login", ""),
                "desc": it.get("description") or "",
                "proj": it.get("html_url") or f"https://github.com/{owner}/{name}",
                "upd": (it.get("pushed_at") or "")[:10],
                "topics": (it.get("topics") or [])[:8],
                "facets": verdict.get("facets") or [],
                "archived": bool(it.get("archived")),
                "stars": it.get("stargazers_count") or 0,
                "source": "github",
                "demo": "", "readme": "", "images": [], "disc": "",
            }

        with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
            for fut in as_completed({ex.submit(_task, it): it for it in netnew.items()}):
                try:
                    res = fut.result()
                    if res:
                        discovered[res[0]] = res[1]
                except Exception:
                    pass

        self._save_discovery_cache(excel_path, cache)
        self._enrich_discovered(discovered, excel_path)  # cached README/demo/images
        logger.info(f"Discovery: {len(discovered)}/{len(netnew)} net-new candidates "
                    f"passed the norns gate")
        return discovered

    # The lines forum (Discourse). NOT self.base_url, which is norns.community —
    # the community-scrape site, whose /tag/norns.json 404s.
    DISCOURSE_FORUM_BASE = "https://llllllll.co"

    def discover_forum_repos(self, known_repos: set, max_pages: int = 5) -> dict:
        """Crawl llllllll.co's `norns` tag for topics whose OP links a GitHub repo
        not already in `known_repos`. Returns {(owner, repo): {disc, topic_id}}.
        Best-effort: any error yields a partial/empty dict (never aborts a run)."""
        base = self.DISCOURSE_FORUM_BASE
        out = {}
        for page in range(0, max_pages):
            try:
                r = self._discourse_get_with_retry(
                    f"{base}/tag/norns.json", params={"page": page}, timeout=20)
                if r.status_code != 200:
                    break
                topics = ((r.json() or {}).get("topic_list") or {}).get("topics") or []
            except Exception as e:
                logger.debug(f"Forum discovery: tag page {page} error: {e}")
                break
            if not topics:
                break
            for t in topics:
                tid = t.get("id")
                if not tid:
                    continue
                try:
                    tr = self._discourse_get_with_retry(f"{base}/t/{tid}.json", timeout=20)
                    if tr.status_code != 200:
                        continue
                    posts = ((tr.json() or {}).get("post_stream") or {}).get("posts") or []
                    op = posts[0].get("cooked", "") if posts else ""
                except Exception:
                    continue
                gh = self._extract_github_url(op)
                if not gh or gh in known_repos or gh in out:
                    continue
                if f"{gh[0]}/{gh[1]}" in self.GH_BLOCK or gh[1] in self.GH_BLOCK_NAMES:
                    continue
                out[gh] = {"disc": f"{base}/t/{t.get('slug') or tid}/{tid}", "topic_id": tid}
        logger.info(f"Forum discovery: {len(out)} repo-linked norns-tag threads not in catalog")
        return out

    def _scrape_by_community_url(self, community_url: str):
        """Build script details for sync-check by looking up the community.json entry."""
        # Derive slug from URL path
        try:
            slug = community_url.replace("https://norns.community/", "").strip("/")
        except Exception:
            slug = community_url
        json_entry = self._lookup_json_entry_by_slug(slug)
        return self.scrape_script_details(
            community_url,
            slug,
            existing_data=None,
            discover_demo=False,
            json_entry=json_entry,
        )

    def reset_discoverable_demos(self, excel_path: str) -> int:
        """Clear Demo + Playwright Status for every row that isn't a Manual Override.

        Use this after a wave of bad external-search results has polluted the
        xlsx. Manual edits (Playwright Status == 'Manual Override') are
        preserved. Also deletes the external_searches.json sidecar so cached
        false-positives don't reseed on the next run.

        Returns the number of rows whose Demo cell was cleared.
        """
        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found: {excel_path}")
            return 0
        df = pd.read_excel(excel_path)
        if "Demo" not in df.columns or "Playwright Status" not in df.columns:
            logger.error("Excel missing Demo or Playwright Status column")
            return 0

        cleared = 0
        for idx, row in df.iterrows():
            status = row.get("Playwright Status", "")
            status_str = str(status).strip() if pd.notna(status) else ""
            if status_str == "Manual Override":
                continue  # preserve user-curated rows
            demo = row.get("Demo", "")
            had_demo = pd.notna(demo) and str(demo).strip() and str(demo) != "nan"
            if had_demo or status_str:
                df.at[idx, "Demo"] = ""
                df.at[idx, "Playwright Status"] = ""
                if had_demo:
                    cleared += 1

        # Save back via the formatted writer (preserves table style + hyperlinks).
        try:
            self._write_formatted_excel_from_df(df, excel_path)
        except Exception as e:
            logger.warning(f"Formatted save failed; falling back to plain write: {e}")
            df.to_excel(excel_path, index=False, engine="openpyxl")

        # Delete external-search cache so cached false-positives don't reseed.
        sidecar = self._external_search_path(excel_path)
        if os.path.exists(sidecar):
            try:
                os.remove(sidecar)
                logger.info(f"Removed cached external searches: {sidecar}")
            except Exception as e:
                logger.warning(f"Could not remove {sidecar}: {e}")

        return cleared

    def sync_check_only(self, excel_path: str) -> int:
        """Compute 'Out of Sync' for all rows without skipping logic; save back to Excel.

        Returns number of rows whose 'Out of Sync' value changed.
        """
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found: {excel_path}")
                return 0
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel for sync-check: {e}")
            return 0

        # Ensure columns exist
        if "Community URL" not in df.columns:
            logger.error("Excel missing 'Community URL' column")
            return 0
        if "Out of Sync" not in df.columns:
            df["Out of Sync"] = ""
        # Ensure dtype supports string assignment to avoid FutureWarning
        try:
            df["Out of Sync"] = (
                df["Out of Sync"]
                .astype("object")
                .where(pd.notna(df["Out of Sync"]), "")
            )
        except Exception:
            pass

        # Prepare tasks: only rows with a non-empty Community URL
        tasks = []
        for idx, row in df.iterrows():
            community_url = row.get("Community URL", "")
            if isinstance(community_url, str) and community_url.strip():
                tasks.append((idx, community_url.strip()))

        logger.info(f"Sync-check: processing {len(tasks)} row(s) from Excel")

        # Pre-warm caches in the main thread to avoid each worker independently
        # re-fetching community.json and the index page (the lazy `if cache is None`
        # check isn't thread-safe under ThreadPoolExecutor).
        self.fetch_community_json()
        self.fetch_slug_map()
        self._lookup_json_entry_by_slug("__warmup__")

        # Scrape in parallel
        scraped_by_idx = {}
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_idx = {
                executor.submit(self._scrape_by_community_url, url): idx
                for idx, url in tasks
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    scraped = future.result() or {}
                except Exception:
                    scraped = {}
                scraped_by_idx[idx] = scraped

        # Load drift snapshots; pass them through the shared drift helper.
        # Sync-check has no merge step, so the "post-merge xlsx" we hand to
        # _compute_drift is simply the existing xlsx row.
        snapshots = self._load_snapshots(excel_path)

        # Compute and apply Out of Sync
        changed = 0
        for idx, community_url in tasks:
            existing_row = df.iloc[idx]
            existing_dict = existing_row.to_dict()
            scraped = scraped_by_idx.get(idx) or {}
            slug = self._slug_for_row(community_url)
            snap_for_slug = snapshots.setdefault(slug, {}) if slug else None

            new_value = self._compute_drift(
                existing_dict, scraped, existing_dict, snap_for_slug
            )
            old_value = existing_row.get("Out of Sync", "")
            if str(old_value) != str(new_value):
                df.at[idx, "Out of Sync"] = new_value
                changed += 1

        # Save back using the same formatting as main save path
        try:
            self._write_formatted_excel_from_df(df, excel_path)
        except Exception as e:
            logger.error(f"Failed to save formatted Excel after sync-check: {e}")
            try:
                df.to_excel(excel_path, index=False, engine="openpyxl")
            except Exception:
                pass

        # Persist snapshot updates the drift helper made during this sync-check
        self._save_snapshots(excel_path, snapshots)
        return changed

    def discover_demos_unified(self, existing_df):
        """Single function to handle all demo discovery with both standard and Playwright methods"""
        if existing_df is None:
            logger.info("No existing data - skipping demo discovery")
            return

        # Create lookup for existing scripts
        existing_scripts = {}
        for _, row in existing_df.iterrows():
            script_name = row["Name"]
            existing_scripts[script_name] = {
                "has_demo": pd.notna(row["Demo"]) and row["Demo"] != "",
                "existing_demo_url": (
                    row["Demo"] if pd.notna(row["Demo"]) and row["Demo"] != "" else ""
                ),
                "playwright_status": (
                    row["Playwright Status"]
                    if "Playwright Status" in existing_df.columns
                    and pd.notna(row["Playwright Status"])
                    else ""
                ),
            }

        # Filter scraped scripts to those needing demo discovery.
        #
        # The parallel scrape pass (`scrape_script_details_from_json`, called
        # earlier from scrape_all_scripts with discover_demo=True) already runs
        # `discover_demo_video` inline and writes the result into
        # script_data["demo"]. We treat THAT as the authoritative discovery
        # pass; this function exists only to fill rows that the inline pass
        # didn't cover.
        #
        # Why this matters: previously, this function re-ran discovery
        # whenever Playwright Status was blank — which is the state every row
        # is in immediately after `--reset-discoverable-demos`. During a
        # Discourse 429 storm the second pass would 429-fail and overwrite
        # the inline pass's good URL with empty + "Missing Demo". Concrete
        # case from the May-6 run: awake-passersby was discovered cleanly at
        # T+25s (BYjTCWS-B7o), then clobbered to "Missing Demo" at T+8m when
        # the second pass hit awake.json's rate limit. Skipping rows where
        # the inline pass already produced a URL eliminates the overwrite.
        scripts_needing_demos = []
        for script_data in self.script_data:
            project_name = script_data["project_name"]

            inline_demo = (script_data.get("demo") or "").strip()
            if inline_demo:
                # Inline pass found a URL. Stamp a clean (no special-state)
                # status so the saver writes both fields atomically and the
                # next run skips this row instead of triggering re-discovery
                # because of a blank status field.
                script_data["playwright_status"] = ""
                continue

            if project_name in existing_scripts:
                existing = existing_scripts[project_name]
                # Run demo discovery if:
                # 1. No demo URL exists in the loaded xlsx, OR
                # 2. Playwright Status is blank (post-reset rows, or rows
                #    that have never been processed by this scraper).
                needs_demo = (
                    (
                        not existing["has_demo"]
                        or existing["playwright_status"] == ""
                        or pd.isna(existing["playwright_status"])
                    )
                    and script_data["discussion_url"]
                    and pd.notna(script_data["discussion_url"])
                    and script_data["discussion_url"] != ""
                )
            else:
                # New script - check if it needs demo discovery
                needs_demo = (
                    script_data["discussion_url"]
                    and pd.notna(script_data["discussion_url"])
                    and script_data["discussion_url"] != ""
                )

            if needs_demo:
                scripts_needing_demos.append(
                    {
                        "name": project_name,
                        "discussion_url": script_data["discussion_url"],
                        "author": script_data.get("author", "") or "",
                    }
                )

        if not scripts_needing_demos:
            logger.info("No scripts need demo discovery")
            return

        logger.info(
            f"Processing {len(scripts_needing_demos)} scripts for demo discovery with {self.max_workers} workers"
        )

        # Use ThreadPoolExecutor for parallel demo discovery
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks - both regular and playwright
            future_to_script = {}
            for script in scripts_needing_demos:
                discussion_url = script["discussion_url"]

                # Skip if discussion URL is invalid
                if (
                    not discussion_url
                    or pd.isna(discussion_url)
                    or discussion_url == ""
                ):
                    logger.debug(
                        f"Skipping {script['name']} - invalid discussion URL: {discussion_url}"
                    )
                    continue

                # Discourse-only: single discovery method, no parallel Playwright,
                # no conflict reconciliation step. Just submit and harvest.
                # Pass author/script_name through so Phase 2 external search can fire.
                future_to_script[
                    executor.submit(
                        self.discover_demo_video,
                        discussion_url,
                        False,
                        script.get("author", ""),
                        script["name"],
                    )
                ] = script

            demos_found = 0
            completed = 0
            results = {}

            for future in as_completed(future_to_script):
                script = future_to_script[future]
                completed += 1
                try:
                    results[script["name"]] = future.result() or ""
                except Exception as e:
                    logger.warning(
                        f"Error in demo discovery for {script['name']}: {e}"
                    )
                    results[script["name"]] = ""

                if completed % 10 == 0 or completed == len(future_to_script):
                    logger.info(
                        f"Demo discovery progress: {completed}/{len(future_to_script)}"
                    )

            # Apply results. Status field is kept ("Playwright Status" Excel column)
            # for A/B parity and historical row preservation, but we only ever
            # emit "Missing Demo" or "Manual Override" in this variant.
            for script in scripts_needing_demos:
                script_name = script["name"]
                discovered_url = results.get(script_name, "")
                existing_demo_url = ""
                if script_name in existing_scripts:
                    existing_demo_url = existing_scripts[script_name][
                        "existing_demo_url"
                    ]

                # Decide final demo + status
                if discovered_url:
                    final_url = discovered_url
                    status = ""  # found cleanly via API; no special status
                    demos_found += 1
                    logger.info(
                        f"Discovered demo for {script_name}: {discovered_url}"
                    )
                elif existing_demo_url:
                    final_url = existing_demo_url
                    status = "Manual Override"
                    logger.info(
                        f"No demo discovered for {script_name}, "
                        f"preserving existing URL as Manual Override"
                    )
                else:
                    final_url = ""
                    status = "Missing Demo"
                    logger.info(f"No demo discovered for {script_name} (Missing Demo)")

                # Stamp onto self.script_data, updating the existing entry if present.
                existing_script_data = next(
                    (
                        sd
                        for sd in self.script_data
                        if sd["project_name"] == script_name
                    ),
                    None,
                )
                if existing_script_data is not None:
                    existing_script_data["demo"] = final_url
                    existing_script_data["playwright_status"] = status
                else:
                    self.script_data.append(
                        {
                            "project_name": script_name,
                            "project_url": "",
                            "author": "",
                            "description": "",
                            "discussion_url": script["discussion_url"],
                            "tags": "",
                            "demo": final_url,
                            "community_url": "",
                            "playwright_status": status,
                        }
                    )

        logger.info(f"Found {demos_found} demos for scripts needing discovery")

    def test_single_script(self, community_url):
        """Test mode: process only a single script by Community URL - just filter the main flow"""
        # Extract script name from URL
        if not community_url.startswith("https://norns.community/"):
            logger.error(
                f"Invalid Community URL: {community_url}. Must start with https://norns.community/"
            )
            return

        script_name = community_url.replace("https://norns.community/", "").strip("/")
        if not script_name:
            logger.error(f"Could not extract script name from URL: {community_url}")
            return

        logger.info(
            f"Test mode: Processing script '{script_name}' from {community_url}"
        )

        # Use the main scraping flow but filter to only process this one script
        self.scrape_all_scripts(test_filter=script_name)


    def load_existing_data(self, filename="norns_scripts.xlsx"):
        """Load existing Excel data if it exists"""
        try:
            if os.path.exists(filename):
                existing_df = pd.read_excel(filename)
                logger.info(f"Loaded existing data with {len(existing_df)} scripts")
                return existing_df
            else:
                logger.info("No existing Excel file found, will create new one")
                return None
        except Exception as e:
            logger.warning(f"Could not load existing Excel file: {e}")
            return None

    def merge_data(self, new_data, existing_df=None, snapshots=None):
        """Merge new scraped data with existing data, preserving manual corrections.

        If `snapshots` is provided, it's mutated in place with snapshot updates;
        the caller is responsible for persisting it via _save_snapshots.
        """
        if snapshots is None:
            snapshots = {}
        if existing_df is None:
            logger.info("No existing data to merge with")
            # Convert new data to Excel format
            excel_data = []
            for script in new_data:
                excel_script = {}
                for excel_col, internal_key in self.FIELD_MAP.items():
                    excel_script[excel_col] = script.get(internal_key, "")
                excel_data.append(excel_script)
                # First-run path: seed snapshots so future drift detection
                # has a baseline. _compute_drift handles the missing-snapshot
                # case by initializing without flagging anything.
                slug = self._slug_for_row(excel_script)
                if slug:
                    snap_for_slug = snapshots.setdefault(slug, {})
                    self._compute_drift(None, script, excel_script, snap_for_slug)

            # Store summary stats for first run
            added_details = []
            for script in excel_data:
                populated_fields = []
                for excel_col in self.FIELD_MAP.keys():
                    if (
                        excel_col != "Playwright Status" and script[excel_col]
                    ):  # Skip Playwright Status for summary
                        populated_fields.append(excel_col)

                added_details.append(
                    {"name": script["Name"], "fields": populated_fields}
                )

            self.summary_stats = {
                "scripts_added": len(excel_data),
                "scripts_updated": 0,
                "scripts_preserved": 0,
                "total_scripts": len(excel_data),
                "added_details": added_details,
                "updated_details": [],
            }

            return excel_data

        logger.info(
            f"Merging {len(new_data)} new scripts with {len(existing_df)} existing scripts"
        )

        # Convert existing DataFrame to list of dicts for easier processing
        existing_scripts = existing_df.to_dict("records")

        # Create lookup dictionaries for existing scripts
        existing_by_name = {
            script.get("Name", ""): script for script in existing_scripts
        }
        existing_by_url = {}
        for script in existing_scripts:
            community_url = script.get("Community URL", "")
            if community_url and isinstance(community_url, str):
                url_path = community_url.replace("https://norns.community/", "").strip(
                    "/"
                )
                if url_path:
                    existing_by_url[url_path] = script

        merged_data = []
        added_count = 0
        updated_count = 0
        preserved_count = 0
        added_details = []
        updated_details = []

        # Helper to merge two row dicts (existing + new) with rules above
        def _merge_rows(existing_script, new_script):
            nonlocal preserved_count, updated_count
            merged_script = {}
            updated_fields = []
            for excel_col, internal_key in self.FIELD_MAP.items():
                existing_value = (
                    existing_script.get(excel_col, "") if existing_script else ""
                )
                new_value = new_script.get(internal_key, "")

                if excel_col == "Playwright Status":
                    if new_value is not None and str(new_value).strip() != "":
                        if str(existing_value) != str(new_value):
                            updated_fields.append(excel_col)
                        merged_script[excel_col] = new_value
                    else:
                        merged_script[excel_col] = existing_value
                    continue

                # Always prefer newly computed Last Updated when available
                if excel_col == "Last Updated":
                    if str(new_value).strip():
                        if str(existing_value) != str(new_value):
                            updated_fields.append(excel_col)
                        merged_script[excel_col] = new_value
                    else:
                        merged_script[excel_col] = existing_value
                    continue

                if (
                    pd.isna(existing_value)
                    or existing_value == ""
                    or str(existing_value).strip() == ""
                    or str(existing_value) == "nan"
                ):
                    merged_script[excel_col] = new_value
                    if new_value != "":
                        updated_fields.append(excel_col)
                else:
                    merged_script[excel_col] = existing_value
                    preserved_count += 1
            return merged_script, updated_fields

        # Drift detection moved to self._compute_drift (snapshot-aware, three-way diff).
        # Snapshot mutations happen in place on the `snapshots` dict.

        # Process new scraped data (prefer URL-path keyed matching)
        processed_keys = set()  # Prefer URL path as key; fallback to Name
        for new_script in new_data:
            script_name = new_script.get("project_name", "")
            community_url = new_script.get("community_url", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )

            existing_script = None
            key_used = None
            if url_key and url_key in existing_by_url:
                existing_script = existing_by_url[url_key]
                key_used = ("url", url_key)
            elif script_name in existing_by_name:
                existing_script = existing_by_name[script_name]
                key_used = ("name", script_name)

            # Resolve the snapshot bucket for this row by URL slug
            slug_for_snapshot = url_key or self._slug_for_row(new_script)
            snap_for_slug = (
                snapshots.setdefault(slug_for_snapshot, {})
                if slug_for_snapshot
                else None
            )

            if existing_script is not None:
                merged_script, updated_fields = _merge_rows(existing_script, new_script)
                # Snapshot-aware drift: compares JSON to last-known snapshot, not
                # to xlsx, so user overrides are silent until upstream actually moves.
                merged_script["Out of Sync"] = self._compute_drift(
                    existing_script, new_script, merged_script, snap_for_slug
                )
                if updated_fields:
                    updated_count += 1
                    updated_details.append(
                        {"name": script_name, "fields": updated_fields}
                    )
                merged_data.append(merged_script)
                logger.debug(
                    f"Merged existing script by {'Community URL' if key_used and key_used[0]=='url' else 'Name'}: {script_name}"
                )
            else:
                # New script, add it
                merged_script = {}
                for excel_col, internal_key in self.FIELD_MAP.items():
                    merged_script[excel_col] = new_script.get(internal_key, "")
                # First sighting -> snapshot gets seeded with current JSON,
                # nothing flagged.
                merged_script["Out of Sync"] = self._compute_drift(
                    None, new_script, merged_script, snap_for_slug
                )
                populated_fields = []
                for excel_col in self.FIELD_MAP.keys():
                    if excel_col != "Playwright Status" and merged_script[excel_col]:
                        populated_fields.append(excel_col)
                added_details.append({"name": script_name, "fields": populated_fields})
                merged_data.append(merged_script)
                added_count += 1
                logger.debug(f"Added new script: {script_name}")

            # Track processed key
            if url_key:
                processed_keys.add(("url", url_key))
            else:
                processed_keys.add(("name", script_name))

        # Add any existing scripts that weren't in the new data (key by URL path when present)
        for existing_script in existing_scripts:
            community_url = existing_script.get("Community URL", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )
            key = (
                ("url", url_key)
                if url_key
                else ("name", existing_script.get("Name", ""))
            )
            if key not in processed_keys:
                merged_data.append(existing_script)
                processed_keys.add(key)
                logger.debug(
                    f"Preserved existing script not in new data: {existing_script.get('Name','')}"
                )

        # Final deduplication by Community URL path to ensure no duplicates slip through
        deduped = {}
        order = []
        for row in merged_data:
            community_url = row.get("Community URL", "")
            url_key = (
                community_url.replace("https://norns.community/", "").strip("/")
                if community_url
                else ""
            )
            if not url_key:
                # Keep rows without URL as-is keyed by name to avoid accidental drop
                url_key = f"__no_url__::{row.get('Name','')}"
            if url_key not in deduped:
                deduped[url_key] = row
                order.append(url_key)
            else:
                # Merge into existing using non-empty preference; Playwright Status pref new non-empty
                base = deduped[url_key]
                for excel_col in self.FIELD_MAP.keys():
                    if excel_col == "Playwright Status":
                        if str(row.get(excel_col, "")).strip():
                            base[excel_col] = row[excel_col]
                        continue
                    if (
                        not str(base.get(excel_col, "")).strip()
                        and str(row.get(excel_col, "")).strip()
                    ):
                        base[excel_col] = row[excel_col]
        merged_data = [deduped[k] for k in order]

        logger.info(
            f"Merge complete: {added_count} added, {updated_count} updated, {preserved_count} preserved"
        )

        # Store summary stats for final report
        self.summary_stats = {
            "scripts_added": added_count,
            "scripts_updated": updated_count,
            "scripts_preserved": preserved_count,
            "total_scripts": len(merged_data),
            "added_details": added_details,
            "updated_details": updated_details,
        }

        return merged_data

    def parse_statuses_from_log(self, log_path):
        """Parse Playwright Status assignments from a run log.

        Returns a dict: script name -> status.
        Heuristics:
          - Explicit lines for Missing Demo and Manual Override are authoritative.
          - "Demo URL conflict for <name>: ... vs ..." implies Extract Preferred.
          - For remaining "Discovered demo for script <name>: <url>" without a preceding conflict,
            mark as Playwright Preferred when the URL looks like an embed-only provider
            (player.vimeo.com, w.soundcloud.com, youtube.com/embed), otherwise No Conflict.
        """
        status_by_name = {}
        try:
            with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()
        except Exception as e:
            logger.error(f"Failed to read log file for status parsing: {e}")
            return status_by_name

        import re as _re

        # Precompile regexes
        rx_missing = _re.compile(
            r"Updated existing script data for (.+?) with Missing Demo status"
        )
        rx_manual = _re.compile(
            r"Updated existing script data for (.+?) with Manual Override status"
        )
        rx_conflict = _re.compile(r"Demo URL conflict for (.+?): ")
        rx_discovered = _re.compile(r"Discovered demo for script (.+?):\s*(\S+)")

        # First pass: direct status lines
        for line in lines:
            m = rx_missing.search(line)
            if m:
                status_by_name[m.group(1)] = "Missing Demo"
                continue
            m = rx_manual.search(line)
            if m:
                status_by_name[m.group(1)] = "Manual Override"
                continue
            m = rx_conflict.search(line)
            if m:
                # Conflict defaults to Extract Preferred in code until user resolution
                name = m.group(1)
                # Don't overwrite explicit statuses already parsed later
                status_by_name.setdefault(name, "Extract Preferred")

        # Second pass: discovered demos without explicit conflict/manual/missing
        # Build a set of names that already have explicit statuses
        already_set = set(status_by_name.keys())
        for line in lines:
            m = rx_discovered.search(line)
            if not m:
                continue
            name = m.group(1)
            url = m.group(2)
            if name in already_set:
                continue

            lower_url = url.lower()
            if (
                "player.vimeo.com/" in lower_url
                or "w.soundcloud.com/" in lower_url
                or "youtube.com/embed/" in lower_url
            ):
                status_by_name[name] = "Playwright Preferred"
            else:
                # Could be No Conflict or Extract Preferred; favor No Conflict when not identifiable
                status_by_name[name] = "No Conflict"

        logger.info(
            f"Parsed {len(status_by_name)} Playwright Status assignments from log"
        )
        return status_by_name

    def apply_status_updates_from_log(self, log_path, excel_path="norns_scripts.xlsx"):
        """Apply Playwright Status updates parsed from a log file directly to the Excel sheet."""
        status_map = self.parse_statuses_from_log(log_path)
        if not status_map:
            logger.warning("No statuses parsed from log; nothing to apply")
            return 0

        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found: {excel_path}")
                return 0
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel file for applying statuses: {e}")
            return 0

        if "Name" not in df.columns:
            logger.error("Excel file missing required 'Name' column")
            return 0

        # Ensure Playwright Status column exists
        if "Playwright Status" not in df.columns:
            df["Playwright Status"] = ""

        # Ensure compatible dtype to avoid pandas warning when assigning strings
        try:
            df["Playwright Status"] = (
                df["Playwright Status"]
                .astype("object")
                .where(pd.notna(df["Playwright Status"]), "")
            )
        except Exception:
            # Fallback: leave as-is if conversion fails
            pass

        # Build a normalization helper for better matching tolerance
        import re as _re2
        import unicodedata as _unicodedata

        def _normalize_name(name: str) -> str:
            if not isinstance(name, str):
                name = "" if pd.isna(name) else str(name)
            # Normalize unicode (convert curly quotes to ASCII equivalents)
            name = (
                name.replace("’", "'")
                .replace("‘", "'")
                .replace("“", '"')
                .replace("”", '"')
            )
            name = _unicodedata.normalize("NFKD", name)
            # Lower, remove diacritics by encoding to ASCII
            try:
                name = name.encode("ascii", "ignore").decode("ascii")
            except Exception:
                pass
            name = name.lower()
            # Collapse punctuation and whitespace to single hyphens/spaces for robust matching
            name = _re2.sub(r"[^a-z0-9\s_\-]", "", name)
            name = _re2.sub(r"[\s\-]+", "-", name).strip("-")
            return name

        # Build fast lookup maps for exact and normalized names
        exact_to_index = {}
        normalized_to_index = {}
        for idx, row in df.iterrows():
            nm = row.get("Name", "")
            exact_to_index[str(nm)] = idx
            normalized_to_index[_normalize_name(str(nm))] = idx

        # Apply updates
        updated_rows = 0
        for raw_name, new_status in status_map.items():
            target_idx = None
            # Prefer exact match first
            if raw_name in exact_to_index:
                target_idx = exact_to_index[raw_name]
            else:
                # Try normalized name match
                norm = _normalize_name(raw_name)
                if norm in normalized_to_index:
                    target_idx = normalized_to_index[norm]

            if target_idx is None:
                continue

            current_value = df.at[target_idx, "Playwright Status"]
            if str(current_value) != str(new_status):
                df.at[target_idx, "Playwright Status"] = new_status
                updated_rows += 1

        if updated_rows == 0:
            logger.info("No Playwright Status cells changed based on log parsing")
            return 0

        # Save back using pandas to minimize side-effects
        try:
            df.to_excel(excel_path, index=False, engine="openpyxl")
            logger.info(
                f"Applied {updated_rows} Playwright Status update(s) from log to {excel_path}"
            )
        except Exception as e:
            logger.error(f"Failed saving Excel after applying statuses: {e}")
            return 0

        return updated_rows

    def _write_formatted_excel_from_df(self, df: pd.DataFrame, filename: str):
        """Write the given DataFrame to Excel with the project's formatting and table styling."""
        # Ensure all required columns exist
        required_columns = list(self.FIELD_MAP.keys())
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""

        # Sort by Name
        df = df.sort_values(by="Name", na_position="last").reset_index(drop=True)

        # Create formatted workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Norns Scripts"

        # Headers
        headers = list(self.FIELD_MAP.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=14)

        # Rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            # Name
            cell = ws.cell(row=row_idx, column=1, value=row["Name"])
            cell.font = Font(size=14)

            # Author
            cell = ws.cell(row=row_idx, column=2, value=row["Author"])
            cell.font = Font(size=14)

            # Tags
            cell = ws.cell(row=row_idx, column=3, value=row["Tags"])
            cell.font = Font(size=14)

            # Description
            cell = ws.cell(row=row_idx, column=4, value=row["Description"])
            cell.font = Font(size=14)

            # Demo (multi-URL aware)
            self._write_demo_cell(ws, row_idx, row.get("Demo"))

            # Discussion URL
            discussion_value = (
                row["Discussion URL"] if pd.notna(row["Discussion URL"]) else ""
            )
            if discussion_value and str(discussion_value).strip():
                cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                cell.hyperlink = str(discussion_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                cell.font = Font(size=14)

            # Project URL
            project_value = row["Project URL"] if pd.notna(row["Project URL"]) else ""
            if project_value and str(project_value).strip():
                cell = ws.cell(row=row_idx, column=7, value=project_value)
                cell.hyperlink = str(project_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=7, value=project_value)
                cell.font = Font(size=14)

            # Documentation URL
            doc_value = (
                row["Documentation URL"]
                if pd.notna(row.get("Documentation URL"))
                else ""
            )
            if doc_value and str(doc_value).strip():
                cell = ws.cell(row=row_idx, column=8, value=doc_value)
                cell.hyperlink = str(doc_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=8, value=doc_value)
                cell.font = Font(size=14)

            # Community URL
            community_value = (
                row["Community URL"] if pd.notna(row["Community URL"]) else ""
            )
            if community_value and str(community_value).strip():
                cell = ws.cell(row=row_idx, column=9, value=community_value)
                cell.hyperlink = str(community_value)
                cell.font = Font(size=14, color="0000FF", underline="single")
            else:
                cell = ws.cell(row=row_idx, column=9, value=community_value)
                cell.font = Font(size=14)

            # Playwright Status
            playwright_status = (
                row["Playwright Status"] if pd.notna(row["Playwright Status"]) else ""
            )
            cell = ws.cell(row=row_idx, column=10, value=playwright_status)
            cell.font = Font(size=14)

            # Last Updated
            last_updated_value = (
                row["Last Updated"] if pd.notna(row.get("Last Updated")) else ""
            )
            cell = ws.cell(row=row_idx, column=11, value=last_updated_value)
            cell.font = Font(size=14)

            # Out of Sync
            out_of_sync_value = (
                row["Out of Sync"] if pd.notna(row.get("Out of Sync")) else ""
            )
            cell = ws.cell(row=row_idx, column=12, value=out_of_sync_value)
            cell.font = Font(size=14)

        # Auto sizing
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        content_length = len(str(cell.value))
                        adjusted_length = int(content_length * 1.2)
                        if adjusted_length > max_length:
                            max_length = adjusted_length
                except:
                    pass
            adjusted_width = max(min(max_length + 2, 80), 10)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Table
        last_row = len(df) + 1
        table_range = f"A1:L{last_row}"
        table = Table(displayName="NornsScripts", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleDark11",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Save
        wb.save(filename)

    def save_to_excel(self, filename="norns_scripts.xlsx"):
        """Save scraped data to Excel file with clickable hyperlinks, merging with existing data"""
        if not self.script_data:
            logger.error("No data to save")
            return

        logger.info(f"Processing {len(self.script_data)} scraped scripts")

        # Convert tags list to comma-separated string for Excel
        for script in self.script_data:
            script["tags"] = ", ".join(script["tags"])

        # Load existing data + drift snapshots (sidecar JSON) and merge.
        # Snapshots are mutated in place by merge_data; we save them after the
        # xlsx is successfully written below.
        existing_df = self.load_existing_data(filename)
        snapshots = self._load_snapshots(filename)
        merged_data = self.merge_data(self.script_data, existing_df, snapshots)

        # Compute GitHub-based Last Updated for all rows (based on Project URL).
        # `filename` enables the committed per-repo Last-Updated cache so nightly
        # runs only re-page commits for repos that were actually pushed.
        try:
            merged_data = self._apply_last_updated(merged_data, filename)
        except Exception as e:
            logger.warning(f"Failed applying Last Updated enrichment: {e}")

        if not merged_data:
            logger.error("No data to save after merging")
            return

        # Create DataFrame from merged data
        df = pd.DataFrame(merged_data)

        # Ensure all required columns exist
        required_columns = list(self.FIELD_MAP.keys())

        for col in required_columns:
            if col not in df.columns:
                df[col] = ""

        # Sort DataFrame by Name column alphabetically
        df = df.sort_values(by="Name", na_position="last").reset_index(drop=True)

        # Create Excel workbook with hyperlinks
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Norns Scripts"

            # Add headers
            headers = list(self.FIELD_MAP.keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=14)

            # Add data with hyperlinks
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                # Name
                cell = ws.cell(row=row_idx, column=1, value=row["Name"])
                cell.font = Font(size=14)

                # Author
                cell = ws.cell(row=row_idx, column=2, value=row["Author"])
                cell.font = Font(size=14)

                # Tags
                cell = ws.cell(row=row_idx, column=3, value=row["Tags"])
                cell.font = Font(size=14)

                # Description
                cell = ws.cell(row=row_idx, column=4, value=row["Description"])
                cell.font = Font(size=14)

                # Demo (multi-URL aware: newline-separated cell, first hyperlinked)
                self._write_demo_cell(ws, row_idx, row.get("Demo"))

                # Discussion URL (as hyperlink)
                discussion_value = (
                    row["Discussion URL"] if pd.notna(row["Discussion URL"]) else ""
                )
                if discussion_value and str(discussion_value).strip():
                    cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                    cell.hyperlink = str(discussion_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=6, value=discussion_value)
                    cell.font = Font(size=14)

                # Project URL (as hyperlink)
                project_value = (
                    row["Project URL"] if pd.notna(row["Project URL"]) else ""
                )
                if project_value and str(project_value).strip():
                    cell = ws.cell(row=row_idx, column=7, value=project_value)
                    cell.hyperlink = str(project_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=7, value=project_value)
                    cell.font = Font(size=14)

                # Documentation URL (as hyperlink)
                doc_value = (
                    row["Documentation URL"]
                    if pd.notna(row.get("Documentation URL"))
                    else ""
                )
                if doc_value and str(doc_value).strip():
                    cell = ws.cell(row=row_idx, column=8, value=doc_value)
                    cell.hyperlink = str(doc_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=8, value=doc_value)
                    cell.font = Font(size=14)

                # Community URL (as hyperlink)
                community_value = (
                    row["Community URL"] if pd.notna(row["Community URL"]) else ""
                )
                if community_value and str(community_value).strip():
                    cell = ws.cell(row=row_idx, column=9, value=community_value)
                    cell.hyperlink = str(community_value)
                    cell.font = Font(size=14, color="0000FF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=9, value=community_value)
                    cell.font = Font(size=14)

                # Playwright Status (text only)
                playwright_status = (
                    row["Playwright Status"]
                    if pd.notna(row["Playwright Status"])
                    else ""
                )
                cell = ws.cell(row=row_idx, column=10, value=playwright_status)
                cell.font = Font(size=14)

                # Last Updated (text only, YYYY-MM-DD)
                last_updated_value = (
                    row["Last Updated"] if pd.notna(row.get("Last Updated")) else ""
                )
                cell = ws.cell(row=row_idx, column=11, value=last_updated_value)
                cell.font = Font(size=14)

                # Out of Sync (text only)
                out_of_sync_value = (
                    row["Out of Sync"] if pd.notna(row.get("Out of Sync")) else ""
                )
                cell = ws.cell(row=row_idx, column=12, value=out_of_sync_value)
                cell.font = Font(size=14)

            # Auto-adjust column widths to fit all content
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            # Calculate length considering font size and content
                            content_length = len(str(cell.value))
                            # Adjust for font size 14 (roughly 1.2x multiplier)
                            adjusted_length = int(content_length * 1.2)
                            if adjusted_length > max_length:
                                max_length = adjusted_length
                    except:
                        pass
                # Set width with minimum and maximum bounds
                adjusted_width = max(min(max_length + 2, 80), 10)  # Min 10, Max 80
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create table with Orange Table Style Dark 11
            last_row = len(df) + 1
            table_range = f"A1:L{last_row}"  # 12 columns through Out of Sync
            table = Table(displayName="NornsScripts", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleDark11",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # Freeze the top row
            ws.freeze_panes = "A2"

            # Save the workbook
            wb.save(filename)
            logger.info(f"Successfully saved to {filename} with clickable hyperlinks")

        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
            # Fallback to pandas method if openpyxl fails
            try:
                df.to_excel(filename, index=False, engine="openpyxl")
                logger.info(f"Fallback save successful: {filename}")
            except Exception as e2:
                logger.error(f"Fallback save also failed: {e2}")

        # Persist drift snapshots (sidecar JSON) — mutated in place by merge_data
        self._save_snapshots(filename, snapshots)

        # Persist Discourse URL resolutions (sidecar JSON) so future runs reuse
        # them without re-resolving stale URLs from upstream community.json.
        if self._discourse_resolutions:
            self._save_discourse_resolutions(filename, self._discourse_resolutions)

        # Persist external-search cache so next run skips successful lookups.
        if hasattr(self, "_external_search_cache") and self._external_search_cache:
            self._save_external_searches(filename, self._external_search_cache)

        # Emit feed.json for ingenue (B7). Best-effort; runs after the xlsx is
        # safely written. GitHub enrichment is cached per-repo so this is cheap
        # on nights where few repos changed. Disable with --no-feed.
        if getattr(self, "feed_enabled", True):
            self.write_feed_json(merged_data, filename, getattr(self, "feed_output", None))

        # Emit catalog.json — the canonical full-rows catalog that build_data.py
        # prefers over the xlsx. Best-effort; the xlsx is already safely written.
        # With discovery enabled, fold in net-new GitHub repos (source='github');
        # the xlsx stays community-only (this only widens catalog.json + the site).
        if getattr(self, "catalog_enabled", True):
            discovered = (
                self._run_discovery(merged_data, filename)
                if getattr(self, "discover_enabled", False)
                else None
            )
            self.write_catalog_json(
                merged_data, filename, getattr(self, "catalog_output", None), discovered
            )

    def print_summary(self):
        """Print a summary of what was accomplished during scraping"""
        print("\n" + "=" * 60)
        print("SCRAPING SUMMARY")
        print("=" * 60)
        print(f"Scripts scraped from norns.community: {len(self.script_data)}")

        if hasattr(self, "summary_stats") and self.summary_stats:
            print(f"Scripts added to Excel: {self.summary_stats['scripts_added']}")
            print(f"Scripts updated in Excel: {self.summary_stats['scripts_updated']}")
            print(
                f"Scripts preserved (manual edits): {self.summary_stats['scripts_preserved']}"
            )
            print(f"Total scripts in Excel file: {self.summary_stats['total_scripts']}")

            # Show detailed information about added scripts
            if self.summary_stats.get("added_details"):
                print("\nAdded scripts:")
                for detail in self.summary_stats["added_details"]:
                    fields_str = (
                        ", ".join(detail["fields"]) if detail["fields"] else "no fields"
                    )
                    print(
                        f"  Added script '{detail['name']}' with field(s): {fields_str}"
                    )

            # Show detailed information about updated scripts
            if self.summary_stats.get("updated_details"):
                print("\nUpdated scripts:")
                for detail in self.summary_stats["updated_details"]:
                    fields_str = (
                        ", ".join(detail["fields"]) if detail["fields"] else "no fields"
                    )
                    print(
                        f"  Updated script '{detail['name']}' with field(s): {fields_str}"
                    )

        # Count discovered demos
        demos_found = sum(
            1 for script in self.script_data if script.get("demo", "").strip()
        )
        print(f"\nDemo discovery: Found {demos_found} demo videos")

        # Surface stale Discourse URLs that we transparently recovered from.
        # Each one is a candidate for PRing back to community.json upstream.
        if self._discourse_resolution_log:
            print(
                f"\nDiscourse URL resolutions this run: "
                f"{len(self._discourse_resolution_log)}"
            )
            print(
                "(community.json's discussion_url is stale for these — "
                "consider PRing fixes)"
            )
            for entry in self._discourse_resolution_log:
                print(f"  {entry['stale']}")
                print(f"    -> {entry['resolved']}")
            print(
                "  Source: https://github.com/monome/norns-community/blob/main/community.json"
            )

        # Surface schema deviations in community.json so they can be PR'd upstream.
        # We still ingest typo'd values defensively, so this is informational only.
        if self.schema_deviations:
            # Group by entry for readability
            by_entry = {}
            for d in self.schema_deviations:
                by_entry.setdefault(d["entry"], []).append(d)

            print(
                f"\ncommunity.json schema deviations: {len(self.schema_deviations)} "
                f"across {len(by_entry)} entries"
            )
            print("(values are still ingested via typo aliases; consider PRing upstream)")
            for entry_name in sorted(by_entry.keys()):
                print(f"  {entry_name}")
                for d in by_entry[entry_name]:
                    if d["kind"] == "typo" and d["suggestion"]:
                        print(f"    typo: '{d['field']}' -> likely '{d['suggestion']}'")
                    else:
                        print(
                            f"    unknown field: '{d['field']}' (no standard equivalent)"
                        )
            print(
                "  Source: https://github.com/monome/norns-community/blob/main/community.json"
            )

        print("=" * 60)


def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="Scrape norns.community scripts")
    parser.add_argument(
        "--workers",
        type=int,
        default=10,
        help="Number of parallel workers (default: 10)",
    )
    parser.add_argument(
        "--demo-delay",
        type=float,
        default=0.5,
        help="Delay in seconds between demo discovery requests (default: 0.5)",
    )
    parser.add_argument(
        "--test",
        type=str,
        help="Test mode: process only the specified Community URL (e.g., https://norns.community/scriptname)",
    )
    parser.add_argument(
        "--status-log",
        type=str,
        help="Apply Playwright Status updates by parsing the given run log file and updating the Excel sheet",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="norns_scripts_discourse.xlsx",
        help="Target Excel file to read/write "
        "(default: norns_scripts_discourse.xlsx — distinct from the original "
        "scraper's default for A/B comparison)",
    )
    parser.add_argument(
        "--dedupe",
        action="store_true",
        help="Only deduplicate the target Excel file by Community URL path and exit",
    )
    parser.add_argument(
        "--sync-check",
        action="store_true",
        help="Only compute and update 'Out of Sync' for all rows without skipping logic",
    )
    parser.add_argument(
        "--reset-discoverable-demos",
        action="store_true",
        help=(
            "Clear Demo + Playwright Status for every row that ISN'T marked "
            "'Manual Override'. Also deletes the external_searches sidecar. "
            "Use this to start fresh after bad search results have been "
            "accumulated. Manual edits are preserved."
        ),
    )
    parser.add_argument(
        "--no-feed",
        action="store_true",
        help="Skip emitting feed.json (the ingenue static data feed) on a full run",
    )
    parser.add_argument(
        "--feed-only",
        action="store_true",
        help=(
            "Rebuild feed.json from the existing Excel file + enrichment cache "
            "without re-scraping. Only repos changed since the cache was built "
            "re-hit GitHub. Useful for fast iteration on the feed."
        ),
    )
    parser.add_argument(
        "--feed-output",
        type=str,
        default=None,
        help=(
            "Path to write feed.json (default: 'feed.json' next to the Excel "
            "file). Point this at ../ingenue/web/feed.json to write in place."
        ),
    )
    parser.add_argument(
        "--no-catalog",
        action="store_true",
        help="Skip emitting catalog.json (the canonical full-rows catalog) on a full run",
    )
    parser.add_argument(
        "--catalog-only",
        action="store_true",
        help=(
            "Rebuild catalog.json from the existing Excel file without re-scraping. "
            "Use this to seed/refresh the canonical catalog from the spreadsheet."
        ),
    )
    parser.add_argument(
        "--catalog-output",
        type=str,
        default=None,
        help="Path to write catalog.json (default: 'catalog.json' next to the Excel file).",
    )
    parser.add_argument(
        "--discover",
        action="store_true",
        help=(
            "Enable GitHub discovery: union search strategies, gate net-new repos on "
            "the norns fingerprint, and fold them into catalog.json as source='github' "
            "(the xlsx stays community-only). Off by default."
        ),
    )
    parser.add_argument(
        "--discover-only",
        action="store_true",
        help="Rebuild catalog.json from the existing xlsx WITH discovery, no re-scrape.",
    )
    parser.add_argument(
        "--discover-max-authors",
        type=int,
        default=None,
        help="Cap the per-author discovery sweep to the first N authors (default: all).",
    )

    args = parser.parse_args()

    # You can adjust max_workers based on your system and network
    # More workers = faster scraping, but be respectful to the server
    scraper = NornsScraper(
        max_workers=args.workers,
        demo_delay=args.demo_delay,
        excel_path=args.excel,
    )
    scraper.feed_enabled = not args.no_feed
    scraper.feed_output = args.feed_output
    scraper.catalog_enabled = not args.no_catalog
    scraper.catalog_output = args.catalog_output
    scraper.discover_enabled = args.discover
    scraper.discover_max_authors = args.discover_max_authors

    # Optional fast-path: rebuild feed.json from existing xlsx + cache and exit
    if args.feed_only:
        scraper.regenerate_feed_only(args.excel, args.feed_output)
        return

    # Optional fast-path: rebuild catalog.json from existing xlsx and exit
    if args.catalog_only or args.discover_only:
        scraper.regenerate_catalog_only(
            args.excel, args.catalog_output, discover=args.discover_only
        )
        return

    # Optional fast-path: apply statuses from a log and exit
    if args.status_log:
        updated = scraper.apply_status_updates_from_log(args.status_log, args.excel)
        if updated:
            logger.info(
                f"Status application complete. {updated} row(s) updated in {args.excel}"
            )
        else:
            logger.info("No rows updated from status log application")
        return

    # Optional fast-path: clear auto-discovered demos and exit
    if args.reset_discoverable_demos:
        try:
            n = scraper.reset_discoverable_demos(args.excel)
            logger.info(
                f"Reset complete. Cleared Demo + Status for {n} row(s) in {args.excel}. "
                f"Manual Override rows preserved. Run the scraper to re-discover."
            )
        except Exception as e:
            logger.error(f"Reset failed: {e}")
        return

    # Optional fast-path: compute Out of Sync only and exit
    if args.sync_check:
        try:
            updated = scraper.sync_check_only(args.excel)
            if updated:
                logger.info(
                    f"Sync-check complete. Updated 'Out of Sync' for {updated} row(s) in {args.excel}"
                )
            else:
                logger.info("Sync-check completed; no changes were necessary")
        except Exception as e:
            logger.error(f"Sync-check failed: {e}")
        return

    # Optional fast-path: deduplicate an existing Excel file and exit
    if args.dedupe:
        try:
            if not os.path.exists(args.excel):
                logger.error(f"Excel file not found: {args.excel}")
                return
            df = pd.read_excel(args.excel)
        except Exception as e:
            logger.error(f"Failed to load Excel for dedupe: {e}")
            return

        before = len(df)

        # Derive URL path key
        def url_key(url: str) -> str:
            if not isinstance(url, str):
                return ""
            return url.replace("https://norns.community/", "").strip("/")

        df["__key__"] = df.get("Community URL", "").apply(url_key)
        # Keep first by URL key; for empties, de-dupe by Name
        non_empty = df[df["__key__"] != ""].copy()
        empty = df[df["__key__"] == ""].copy()
        non_empty = non_empty.drop_duplicates(subset=["__key__"], keep="first")
        empty = empty.drop_duplicates(subset=["Name"], keep="first")
        out = pd.concat([non_empty, empty], ignore_index=True).drop(
            columns=["__key__"], errors="ignore"
        )
        after = len(out)
        removed = before - after
        try:
            out.to_excel(args.excel, index=False, engine="openpyxl")
            logger.info(
                f"Deduplicated {removed} row(s) by Community URL path; saved to {args.excel}"
            )
        except Exception as e:
            logger.error(f"Failed saving Excel after dedupe: {e}")
        return

    if args.test:
        logger.info(f"Test mode: Processing single script from {args.test}")
        scraper.test_single_script(args.test)
    else:
        logger.info(
            f"Starting norns.community scraper with parallel processing, demo discovery, and Playwright..."
        )
        scraper.scrape_all_scripts()

    if scraper.script_data:
        scraper.save_to_excel(args.excel)
        logger.info(f"Scraping complete! Found {len(scraper.script_data)} scripts.")
        scraper.print_summary()
    else:
        logger.error("No scripts were scraped successfully")


if __name__ == "__main__":
    main()
