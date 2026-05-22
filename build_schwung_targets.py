#!/usr/bin/env python3
"""
Generate norns_to_schwung.html — a triage + adaptation-design report for
porting norns scripts to Schwung (Ableton Move).

Data sources:
  - norns_scripts_discourse.xlsx  (name, author, tags, description, urls)
  - local clones at ~/gits/norns-dev-lab/library/clones (source-inspected for FEATURED)
  - Schwung module-catalog.json (80 modules) for de-duplication

Output: norns_to_schwung.html  (single self-contained file, no deps)

Methodology note baked into the page:
  - The FEATURED ~30 are source-verified (engine/softcut/crow/grid inspected).
  - The full 344-row triage is curated + heuristic (tags/descriptions + dedupe
    against Schwung's catalog), NOT source-verified per row.
"""
import openpyxl, json, html, os

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "norns_scripts_discourse.xlsx")
OUT  = os.path.join(ROOT, "norns_to_schwung.html")

# ---------------------------------------------------------------------------
# Load sheet
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb["Norns Scripts"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
SCRIPTS = []
for r in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr, r))
    SCRIPTS.append({
        "name": str(d.get("Name") or "").strip(),
        "author": str(d.get("Author") or "").strip(),
        "tags": [t.strip().lower() for t in str(d.get("Tags") or "").replace(";", ",").split(",") if t.strip()],
        "desc": str(d.get("Description") or "").strip(),
        "url": str(d.get("Project URL") or d.get("Discussion URL") or "").strip(),
    })
BYNAME = {s["name"].lower(): s for s in SCRIPTS}

def sheet(name):
    return BYNAME.get(name.lower(), {"name": name, "author": "", "tags": [], "desc": "", "url": ""})

# ---------------------------------------------------------------------------
# Buckets / classes
# ---------------------------------------------------------------------------
# verdict buckets
PORT="port"; ADAPT="adapt"; NICHE="niche"; DUP="dup"; SKIP="skip"
BUCKET_META = {
    PORT:  ("Port",            "🟢", "Strong target, clear path to a faithful Schwung module."),
    ADAPT: ("Adapt",           "🟡", "Viable but needs reimagining (grid→tool/overtake, crow→MIDI-FX)."),
    NICHE: ("Niche / Maybe",   "🔵", "Works, but lower value or overlaps an existing Schwung module."),
    DUP:   ("Skip — duplicate","⚪", "Schwung (or a community module) already covers this well."),
    SKIP:  ("Skip — impractical","⚫","Needs physical CV/HW with nothing interesting left, or is norns-internal/visual-only."),
}
# portability classes
P_LUA   = ("lua",     "Lua-only",     "Pure logic — reimplement in QuickJS/C. No DSP risk.")
P_SCS   = ("sc-stock","SC engine",    "Has a SuperCollider engine — port to a native DSP .so (or run scsynth via sc-plugins-arm64).")
P_SOFT  = ("softcut", "Softcut",      "Built on norns softcut buffers — needs a varispeed sample-buffer DSP.")
P_CV    = ("cv",      "CV-reinterp",  "Crow/JF logic kept as MIDI/CC; the physical CV jacks are dropped.")
P_NA    = ("na",      "N/A",          "Norns-internal, hardware-companion, or visual-only — not portable in a meaningful way.")

# build status (work done so far) — keyed into BUILT below
STATUS_META = {
    "ready": ("✅ Built — test now", "st-ready"),
    "built": ("🔨 Built — other branch", "st-built"),
}

FORM_META = {
    "midi_fx":          ("MIDI FX",          "Sits in a chain slot before a sound generator; rewrites or generates MIDI."),
    "sound_generator":  ("Sound Generator",  "A voice/instrument — slot synth, chainable."),
    "audio_fx":         ("Audio FX",          "Processes audio in a chain slot or Master FX."),
    "tool":             ("Tool",              "Standalone workflow that outputs MIDI to Move tracks / Schwung synths."),
    "overtake":         ("Overtake",          "Takes full UI control: 128×64 + 4×8 pads + transport."),
}

# ---------------------------------------------------------------------------
# FEATURED — 30 source-verified deep-dives (6 per form factor)
# Each: id, form, chainable, port(class), effort, additive(html), why_modules,
#       mapping(html), mockup(text), crow(optional), bucket
# ---------------------------------------------------------------------------
FEATURED = [
# ============================ MIDI FX ============================
{
 "name":"changes","form":"midi_fx","chainable":True,"port":P_LUA,"effort":"Low","bucket":PORT,
 "additive":"Schwung has no modulation-source MIDI FX at all. <code>changes</code> is eight phase-linked sine LFOs streamed out as MIDI CC — drop it before any slot synth (or aim it at Move's track macros) and you get hands-free evolving motion.",
 "why":"genera / euclidrum / superarp generate <em>notes</em>; nothing in the catalog generates continuous <em>modulation</em>.",
 "mapping":"8 encoders = the 8 LFO rates (1:1 — this is the dream mapping; norns crammed all 8 into menus). Pads col 1–8 top row = per-LFO on/off; row 2 = bipolar/unipolar. Hold a pad + encoder = set that LFO's destination CC. Display draws all 8 phasors as scrolling sines.",
 "mockup":"""┌─ changes · 8 linked LFOs → CC ─────────┐
│ 1∿ 2∿ 3∿ 4∿ 5∿ 6∿ 7∿ 8∿   tempo-sync │
│ ╱╲    ╱╲   ___   ╱╲   .·.   ╱╲   ___  │
│╱  ╲__╱  ╲ ╱   ╲ ╱  ╲ ·   · ╱  ╲ ╱   ╲ │
│ rate .25  .5  1  2  .33 1.5  1  4 bars│
│ →CC  74  71  1  7  93  10  91  79     │
└────────────────────────────────────────┘
ENC 1-8 : LFO 1-8 rate          (1 knob each!)
PADS r1 : LFO on/off  r2 : uni/bi-polar
HOLD pad+ENC : pick destination CC
KEYS    : Shift = link/unlink phases""",
},
{
 "name":"clockabout","form":"midi_fx","chainable":True,"port":P_LUA,"effort":"Med","bucket":PORT,
 "additive":"A <em>non-linear</em> MIDI clock: swing, accelerando, exponential and custom groove curves applied to the beat itself. Schwung's clock is linear; this reshapes time for everything downstream.",
 "why":"No groove/swing-curve engine exists; superarp has rhythm presets but not a global non-linear clock.",
 "mapping":"E1 = pattern (linear/swing/exp/sine/random), E2 = strength, E3 = period (per-beat / per-bar), E4 = phase. Pads draw the 16th-note grid warped by the active curve — tap a step to nudge it. A live 'beat ball' rolls across the display at the warped rate.",
 "mockup":"""┌─ clockabout · non-linear MIDI clock ───┐
│ curve: SWING 62%      period: 1 bar    │
│ ●            ●        ●            ●    │
│  ·  ·   · ·     ·  ·   · ·    ·  ·  ·   │
│ |....:....|....:....|  warped 16ths    │
│ out: 24ppqn ✓   bpm src: MIDI-clock    │
└────────────────────────────────────────┘
ENC 1 curve  2 strength  3 period  4 phase
PADS 16 = warped step grid (tap to nudge)
KEYS Play=reset phase  Shift=A/B curve""",
},
{
 "name":"nørgård","form":"midi_fx","chainable":True,"port":P_LUA,"effort":"Low","bucket":PORT,
 "additive":"Generates Per Nørgård's 'infinity series' — a self-similar melodic sequence with deep fractal structure. A one-of-a-kind generative note source you can quantize to scale and fire at any slot synth.",
 "why":"genera is mode-based generative; the infinity-series algorithm is a distinct compositional engine not present anywhere.",
 "mapping":"E1 = root, E2 = scale, E3 = sequence length, E4 = step rate, E5 = octave span, E6 = transpose offset into the series. Pads show the next 32 series degrees as a heat-map; tap to jump the read-head. Genuinely additive and tiny to port.",
 "mockup":"""┌─ nørgård · infinity series ────────────┐
│ root C  scale DORIAN   len 32  ÷ 1/16  │
│ 0 1 -1 2 1 0 -2 3  ← series degrees    │
│ ▁▃▁▅▃▁█▂▁▄▆▁▃▇▂▁  read-head ►          │
│ next: E3  → Dorian-quantized           │
└────────────────────────────────────────┘
ENC 1 root 2 scale 3 length 4 rate 5 oct 6 offset
PADS = upcoming degrees (tap = move read-head)
KEYS Play start/stop  Shift = reseed""",
},
{
 "name":"serialism","form":"midi_fx","chainable":True,"port":P_CV,"effort":"Low","bucket":PORT,
 "additive":"A twelve-tone row composer — define a tone row, then run prime / retrograde / inversion / retrograde-inversion transforms in real time. Schoenberg in a chain slot.",
 "why":"No serial/12-tone tool exists; complements (doesn't duplicate) the scale-based generators.",
 "mapping":"Pads top two rows (12 of 16) = enter/edit the row chromatically. E1 = transform (P/R/I/RI), E2 = transposition, E3 = rate, E4 = octave. Display shows the row as a clock-face matrix. Crow note-out becomes MIDI; the physical CV is simply dropped.",
 "crow":"Original offers crow v/oct out — we keep the row logic and emit MIDI instead.",
 "mockup":"""┌─ serialism · 12-tone row ──────────────┐
│ row: 0 11 7 8 3 1 2 10 6 5 4 9         │
│ form: INVERSION   T+3   ÷1/8           │
│   ·11· 0 ·                              │
│  9      1   matrix (P/R/I/RI)          │
│   ·····7···                             │
└────────────────────────────────────────┘
ENC 1 form 2 transpose 3 rate 4 octave
PADS r1-2 = enter row   r3 = transform pick
KEYS Shift = lock row / edit row""",
},
{
 "name":"magpie","form":"midi_fx","chainable":True,"port":P_CV,"effort":"Med","bucket":ADAPT,
 "additive":"Signal routing + a <em>modulated note echo</em>: incoming notes are re-emitted with evolving delay, transposition and probability — a melodic diffuser, not an audio delay.",
 "why":"superarp arpeggiates; this echoes/diffuses an arbitrary incoming melody. Different mechanism, no overlap.",
 "mapping":"E1 = echo time, E2 = feedback (repeat count), E3 = transpose-per-echo, E4 = probability, E5 = humanize. The crow/CV routing half is dropped; the MIDI note-echo matrix is the keeper. Pads = tap a routing cell (source→dest channel).",
 "crow":"crow=8 refs are CV routing — we keep only the MIDI note-echo/transpose engine.",
 "mockup":"""┌─ magpie · modulated note echo ─────────┐
│ in ──► [echo] ──► out   reps 4         │
│ time 3/16  fb 60%  +st +5  prob 80%    │
│ C4 · · E4· · G4 · · (echoes fan out)   │
│  ░▒▓ trailing transposed repeats       │
└────────────────────────────────────────┘
ENC 1 time 2 reps 3 transpose 4 prob 5 humanize
PADS routing grid (src row → dest ch)
KEYS Shift = freeze pattern""",
},
{
 "name":"justharmonicon","form":"midi_fx","chainable":True,"port":P_CV,"effort":"Med","bucket":ADAPT,
 "additive":"A subharmonic sequencer (built for Just Friends) — generates undertone-series chords. Reinterpreted as MIDI, it drives any Schwung synth with rich subharmonic voicings you can't dial up elsewhere.",
 "why":"impressive-chords builds overtone/standard chords; subharmonic (undertone) voicing is a distinct harmonic world.",
 "mapping":"E1 = fundamental, E2–E5 = four subharmonic divisors (the Harmonic-Synth move), E6 = rate. Pads = step on/off + per-step divisor latch. The JF voice allocation maps to MIDI channels / a poly slot synth.",
 "crow":"JF subharmonic voicing → MIDI chords aimed at a poly slot synth (e.g. OB-Xd, Surge).",
 "mockup":"""┌─ justharmonicon · subharmonics ────────┐
│ fund A2   ÷2 ÷3 ÷5 ÷7   rate 1/8       │
│ A1  E1  C#1  ~G0     (undertone chord) │
│ step ■ □ ■ ■ □ ■ □ ■  divisor-latch    │
│ → MIDI poly  ch 1-4                     │
└────────────────────────────────────────┘
ENC 1 fund  2-5 divisors  6 rate
PADS = step on/off (+ hold = latch divisor)
KEYS Shift = glide on/off""",
},
# ======================== SOUND GENERATOR ========================
{
 "name":"benjolis","form":"sound_generator","chainable":True,"port":P_SCS,"effort":"Low","bucket":PORT,
 "additive":"A Benjolin/Blippoo-style chaotic synth — cross-modulating oscillators + a rungler shift-register + state-variable filter. Schwung has no chaos/feedback-noise voice. Tiny engine (366 LOC of SC).",
 "why":"Surge/OB-Xd/Plaits are 'tame' macro/VA voices; the Benjolin's self-patching chaos is a genuinely missing flavour.",
 "mapping":"All 8 encoders map cleanly: osc1/osc2 freq, rungle rate, loop/chaos, filter cutoff, resonance, filter type, output. Pads = preset chaos states + a 'freeze rungler' latch. This is a poster child for the 3-knob→8-encoder expansion.",
 "mockup":"""┌─ benjolis · chaotic / rungler ─────────┐
│ o1▟  o2▙   rungle ▚▞▚  filter SVF      │
│ freq1 ███▒  freq2 ██▒▒  rate ████▒     │
│ cut ██████  res ███▒  loop:CHAOS       │
│ ~~ unstable scope ~~                    │
└────────────────────────────────────────┘
ENC 1 f1 2 f2 3 rungle 4 chaos 5 cut 6 res 7 type 8 out
PADS = chaos snapshots  HOLD = freeze rungler
PORT  : Engine_Benjolis.sc → native DSP .so""",
},
{
 "name":"overtones","form":"sound_generator","chainable":True,"port":P_SCS,"effort":"Med","bucket":PORT,
 "additive":"True additive synthesis — sculpt a tone partial-by-partial. Schwung covers FM, VA, wavetable, sample, physical-model — but not additive. Grid-native partial editing maps perfectly to the 4×8 pads.",
 "why":"Dexed=FM, Surge=wavetable, NuSaw=saw-stack; none let you draw a harmonic spectrum directly.",
 "mapping":"The 4×8 pad grid = the 32 partial amplitudes (a literal spectral drawing surface). E1 = number of partials, E2 = spectral tilt, E3 = inharmonicity, E4 = attack, E5 = release, E6 = spread. Display shows the live spectrum + waveform.",
 "mockup":"""┌─ overtones · additive ─────────────────┐
│ partials 24  tilt -3dB/oct  inharm 0.02│
│ █▆▅▄▃▃▂▂▂▁▁▁▁▁▁▁  (spectrum)           │
│ ∿∿∿  resulting wave                     │
│ env  A▕‾‾‾▏R    spread 12%              │
└────────────────────────────────────────┘
PADS 4×8 = 32 partial amplitudes (draw!)
ENC 1 #part 2 tilt 3 inharm 4 atk 5 rel 6 spread
PORT  : Engine_Overtones.sc → native DSP .so""",
},
{
 "name":"mouse","form":"sound_generator","chainable":False,"port":P_SOFT,"effort":"High","bucket":ADAPT,
 "additive":"Laurie Spiegel's <em>Music Mouse</em> — an algorithmic improvisation instrument where one pointer drives intelligent harmonized voices. A canonical piece of computer-music history with no equivalent on Move.",
 "why":"Nothing in the catalog is a real-time algorithmic-harmony improv instrument; it's its own category.",
 "mapping":"The 4×8 pads become the X/Y 'mouse' field (treat pad position as the pointer; pad pressure = dynamics). E1 = harmonic mode (the famous treatments), E2 = voice count, E3 = pattern, E4 = transposition. Best as a standalone tool/instrument outputting MIDI; softcut voice can be dropped in favour of a slot synth.",
 "crow":"—",
 "mockup":"""┌─ mouse · Music Mouse ──────────────────┐
│ mode CHROMATIC  voices 4  patt ▟▙       │
│  · · · · ◆ · · ·    ◆ = pointer        │
│  · · · ·│· · · ·    cross-hair voices  │
│  treatment: 'lines'   transp +2        │
└────────────────────────────────────────┘
PADS 4×8 = X/Y pointer field (pressure=dyn)
ENC 1 mode 2 voices 3 pattern 4 transpose
NOTE  : softcut voice → route to slot synth""",
},
{
 "name":"krill","form":"sound_generator","chainable":True,"port":P_SCS,"effort":"High","bucket":ADAPT,
 "additive":"A Lorenz strange-attractor sequencer feeding a physical-modelling voice, with a full modulation matrix. The chaos-as-melody + mod-matrix workflow is the prize; the PM engine is a bonus.",
 "why":"Plaits has a modal/PM mode, but krill's attractor sequencer + patchable mod-matrix is a unique generative front-end.",
 "mapping":"E1–E3 = Lorenz σ/ρ/β (the attractor shape), E4 = read rate, E5 = scale-quantize, E6 = PM brightness. Pads = mod-matrix cells (source→dest). Big port (5.8k LOC, crow-heavy) — keep the attractor+matrix, drop crow outs.",
 "crow":"crow=9 refs are CV destinations — keep attractor/matrix, emit to the internal voice + MIDI.",
 "mockup":"""┌─ krill · Lorenz seq + PM voice ────────┐
│  σ 10  ρ 28  β 2.6     ∞-attractor      │
│    ╭─╮      ·°·                          │
│   ( ⟁ )    ·   ·   quantized → notes    │
│ mod-matrix [src→dst]  PM bright 0.6     │
└────────────────────────────────────────┘
ENC 1 σ 2 ρ 3 β 4 rate 5 quant 6 bright
PADS = mod-matrix routing grid
PORT  : Engine_Krill.sc → DSP; large logic port""",
},
{
 "name":"molly_the_poly","form":"sound_generator","chainable":True,"port":P_SCS,"effort":"Med","bucket":ADAPT,
 "additive":"The iconic norns classic poly-synth — and its 'solar-system' randomized patch <em>creator</em> is the additive hook: a generative patch designer Schwung's VA synths don't offer.",
 "why":"OB-Xd/Surge/NuSaw cover VA polysynth tone; molly's auto-patch-generator workflow is the differentiator. Borderline duplicate on sound, additive on workflow.",
 "mapping":"8 encoders = osc/sub mix, shape, cutoff, res, env amount, LFO, drive, verb. Pads top row = 'randomize zone' (osc / filter / env / fx) for the solar-system generator; bottom rows = patch slots. Display = the orbiting patch-creator viz.",
 "mockup":"""┌─ molly the poly · + patch creator ─────┐
│   ☉  ·  ●     · °   solar-system gen    │
│ osc▥ sub▤ shape▒ cut███ res▒  auto      │
│ env ▕‾‾\\__  lfo ∿  drive 2  verb 20%   │
│ [randomize: OSC FILT ENV FX]            │
└────────────────────────────────────────┘
ENC 1-8 = osc/sub · shape · cut · res · env · lfo · drive · verb
PADS r1 = randomize-zone   r2-4 = patch slots
PORT  : Engine_MollyThePoly.sc → DSP .so""",
},
{
 "name":"dronecaster","form":"sound_generator","chainable":True,"port":P_SCS,"effort":"Med","bucket":ADAPT,
 "additive":"Cast a drone, then 'record whatever returns' — a drone synth fused with a capture workflow. The build-a-drone-and-bottle-it loop is novel against Schwung's play-only synths.",
 "why":"Schwung synths play; none pair a drone designer with an integrated capture-the-result step. Synergizes with the Quantized Sampler.",
 "mapping":"8 encoders = up to 8 drone partials/voices (detune, level). Pads = voice on/off + 'cast' (start) / 'return' (capture to sample). Display = the layered drone spectrum. The two SC engines (synth + socket) port to one DSP.",
 "mockup":"""┌─ dronecaster ──────────────────────────┐
│ v1▓ v2▒ v3▓ v4░ v5▒ v6 v7 v8           │
│ detune  -7 +3 0 +12 -5 cents           │
│ ▟▙▟▙▟▙  layered drone                   │
│ ◉ CAST   ⤓ capturing 00:07...          │
└────────────────────────────────────────┘
ENC 1-8 = voice detune/level
PADS r1 voice on/off  CAST / RETURN(capture)
PORT  : 2 SC engines → one native DSP .so""",
},
# =========================== AUDIO FX ===========================
{
 "name":"phyllis","form":"audio_fx","chainable":True,"port":P_SCS,"effort":"Low","bucket":PORT,
 "additive":"A standalone digitally-modelled analog filter (with drive). Schwung has filters buried inside synths but no dedicated, sweepable filter <em>FX</em> block for a chain or Master FX.",
 "why":"spectra is a resonator-bank, usefulity is a utility; neither is a musical resonant LP/HP/BP filter you can drop on any audio.",
 "mapping":"E1 = cutoff, E2 = resonance, E3 = type (LP/HP/BP/notch), E4 = drive, E5 = env-follow amount, E6 = mix. Pads = filter-type select + cutoff 'snap' notes. Tiny engine — a clean first DSP port. Fully chainable + Master-FX friendly.",
 "mockup":"""┌─ phyllis · analog filter ──────────────┐
│ type LP24   cut 1.2kHz  res 0.7        │
│      ╲                                  │
│ ─────╲___  drive 1.8  env→ 30%         │
│ mix 100%   (modeled ladder)            │
└────────────────────────────────────────┘
ENC 1 cut 2 res 3 type 4 drive 5 env 6 mix
PADS = LP/HP/BP/notch + cutoff snap-notes
PORT  : Engine_Phyllis.sc → native DSP .so""",
},
{
 "name":"glaciers","form":"audio_fx","chainable":True,"port":P_SCS,"effort":"Med","bucket":ADAPT,
 "additive":"Extreme sound stretcher + harmoniser — freeze a moment and smear it into a glassy chord. Distinct from Schwung's tape stretch and Clouds granular: this is spectral freeze + harmony.",
 "why":"stretch=Bungee time-stretch, verglas=Clouds granular; glaciers' freeze-and-harmonise spectral pad is a different texture.",
 "mapping":"E1 = stretch/freeze amount, E2 = harmony interval set, E3 = spread, E4 = tone, E5 = feedback, E6 = mix. Pads top rows = harmony chord shape (drop intervals into the freeze). Softcut buffer + SC engine → one capture+process DSP.",
 "mockup":"""┌─ glaciers · stretch + harmonise ───────┐
│ freeze ███████▒  smear 8.0s            │
│ harmony [ -12 -5 0 +7 +12 ]            │
│ ≈≈≈≈≈≈≈≈≈  glassy pad                  │
│ tone ▒▒██  fb 40%  mix 70%             │
└────────────────────────────────────────┘
ENC 1 freeze 2 harmony 3 spread 4 tone 5 fb 6 mix
PADS = harmony interval chord builder
PORT  : Engine_Glacial.sc + softcut → DSP""",
},
{
 "name":"oooooo","form":"audio_fx","chainable":False,"port":P_SOFT,"effort":"High","bucket":PORT,
 "additive":"Six digital tape loops — the beloved varispeed looper. <strong>Schwung ships no looper at all</strong>, so this fills a genuine, much-requested hole. Best as an overtake/tool given it needs transport + 6 loop lanes.",
 "why":"dj is a stem player, slicer/twinsampler are samplers; none is a live multi-track varispeed loop recorder.",
 "mapping":"Pads: 6 columns = 6 loops, rows = rec / play / reverse / clear. E1–E6 = per-loop speed (varispeed = the whole point). Shift+pad = loop length. Transport keys = global play/stop. This is the strongest argument for a softcut-equivalent buffer DSP in Schwung.",
 "crow":"—",
 "mockup":"""┌─ oooooo · 6 tape loops ────────────────┐
│ L1 ▶ L2 ● L3 ◀ L4 · L5 ▶ L6 ·          │
│ spd 1.0 .5  -1  --  2.0 --             │
│ ▓▓▓▓░░ ▓▓░░░░ ▓▓▓▓▓░ len               │
│ rec ● play ▶ rev ◀ clear ✕            │
└────────────────────────────────────────┘
PADS  cols=6 loops  rows= rec/play/rev/clear
ENC 1-6 = per-loop varispeed
PORT  : softcut → varispeed buffer DSP (key dep)""",
},
{
 "name":"massif","form":"audio_fx","chainable":True,"port":P_SCS,"effort":"Low","bucket":NICHE,
 "additive":"An 8-band resonator — pitched resonant bandpass bank that turns any input into a tuned, ringing chord. Overlaps spectra somewhat, but massif's fixed-8-band tuned-chord approach is simpler and pad-playable.",
 "why":"spectra is a pitch-tracking SVF resonator; massif is a fixed tuned bank — adjacent, not identical. Marked Niche for honesty.",
 "mapping":"8 encoders = the 8 band frequencies (1 knob each — perfect fit). Pads = tune all 8 to a chord/scale at once; rows = decay/Q presets. Tiny 272-LOC engine.",
 "mockup":"""┌─ massif · 8-band resonator ────────────┐
│ f: C2 G2 C3 E3 G3 B3 D4 G4   (Cmaj9)   │
│ ║ ║ ║ ║ ║ ║ ║ ║   ringing bands        │
│ Q 0.8   decay 1.4s   mix 60%           │
└────────────────────────────────────────┘
ENC 1-8 = band 1-8 frequency
PADS = snap bank to chord/scale; Q/decay presets
PORT  : Engine_ResonatorBank.sc → DSP .so""",
},
{
 "name":"raft","form":"audio_fx","chainable":True,"port":P_SOFT,"effort":"Med","bucket":ADAPT,
 "additive":"A scene-setting, softcut-based modulated delay <em>plus</em> a noise/ocean generator — an ambient texture box, not a clean delay. The 'add atmosphere to a chain' use-case is unserved.",
 "why":"Schwung's delays (tapedelay etc.) are clean utilities; raft's role is generative ambience + modulated echo, a texture tool.",
 "mapping":"E1 = delay time, E2 = feedback, E3 = modulation depth, E4 = noise/ocean level, E5 = tone, E6 = mix. Pads = scene morph snapshots. Softcut delay + Engine_Ocean noise → one DSP.",
 "mockup":"""┌─ raft · modulated delay + ocean ───────┐
│ ≈≈≈ ocean ≈≈≈   lvl ▒▒██               │
│ delay 3/8  fb 65%  mod ∿ 30%           │
│ echoes ░▒▓ drifting                     │
│ scene: [A] B  C  D   morph             │
└────────────────────────────────────────┘
ENC 1 time 2 fb 3 mod 4 ocean 5 tone 6 mix
PADS = scene snapshots / morph
PORT  : softcut + Engine_Ocean.sc → DSP""",
},
{
 "name":"sway","form":"audio_fx","chainable":True,"port":P_SCS,"effort":"Med","bucket":ADAPT,
 "additive":"Analysis-driven live processing: it listens to the input (amplitude/pitch/onset) and steers its own effects from that analysis. An <em>adaptive</em> FX — none of Schwung's effects self-modulate from the signal.",
 "why":"keydetect analyses but doesn't process; every other FX is statically dialled. Adaptive/auto-reactive processing is new.",
 "mapping":"E1 = sensitivity, E2–E4 = analysis→FX routings (amp→cutoff, onset→grain, pitch→delay), E5 = depth, E6 = mix. Pads = enable/disable each analysis→effect link. Three small SC engines → one DSP.",
 "mockup":"""┌─ sway · analysis-driven FX ────────────┐
│ in ▂▅█▃▁  amp→cutoff ✓  onset→fx ✓     │
│ pitch trk ♪A3   sens 0.6               │
│ links: [amp▸cut][onset▸grain][pit▸dly]│
│ depth ███▒  mix 50%                     │
└────────────────────────────────────────┘
ENC 1 sens 2-4 routings 5 depth 6 mix
PADS = toggle analysis→effect links
PORT  : 3 SC engines → native DSP .so""",
},
# ============================= TOOL =============================
{
 "name":"orca","form":"tool","chainable":False,"port":P_SOFT,"effort":"High","bucket":ADAPT,
 "additive":"Hundred Rabbits' <em>ORCΛ</em> — a 2D esoteric programming language for building procedural sequencers live. An entire generative-composition paradigm; nothing remotely like it on Move.",
 "why":"davebox/eucalypso are conventional sequencers; ORCA is a Turing-complete grid-language. Category of one.",
 "mapping":"The 128×64 shows a viewport of the ORCA grid (chars). The 4×8 pads = a cursor/navigation + operator palette; jog encoder scrolls the cursor, E1 picks operator, E2 edits value. Outputs MIDI to Move tracks/Schwung synths on bang. A flagship overtake-style tool — ambitious but iconic.",
 "crow":"crow refs = extra outs; MIDI out is the primary target on Move.",
 "mockup":"""┌─ orca ─────────────────────────────────┐
│ . . D8 . . . . . . . . . . . . . . . . │
│ . . : 3C. . . aE5. . . . . . . . . . . │
│ . . . . . . . :48 . . . . . . . . . .  │
│ . . . *. . . . . . . . . . ▣cursor . . │
│ 120bpm  ► out:MIDI ch1-4   frame 0148  │
└────────────────────────────────────────┘
JOG = move cursor   ENC 1 operator  2 value
PADS = operator palette / nav   Play = run
PORT  : softcut→sample outs optional; MIDI core""",
},
{
 "name":"takt","form":"tool","chainable":False,"port":P_LUA,"effort":"Med","bucket":ADAPT,
 "additive":"A parameter-locking step sequencer — Elektron-style p-locks where every step can override any parameter. The killer hook: drop the Timber sample engine and p-lock <em>Schwung synth params</em> per step.",
 "why":"davebox has per-clip FX chains; superarp/eucalypso are generative. None offers per-step parameter locks across an arbitrary target's params. Huge workflow.",
 "mapping":"Pads = 16-step lanes (4 tracks × 16 via page). Hold a step + turn any encoder = lock that param for that step. E1–E8 = the 8 locked params (live readout). Display = step grid + the param being locked. Engine becomes 'send MIDI/param to the slot synth'.",
 "crow":"—",
 "mockup":"""┌─ takt · p-lock sequencer ──────────────┐
│ T1  ■ □ ■◆■ □ ■ □ ■ □ ■◆□ ■ □ ■ □      │
│ ◆ = locked step   page 1/1             │
│ HOLD st7 + ENC3 → cutoff = 0.62 ✎      │
│ locks: cut filt pan ... → slot synth   │
└────────────────────────────────────────┘
PADS = 16 steps (× 4 track pages)
HOLD step + ENC = p-lock that parameter
ENC 1-8 = the 8 lockable params (live)""",
},
{
 "name":"yggdrasil","form":"tool","chainable":False,"port":P_LUA,"effort":"High","bucket":ADAPT,
 "additive":"A 'cyberdeck' tracker — classic tracker workflow (pattern rows, effect columns) on norns. <strong>Schwung has no tracker</strong>; this brings a whole composition idiom to Move.",
 "why":"All Schwung sequencers are step/pad based; tracker (vertical pattern + hex effects) is a distinct, deep workflow.",
 "mapping":"128×64 = the pattern editor (rows scroll under the play-head). Pads = note entry (chromatic) + transport; jog scrolls rows; E1 = track, E2 = column (note/inst/fx), E3 = value, E4 = octave. Drop the YggdrasilSynth engine → route rows to slot synths via MIDI.",
 "crow":"crow=9 refs are extra outs; MIDI to slots is primary.",
 "mockup":"""┌─ yggdrasil · tracker ──────────────────┐
│   T1     T2     T3     T4              │
│ 00 C-4 5│--- -│G-3 2│--- -            │
│ 01 --- -│E-4 5│--- -│C-2 4  ◄play     │
│ 02 D#4 7│--- -│G-3 -│--- -            │
│ 03 ▣    │     │     │     ptn 01/08   │
└────────────────────────────────────────┘
JOG = scroll rows  PADS = note entry + transport
ENC 1 track 2 column 3 value 4 octave
PORT  : drop synth engine → MIDI to slots""",
},
{
 "name":"pitfalls","form":"tool","chainable":False,"port":P_LUA,"effort":"Med","bucket":ADAPT,
 "additive":"A microtonal scale explorer / arpeggiator / grid-keyboard — design non-12-TET tunings and play/sequence them. Schwung is firmly 12-TET; this opens xenharmonic territory.",
 "why":"chorddex/impressive-chords are 12-TET; no tuning explorer or microtonal keyboard exists. Outputs MIDI (incl. MPE pitch-bend for true microtonality).",
 "mapping":"4×8 pads = isomorphic microtonal keyboard (configurable layout). E1 = EDO/scale system, E2 = root, E3 = layout offset, E4 = arp pattern, E5 = rate. Display = the scale ruler + cents readout. MPE pitch-bend per note carries the microtones to any synth.",
 "crow":"—",
 "mockup":"""┌─ pitfalls · microtonal ────────────────┐
│ 19-EDO   root C   layout +3/+1          │
│ cents: 0 63 126 189 ... 1137           │
│ ▢▢▢▢▢▢▢▢  isomorphic pads              │
│ arp ▟▙  rate 1/8   out: MPE bend       │
└────────────────────────────────────────┘
PADS 4×8 = isomorphic microtonal keyboard
ENC 1 EDO 2 root 3 layout 4 arp 5 rate
PORT  : pure Lua; MPE pitch-bend out""",
},
{
 "name":"dreamsequence","form":"tool","chainable":False,"port":P_LUA,"effort":"High","bucket":ADAPT,
 "additive":"A chord-progression composer with an arpeggiator and a harmonizer — write a progression, then generate arps/harmonized lines locked to it. A full song-sketching workflow.",
 "why":"impressive-chords makes chords; genera generates notes. dreamsequence is the conductor that ties a <em>progression</em> + arp + harmony together — closer to a mini-DAW arranger.",
 "mapping":"Pads top rows = the chord progression timeline (one chord per pad); bottom rows = arp pattern editor. E1 = key, E2 = chord quality, E3 = arp style, E4 = rate, E5 = harmonizer interval, E6 = swing. Multi-output: chord to one slot, arp to another.",
 "crow":"crow=6 refs are extra outs; multi-channel MIDI to slots is primary.",
 "mockup":"""┌─ dreamsequence · progression ──────────┐
│ key Am   | Am  F   C   G  | ◄ now      │
│ chord timeline ■ □ □ □                  │
│ arp: up-down 1/16  harm +3rd            │
│ → ch1 chords  ch2 arp  ch3 harmony      │
└────────────────────────────────────────┘
PADS r1-2 = progression  r3-4 = arp editor
ENC 1 key 2 quality 3 arp 4 rate 5 harm 6 swing
PORT  : pure Lua; multi-channel MIDI out""",
},
{
 "name":"tmi","form":"tool","chainable":False,"port":P_LUA,"effort":"Med","bucket":NICHE,
 "additive":"'Text Music Interface' — sequence MIDI by typing terse text commands. A keyboard/notation-driven sequencer that pairs naturally with Move's text-entry; niche but unlike anything in the catalog.",
 "why":"All Schwung sequencers are graphical; a text/notation sequencer is a distinct authoring style (and great for precise, copyable patterns).",
 "mapping":"Display = the text buffer + parsed timeline. Uses Schwung's text_entry.mjs for input; pads = quick tokens (note names, durations, rests); E1 = cursor, E2 = transpose, E3 = tempo. Marked Niche — powerful but a narrower audience.",
 "crow":"—",
 "mockup":"""┌─ tmi · text MIDI ──────────────────────┐
│ > c4 e4 g4 . c5/8 r r [a3 c4 e4]       │
│ parsed: ▮▮▮ _ ▮ __ ▮▮▮  loop ⟳         │
│ cursor ▏  transpose +0  bpm 120         │
│ ch 1   quantize 1/16                    │
└────────────────────────────────────────┘
TEXT-ENTRY buffer (text_entry.mjs)
PADS = token shortcuts (notes/durations/rests)
ENC 1 cursor 2 transpose 3 tempo""",
},
# =========================== OVERTAKE ===========================
{
 "name":"mlr","form":"overtake","chainable":False,"port":P_SOFT,"effort":"High","bucket":ADAPT,
 "additive":"The original monome <em>mlr</em> — live sample-cutting where pads scrub a loop in real time, and you can <strong>record the act of cutting</strong> as a pattern. That performance-recording loop is what twinsampler doesn't do.",
 "why":"twinsampler/slicer slice & trigger; mlr's pattern-recorders (capture + overdub your pad performance) are the additive piece.",
 "mapping":"Each pad row = a loop, the 8 columns = scrub positions across it (tap = jump-and-play). Dedicated pads = pattern record/overdub/play. E1 = loop selection, E2 = speed, E3 = pattern bank. Needs the softcut-equivalent buffer DSP (shared with oooooo).",
 "crow":"—",
 "mockup":"""┌─ mlr · live cut + pattern rec ─────────┐
│ row1 ▣□□□□□□□  loop A  spd 1.0          │
│ row2 □□□▣□□□□  loop B  spd 0.5          │
│ row3 □□□□□□□□  (rec armed ●)            │
│ PATTERNS: P1▶ P2● rec  P3·  overdub    │
└────────────────────────────────────────┘
PADS rows=loops  cols=scrub positions
PADS r4 = pattern record / overdub / play
PORT  : softcut buffer DSP (shared w/ oooooo)""",
},
{
 "name":"arcologies","form":"overtake","chainable":False,"port":P_SOFT,"effort":"High","bucket":ADAPT,
 "additive":"An interactive 2D environment where you place 'structures' on a grid that emit/route notes — designing self-playing sound cities. A generative <em>playground</em>, not a step sequencer.",
 "why":"No spatial/automata composition surface exists; the closest (branchage/genera) are still linear generators.",
 "mapping":"4×8 pads = the world map; place/erase structures with pad taps, jog to move the cursor, E1 = structure type, E2 = signal strength, E3 = tempo. JF voices → Schwung slot synths / MIDI channels. Display = the living map + signal pulses.",
 "crow":"JF/crow voices → MIDI to slot synths; the spatial logic is untouched.",
 "mockup":"""┌─ arcologies ───────────────────────────┐
│ · ◇ · · → ● · ·   ◇=structure ●=pulse  │
│ · · · ▣ · · · ·   ▣=cursor             │
│ · ↑ · · · · ◆ ·   signals propagate    │
│ struct: HARP  str 4  120bpm → MIDI     │
└────────────────────────────────────────┘
PADS 4×8 = world map (place/erase)  JOG = cursor
ENC 1 struct-type 2 strength 3 tempo
PORT  : JF voices → slot synths; softcut opt""",
},
{
 "name":"meadowphysics","form":"overtake","chainable":False,"port":P_LUA,"effort":"Med","bucket":PORT,
 "additive":"The monome <em>meadowphysics</em> cascading-counter sequencer — rows count down at independent rates and trigger each other, producing rhizomatic, ever-shifting rhythms. A unique generative mechanism.",
 "why":"eucalypso/euclidrum are euclidean; meadowphysics' inter-triggering counters are a different generative model entirely. Pure-Lua logic.",
 "mapping":"4×8 pads = 8 counter rows × position; tap to set each row's reset point and trigger links. E1 = global speed, E2 = row select, E3 = count, E4 = rule. Display = the counters mid-cascade. Drop crow outs, emit MIDI per row.",
 "crow":"crow=2 refs are trigger outs → become MIDI notes per row.",
 "mockup":"""┌─ meadowphysics ────────────────────────┐
│ r1 ●○○○○○○○ →r3   speed ███▒           │
│ r2 ○○●○○○○○ →r1                        │
│ r3 ○○○○●○○○ →r2   cascading counters   │
│ r4 ●○○○○○○○        → MIDI notes        │
└────────────────────────────────────────┘
PADS = 8 counter rows (set reset + links)
ENC 1 speed 2 row 3 count 4 rule
PORT  : pure Lua; crow trigs → MIDI""",
},
{
 "name":"n.kria","form":"overtake","chainable":False,"port":P_LUA,"effort":"Med","bucket":PORT,
 "additive":"Native <em>Kria</em> — the legendary monome/ansible step sequencer with independent loop lengths per parameter (trigger / note / octave / duration / ratchet). The per-param-track polymetry is its signature and is missing from Schwung.",
 "why":"euclidrum/eucalypso/genera don't offer Kria's independent-loop-length-per-parameter model. n.kria has no Timber dependency — clean MIDI port.",
 "mapping":"4×8 pads = the active parameter page (16 steps × value). Transport/track keys = pick parameter (TRIG/NOTE/OCT/DUR/RPT) and pattern. E1 = loop length of the current param (the magic), E2 = scale, E3 = tempo, E4 = pattern. Display = current page + per-param loop indicators.",
 "crow":"—",
 "mockup":"""┌─ n.kria · param-track sequencer ───────┐
│ page: NOTE   loop-len 11  scale Dorian │
│ step ▁▃▅▂▆▁▇▃▁▄▆ (×11 then wraps)      │
│ TRIG■NOTE◆OCT DUR RPT   ptn A          │
│ trk1 → MIDI ch1   tempo 124            │
└────────────────────────────────────────┘
PADS = current param page (16 steps × value)
KEYS = pick param page + pattern
ENC 1 loop-length 2 scale 3 tempo 4 pattern""",
},
{
 "name":"zellen","form":"overtake","chainable":False,"port":P_CV,"effort":"Low","bucket":PORT,
 "additive":"A Conway's Game-of-Life sequencer — cellular automata evolve a grid; live cells become notes. Watching rhythm emerge from CA rules is a beloved generative idea with no Schwung equivalent.",
 "why":"No automata-based sequencer exists; small, pure-Lua, and visually perfect for the 4×8 pad grid + display.",
 "mapping":"4×8 pads = the CA board (tap to seed/toggle cells). E1 = generation speed, E2 = scale, E3 = play direction/scan, E4 = ruleset. Display = the evolving board + scan line. Crow out → MIDI notes. Lowest-effort iconic overtake.",
 "crow":"crow=6 refs are note/CV outs → MIDI.",
 "mockup":"""┌─ zellen · game of life seq ────────────┐
│ ░█░░██░░  gen 0142  speed 1/8          │
│ █░██░░█░  scan ▼ col-by-col            │
│ ░░█░█░██  live cell → note (Dorian)    │
│ ██░░░█░░  → MIDI ch1                    │
└────────────────────────────────────────┘
PADS 4×8 = CA board (tap to seed)
ENC 1 speed 2 scale 3 scan-dir 4 ruleset
PORT  : pure Lua; crow → MIDI (very low effort)""",
},
{
 "name":"less-concepts","form":"overtake","chainable":False,"port":P_CV,"effort":"Med","bucket":ADAPT,
 "additive":"A 1D cellular-automata sequencer (Wolfram-style elementary rules) for grid + crow/JF/W. The 1D-rule-number approach (different from zellen's 2D Life) gives a distinct family of evolving patterns.",
 "why":"Pairs with — doesn't duplicate — zellen: 1D elementary CA vs 2D Life. Both are additive automata sequencers with different feels.",
 "mapping":"Top pad row = the current generation (cells); successive rows show history scrolling down. E1 = rule number (0–255), E2 = speed, E3 = scale, E4 = seed. Crow/JF/W outs → MIDI to slot synths. Display = the rule's space-time triangle.",
 "crow":"crow/jf/w outs → MIDI; the 1D-CA logic is the keeper.",
 "mockup":"""┌─ less-concepts · 1D CA ────────────────┐
│ rule 110   speed 1/16   scale Phryg.   │
│ ░░░░█░░░░  gen now                      │
│ ░░░██░░░░                               │
│ ░░██░██░░  space-time history ▼         │
│ ░██░░░██░  live cells → MIDI            │
└────────────────────────────────────────┘
PADS r1 = current gen  r2-4 = history
ENC 1 rule# 2 speed 3 scale 4 seed
PORT  : pure Lua; crow/jf/w → MIDI""",
},
]

FEATURED_NAMES = {f["name"].lower() for f in FEATURED}

# ---------------------------------------------------------------------------
# TRIAGE OVERRIDES for all 344 (curated; featured auto-merged below)
# value = (bucket, form_or_None, port_class, short note)
# ---------------------------------------------------------------------------
OV = {}
def ov(names, bucket, form, port, note):
    for n in names.split(","):
        OV[n.strip().lower()] = (bucket, form, port, note)

# --- Skip: norns-internal / libs / mods / visual-only / hw-companion ---
ov("3d,r,u,arcify,passthrough,interpret,thebangs,combos,ack,timber,seaflex,cryptkeeper", SKIP, None, P_NA, "Library / engine / companion script, not a standalone Move module.")
ov("warmreload,nice-tapes,oblique,semiconductor,monomaniac,combiner,midigrid,choukanzu,folio,internorns,norns.online,nornsfetch,foundry,grid-test,hid_demo", SKIP, None, P_NA, "Norns-system mod/utility — no analogue on Move.")
ov("blue,showers,taweeet,ong,avonlea,fiahod,nc01-drone,nc02-rs,nc03-ds,scryingstone,plasma,p8,waver,esper", SKIP, None, P_NA, "Art / visual / system piece — little musical workflow to port.")
ov("crow_talk,caliper,bowering,path-trace,parrot,vcr,window-crowparator,4-big-knobs,cccccccc,clarck,n16o", SKIP, None, P_NA, "Crow/arc/i2c-only utility — no physical CV/I2C on Move.")
ov("randomizer_typhon,automs70,MDPatch,MDScaler,Mnmpatch", NICHE, "tool", P_LUA, "Vendor-specific CC randomizer — could generalise into a Move CC-randomizer tool.")
ov("cybermidi,midididi", SKIP, None, P_NA, "Norns↔norns / mod-specific MIDI utility.")
ov("nb_drumcrow,nb_mxsynths,nb_plyprc,nb_smpKit", SKIP, None, P_NA, "nb-player mod (norns voice-routing framework) — not applicable.")

# --- Duplicates of existing Schwung modules ---
ov("fm7,orgn,orgnwrms,xD1,bitters,o-o-o", DUP, "sound_generator", P_SCS, "FM synth — Schwung already ships Dexed (6-op FM).")
ov("icarus", DUP, "sound_generator", P_SCS, "Supersaw — covered by NuSaw.")
ov("passersby", DUP, "sound_generator", P_SCS, "West-coast mono — covered by denis.")
ov("mx.samples,mx.synths,moln,beacon", DUP, "sound_generator", P_SCS, "Sample/poly player — covered by SFZ/mrsample and the synth shelf.")
ov("mangl,glut,haze,silos,twine,twins,rangl,easygrain,gemini,graintopia,uhf,here-there,nameless-nightmare,granchild", DUP, "audio_fx", P_SOFT, "Granular — covered by granny / verglas (Clouds) / Boris granular.")
ov("greyhole,delayyyyyyyy,carter's delay,blndr,rpls,eels,bounds,larc,delar", DUP, "audio_fx", P_SOFT, "Delay — Schwung ships tapedelay and others.")
ov("pools", DUP, "audio_fx", P_SCS, "Reverb — many in catalog (CloudSeed, MVerb, Dragonfly...).")
ov("tapedeck", DUP, "audio_fx", P_SCS, "Tape emulation — covered by CHOWTape / TapeScam.")
ov("bline,acid-test", DUP, "sound_generator", P_LUA, "303/acid bassline — covered by 303 + tb3po.")
ov("hachi,cyrene,foulplay,kreislauf,impact,rudiments,supertonic,lorenzos-drums,moonraker,smash,drum_room,makebreakbeat,amenbreak,amen,bakeneko,lamination", DUP, "sound_generator", P_SCS, "Drum machine / euclidean drums — heavily covered (weird-dreams, krautdrums, forge, euclidrum, branchage).")
ov("tuner", DUP, "tool", P_LUA, "Tuner — Schwung ships tuner + guitar-tuner.")
ov("pedalboard,manifold", DUP, "audio_fx", P_SCS, "Live multi-FX chain — that's literally Schwung's Signal Chain + performance-fx.")
ov("clipper,sam,step,eterna,rpmate,ortf,sampswap,piwip", DUP, "tool", P_SOFT, "Sample record/slice/playback — covered by slicer / mrsample / waveform-editor / AutoSample.")
ov("noize,noizeop,demoncore,haven", NICHE, "sound_generator", P_SCS, "Noise synth — partial overlap; demoncore's chaos is the interesting bit.")

# --- Loopers (additive: Schwung has NO looper) ---
ov("reels,wrms,otis,cranes,nydl,samsara,samthree,giro,jiffy,lemniscate,downtown,glitchlooper,sweet bees,footfoot,barcode,bounds", ADAPT, "tool", P_SOFT, "Varispeed/async looper — additive (no looper in Schwung); needs softcut-equivalent buffer DSP.")
ov("nmQuadroDubber,nmSmartyPants,nmRain", NICHE, "audio_fx", P_SOFT, "External-audio tape/loop mangler — additive softcut texture.")

# --- Sequencers worth adapting (grid → tool/overtake) ---
ov("kria_midi,faeng,n.kria", PORT, "overtake", P_LUA, "Kria-family param-track sequencer — signature independent loop-lengths.")
ov("metrix,m18s,skylines,hpns,tviburar,sempra", ADAPT, "tool", P_LUA, "Metropolis/M-185/parc stage sequencer — additive stage-based workflow.")
ov("gridstep,isoseq,descartes,fugu,fugarc,colorwheel,torii,strum,plonky,vials,strides,kolor,beets,ekombi,abacus,delar", ADAPT, "tool", P_SOFT, "Grid step/sample sequencer — adapt to 4×8 pads; many need softcut sample voice.")
ov("loom,traffic,dunes,delinquencer,breakthrough,qfwfq,tulpamancer,lamellae,jala,luck,raindrops,scholastic,grurder,KoiBoi2,Hiswing,SuperBrain,marcovaldo,punchcard,drift,rebound,barycenter,spirals,quence,constellations,fall,fibonacci,streams,patchwork,thorns,groovecats,animator,initenere,corners,euclidigons,skylines", ADAPT, "midi_fx", P_LUA, "Generative/algorithmic MIDI sequencer — strong Lua-logic port to a MIDI-FX or tool.")
ov("forge,flora,endless-stairs,krill,gematria,circles,shfts,ribbons,tambla,nmMelodyMagic,nørgård,loudnumbers_norns,get in the sea!,paranibbles,serialism,nisp,yggdrasil,zxcvbn,paracosms", ADAPT, "tool", P_LUA, "Generative/notation sequencer or tracker — additive composition workflow.")
ov("dreamsequence,mosaic", ADAPT, "tool", P_LUA, "Chord-progression / harmony composer — mini-arranger workflow.")
ov("seeker_ii", NICHE, "tool", P_SOFT, "Deep compositional interface — very large (27k LOC); ambitious long-term port.")
ov("cheat_codes_2,mlre,thirtythree,repl-looper,pitter-patter", NICHE, "tool", P_SOFT, "Big sample-playground — overlaps slicer/dj; pick specific sub-workflows.")
ov("orca", ADAPT, "tool", P_SOFT, "ORCA esoteric grid-language sequencer — iconic, ambitious.")
ov("takt", ADAPT, "tool", P_LUA, "Parameter-locking step sequencer — p-lock Schwung synth params.")
ov("zellen,less-concepts,meadowphysics", PORT, "overtake", P_LUA, "Cellular-automata / cascading-counter sequencer — additive generative grid.")
ov("arcologies,mouse,buoys", ADAPT, "overtake", P_SOFT, "Spatial/algorithmic instrument environment — full UI takeover.")

# --- MIDI FX / modulation / clock (mostly additive) ---
ov("changes,shapes,wobblewobble,lorenz,ufo", PORT, "midi_fx", P_LUA, "LFO / modulation source → MIDI CC — additive (no modulation MIDI-FX exists).")
ov("clockabout", PORT, "midi_fx", P_LUA, "Non-linear / groove MIDI clock — additive.")
ov("magpie,monitor,onehanded,intervaltrainer,tmi", ADAPT, "midi_fx", P_LUA, "MIDI utility / note-echo / trainer — small, additive helpers.")
ov("justharmonicon,just-play,iiitoii,tidbit,acrostic,constellations", ADAPT, "midi_fx", P_CV, "JF/crow-driven harmony or sequence — keep logic, emit MIDI.")
ov("cc-canvas,4CCFader,changes,clockabout", PORT, "midi_fx", P_LUA, "MIDI CC performance controller — additive (control covers banks, not slew/recall).")
ov("4CCFader,cc-canvas", NICHE, "tool", P_LUA, "CC fader/slew controller — partial overlap with control module.")
ov("clockabout", PORT, "midi_fx", P_LUA, "Non-linear MIDI clock — additive groove engine.")
ov("metronome,midiplayer", DUP, "midi_fx", P_LUA, "Metronome / MIDI-file player — covered by host clock + midi-player.")
ov("midi-monitor,midi-review", NICHE, "tool", P_LUA, "MIDI monitor/visualiser — handy dev tool, low end-user value.")
ov("serialism,nørgård", PORT, "midi_fx", P_LUA, "12-tone / infinity-series generator — additive compositional engine.")

# --- Synths worth adapting (SC engine port, additive flavour) ---
ov("benjolis,blippoo", PORT, "sound_generator", P_SCS, "Benjolin/Blippoo chaotic synth — additive chaos voice; tiny engine.")
ov("overtones,sines", PORT, "sound_generator", P_SCS, "Additive synthesis — additive (no additive synth in Schwung).")
ov("molly_the_poly,grendy,triangles,stjörnuíþrótt,grendy", ADAPT, "sound_generator", P_SCS, "Classic poly / drone synth — borderline-additive (workflow or texture).")
ov("dronecaster,sonde,asterion,rachim", ADAPT, "sound_generator", P_SCS, "Drone designer + capture / orchestral drone — additive workflow.")
ov("pitfalls", ADAPT, "tool", P_LUA, "Microtonal scale explorer / arp — additive xenharmonic workflow.")
ov("overtones", PORT, "sound_generator", P_SCS, "Additive synth.")
ov("benjolis", PORT, "sound_generator", P_SCS, "Benjolin chaos voice.")
ov("hs010,krautdrums,time-rhythm,euclidigons", NICHE, "sound_generator", P_SCS, "Bass/lead/duophonic voice — overlaps existing synths.")
ov("schicksalslied,shnthsalslied,tetrabobo,tetrabotis,bitebeet,benjolis,blippoo,demoncore", NICHE, "sound_generator", P_SCS, "Experimental/shbobo voice — niche but characterful.")
ov("molly_the_poly", ADAPT, "sound_generator", P_SCS, "Classic poly + generative patch creator.")
ov("synthy,raindrops,lissadron,høst,støv,støy,benjolis", NICHE, "sound_generator", P_SCS, "Small character synth — niche.")

# --- Audio FX worth adapting ---
ov("leslie", PORT, "audio_fx", P_SCS, "Rotary-speaker / doppler — additive (no Leslie in Schwung). (Clone not in local mirror.)")
ov("phyllis", PORT, "audio_fx", P_SCS, "Modeled analog filter FX — additive standalone filter.")
ov("massif,stack,fixed-filter-banking-c", NICHE, "audio_fx", P_SCS, "Resonator / filter bank — partial overlap with spectra.")
ov("glaciers", ADAPT, "audio_fx", P_SCS, "Spectral freeze + harmonise — additive texture.")
ov("sway", ADAPT, "audio_fx", P_SCS, "Analysis-driven adaptive FX — additive self-modulating processor.")
ov("raft,tunnels", ADAPT, "audio_fx", P_SOFT, "Modulated/uncertain delay + texture — additive ambience.")
ov("dimension-expander,phyllis", NICHE, "audio_fx", P_SCS, "Stereo widener / utility — partial overlap with usefulity.")
ov("b-b-b-b-beat,glitchlets,clcks,d", NICHE, "audio_fx", P_SOFT, "Beat-repeat / glitch — overlaps punchfx / performance-fx.")
ov("leslie,phyllis", PORT, "audio_fx", P_SCS, "Additive FX.")
ov("repeater,blndr", NICHE, "audio_fx", P_SOFT, "Textural / quantized delay — partial overlap.")

# --- Pure granular / drone duplicates handled above; ensure a few extras ---
ov("oooooo,mlr", PORT, "overtake", P_SOFT, "Iconic looper / live-cutter — fills the missing looper category.")

# ---------------------------------------------------------------------------
# Heuristic fallback for anything not overridden
# ---------------------------------------------------------------------------
def heuristic(s):
    t = set(s["tags"])
    d = s["desc"].lower()
    # not-applicable signals
    if t & {"art"} and not (t & {"sequencer","synth","midi","generative","drum","sampler","looper"}):
        return (SKIP, None, P_NA, "Visual/art piece — minimal portable workflow.")
    if t & {"lib","mod","eduscript"} and not (t & {"sequencer","synth","midi","drum"}):
        return (SKIP, None, P_NA, "Library / mod / teaching script.")
    if t & {"audio_fx"}:
        port = P_SOFT if ("delay" in t or "looper" in t) else P_SCS
        return (NICHE, "audio_fx", port, "Audio FX — check overlap with Schwung's deep FX shelf.")
    if t & {"synth","drone","engine"}:
        return (NICHE, "sound_generator", P_SCS, "Synth voice — likely overlaps existing synths; port if flavour is distinct.")
    if t & {"granulator"}:
        return (DUP, "audio_fx", P_SOFT, "Granular — covered by granny/verglas/Boris.")
    if t & {"looper"}:
        return (ADAPT, "tool", P_SOFT, "Looper — additive; needs softcut-equivalent buffer DSP.")
    if t & {"sampler"}:
        return (NICHE, "tool", P_SOFT, "Sampler — overlaps slicer/mrsample; port distinct workflows only.")
    if t & {"sequencer","generative"}:
        form = "overtake" if (t & {"grid","arc"}) else "midi_fx"
        port = P_CV if (t & {"crow","jf","w"}) else P_LUA
        return (ADAPT, form, port, "Sequencer — adapt grid→pads or run as MIDI-FX; logic ports cleanly.")
    if t & {"midi"}:
        return (ADAPT, "midi_fx", P_LUA, "MIDI utility — small Lua-logic port.")
    if t & {"utility"}:
        return (NICHE, "tool", P_LUA, "Utility — assess case by case.")
    return (NICHE, None, P_LUA, "Uncategorised — needs a closer look.")

# Build final triage records
def featured_lookup(name):
    for f in FEATURED:
        if f["name"].lower() == name.lower():
            return f
    return None

TRIAGE = []
for s in SCRIPTS:
    key = s["name"].lower()
    f = featured_lookup(s["name"])
    if f:
        bucket, form, port = f["bucket"], f["form"], f["port"]
        note = "★ FEATURED — see deep-dive above. " + (f["additive"][:90].replace("<code>","").replace("</code>","").replace("<strong>","").replace("</strong>","").replace("<em>","").replace("</em>","") + "…")
    elif key in OV:
        bucket, form, port, note = OV[key]
    else:
        bucket, form, port, note = heuristic(s)
    TRIAGE.append({**s, "bucket": bucket, "form": form, "port": port, "note": note, "featured": bool(f)})

# ---------------------------------------------------------------------------
# Render HTML
# ---------------------------------------------------------------------------
def esc(x): return html.escape(str(x or ""))

# stats
from collections import Counter
bstat = Counter(r["bucket"] for r in TRIAGE)
fstat = Counter(r["form"] for r in TRIAGE if r["form"])
pstat = Counter(r["port"][0] for r in TRIAGE)

def badge(text, cls): return f'<span class="badge {cls}">{esc(text)}</span>'

# ---------------------------------------------------------------------------
# BUILT — modules already implemented (clean-room ports), with hands-on test
# steps. status: "ready" = aarch64 tarball delivered, sideload + test now;
# "built" = built & unit-tested on another branch, not packaged here yet.
# `id` is the Schwung module id (install dir / chain pick name).
# `testing` = real put-it-through-its-paces steps (not "how to load it").
# ---------------------------------------------------------------------------
BUILT = {
 "overtones": {"status":"ready","id":"overtones","slot":"sound generator", "testing":[
   "Load into a Signal Chain <b>sound-generator</b> slot and hold a note. The default patch is a <b>pure sine</b> (snapshot 0 = fundamental only): confirm one clean pitch, no buzz, no overtones.",
   "<b>Waveforms → Snapshot 1</b>: knobs 1–8 are the 8 partial amplitudes. Hold a note and raise knob 1 (fundamental), then 2, 3… one at a time — each adds the next-higher harmonic; the tone should brighten predictably (knob 8 = a faint, high partial).",
   "<b>Morph:</b> make Snapshot 1 bright (several partials) and Snapshot 2 dark (knob 1 only). Morph page → Start 0, End 1, Rate ~4 s, Src 0 (LFO). Hold a note: timbre should slowly sweep bright↔dark every few seconds. Switch Src to 2 (Env): each new note starts at Start and settles to End across attack→sustain (onset timbre ≠ held timbre).",
   "<b>Motion (headphones):</b> raise Pan Width + Pan Rate ~8 → partials wander across the stereo field. Raise Pitch Drift to ~10 Hz → a gentle detuned/beating chorus. Both should be clearly audible.",
   "<b>Polyphony + envelope:</b> Envelope page → Attack ~3 s, Release ~4 s. Swell in a chord; then play <b>more than 8 notes</b> — the 9th should steal the oldest voice (no stuck notes/overflow). Lift hands → tails ring out over the Release time.",
   "<b>Abuse / safety:</b> all partials up + 8-note chord + Level high → must stay clean (no digital clipping, crackle, or dropout). Confirms the output limiter holds.",
 ]},
 "massif": {"status":"ready","id":"massif","slot":"audio FX", "testing":[
   "Place <b>after a sound source</b> in a chain (drum slot, sampler, synth) or as a <b>Master FX</b>. It resonates the <i>input</i> — with no input it is silent by design.",
   "Feed percussive/broadband audio (drum loop or noise). Default = 8 short-ring resonators: the dry sound should gain a metallic, pitched ring. Confirm the input becomes 'tuned.'",
   "<b>Tuning (Frequencies page, 8 knobs):</b> set the 8 peaks to a chord (root/3rd/5th across two octaves). Run noise or a drum loop through → the output should <i>sing that chord</i>; turning a freq knob audibly re-pitches that peak.",
   "<b>Ring time (Ring Times page):</b> raise all rings toward 2–5 s, feed a single hit → it should ring/sustain like a struck string or bell long after the transient (longer ring = longer T60 tail).",
   "<b>Isolate a peak (Amplitudes page):</b> set 7 amps to 0 and one to max → you should hear exactly one resonant frequency. Confirms each peak is independent and on-pitch.",
   "<b>Don't-blow-up test:</b> long rings + all amps up + loud input + Master high. A raw resonator bank would scream/clip — confirm massif stays clean and bounded (the normalized-gain + limiter design).",
 ]},
 "benjolis": {"status":"ready","id":"benjolis","slot":"sound generator", "testing":[
   "Load into a sound-gen slot and <b>hold a note</b> to open the gate — you should hear an evolving <b>chaotic Benjolin texture</b>, not a clean note. Release: it goes silent but keeps churning internally (the next note picks up a new state).",
   "<b>Cross-mod:</b> knobs 1 & 2 = osc1/osc2 freq. Detune them → you should get sidebands/grit from the cross-modulation, not two independent pitches.",
   "<b>Rungler:</b> raise Rungler 1/2 depth + Rungler Filt → a stepped, looping pseudo-random pattern should drive pitch/filter (the signature 'rungler'): a repeating-but-evolving stepped motion.",
   "<b>Loop/chaos:</b> Loop near 1 = self-feeding chaos; lower it → more periodic/tame. Confirm the chaos amount changes.",
   "<b>Filter:</b> sweep Filter Freq, raise Resonance, switch Type LP/HP/BP → the filter should clearly reshape the tone and the three types should sound distinct.",
   "<b>Output select + safety:</b> step the Output param (Tri1 … Rungler … Filter) — each should sound different (Rungler = raw stepped, Filter = smoothed). Then max Rungler depth + Resonance + Gain → must stay bounded (no NaN dropout, no runaway).",
 ]},
 "phyllis": {"status":"ready","id":"phyllis","slot":"audio FX", "testing":[
   "Place after a <b>harmonically rich</b> source (saw synth, drums, noise) in a chain, or as a Master FX. Needs input audio.",
   "<b>Cutoff sweep:</b> sweep Cutoff high→low on a bright input → classic low-pass darkening; highs should roll off as cutoff drops.",
   "<b>Resonance:</b> raise toward 1 → a pronounced peak at the cutoff; near the top it should whistle/ring (approaching self-oscillation).",
   "<b>Type:</b> switch LP↔HP → HP should thin out the lows (the inverse of LP).",
   "<b>Drive:</b> push Drive/gain → the signal should overdrive into the filter (added harmonics/dirt), an analog-style saturation rather than just louder.",
   "<b>Noise + character:</b> small Noise values add a subtle analog hiss. Compared to a clean DSP filter, phyllis should feel slightly nonlinear/'analog,' especially with drive + resonance up.",
 ]},
 "changes": {"status":"built","id":"changes","slot":"MIDI FX","branch":"feat/changes-cc-modulation-bank", "testing":[
   "Place in a chain <b>MIDI-FX</b> slot before a synth, or aim it at Move's own track macros (Pre mode → inject to Move MIDI_IN). It outputs <b>CC, not notes</b>.",
   "Map one of its 8 LFO outputs to a synth param (e.g. a filter-cutoff CC). With the synth holding a note, that param should move on its own in a smooth evolving pattern — confirm hands-free motion.",
   "Change an LFO's rate, depth, and phase offset → confirm the speed/amount/relationship of the motion changes; run several LFOs at different rates → confirm independent, phase-linked movement.",
   "<b>Key unknown to confirm:</b> aim a CC at a Move <i>track macro</i> and verify Move actually responds. The CC-to-Move-track path is unverified — this is the test that settles <code>docs/MOVE_CC_MAP.md</code>.",
 ]},
 "clockabout": {"status":"built","id":"clockabout","slot":"MIDI FX","branch":"feat/norns-adaptations-batch", "testing":[
   "Put before a sequenced synth in a chain MIDI-FX slot. Feed a steady stream of notes and raise swing → off-beats should shift late/early audibly vs the straight input.",
   "Try the different curve shapes (linear / exp / groove) → confirm the timing <i>feel</i> changes between them.",
   "Confirm scope: it can only <b>delay</b> notes passing through it, not advance Move's own master clock (documented limitation) — so it grooves the note stream, not the transport.",
 ]},
 "nørgård": {"status":"built","id":"norgaard","slot":"MIDI FX","branch":"feat/norns-adaptations-batch", "testing":[
   "MIDI-FX slot before a synth. Trigger it (note/clock) → it should emit a self-similar, non-repeating melodic line (the infinity series).",
   "Set Scale + Root → confirm output is quantized to that scale. Change sequence length / step rate → confirm the read-head speed and pattern window change.",
   "Confirm it's deterministic/self-similar: the same start reproduces the same line — a recognizable fractal melody, not randomness.",
 ]},
 "serialism": {"status":"built","id":"serialism","slot":"MIDI FX","branch":"feat/norns-adaptations-batch", "testing":[
   "MIDI-FX slot before a synth. Define a 12-tone row and run it → it should play the row.",
   "Switch transform Prime / Retrograde / Inversion / Retrograde-Inversion → verify by ear or log: Retrograde = row reversed, Inversion = intervals flipped, RI = both.",
 ]},
 "magpie": {"status":"built","id":"magpie","slot":"MIDI FX","branch":"feat/norns-adaptations-batch", "testing":[
   "MIDI-FX slot before a synth. Play notes → magpie should re-emit them with evolving delay/transposition/probability (a melodic echo, not an audio delay).",
   "Raise delay + transpose → echoed notes should be pitch-shifted and time-spread; raise the probability/variance → some echoes should drop or vary. Distinct evolving repeats, not a fixed tap.",
 ]},
 "justharmonicon": {"status":"built","id":"justharmonicon","slot":"MIDI FX","branch":"feat/norns-adaptations-batch", "testing":[
   "MIDI-FX slot before a synth. Trigger it → confirm rich undertone-series (subharmonic) voicings <i>below</i> the played root, not standard overtone chords.",
   "Change the subharmonic divisors → confirm the chord voicing changes accordingly.",
 ]},
}

def testing_block(name):
    b = BUILT.get(name)
    if not b: return ""
    steps = "".join(f"<li>{s}</li>" for s in b["testing"])
    sid = b.get("id","")
    slot = b.get("slot","")
    branch = b.get("branch")
    if b["status"] == "ready":
        head = (f'<b>Sideload:</b> Move Manager (<code>:7700</code>) → Modules → Install Custom → '
                f'Upload <code>{esc(sid)}-module.tar.gz</code>. It installs to the {esc(slot)} category, '
                f'then pick <b>{esc(sid)}</b> in a chain {esc(slot)} slot. <b>Start with a preset</b> '
                f'(scroll the selector at the top of the module menu), then tweak. '
                f'<i>On Move these use the standard knob-editor menu — not the full-screen layout mocked above.</i>')
    else:
        head = (f'On branch <code>{esc(branch)}</code> (not in the delivered tarballs). Build + deploy that branch, '
                f'then pick <b>{esc(sid)}</b> in a chain {esc(slot)} slot.')
    return (f'<div class="testing"><div class="testhead">▶ How to put it through its paces</div>'
            f'<div class="testlead">{head}</div><ol class="teststeps">{steps}</ol></div>')

def status_badge(name):
    b = BUILT.get(name)
    if not b: return ""
    label, cls = STATUS_META[b["status"]]
    return badge(label, cls)

# Build-progress banner: what's been ported so far, linking to each card.
READY_ORDER = ["overtones","massif","benjolis","phyllis"]
BUILT_ORDER = ["changes","nørgård","serialism","justharmonicon","magpie","clockabout"]
def _pchip(name, extra=""):
    b = BUILT[name]
    return (f'<a class="pchip {extra}" href="#feat-{esc(name)}">'
            f'<span class="pn">{esc(name)}</span><span class="ps">{esc(b["slot"])}</span></a>')
def progress_section():
    ready = "".join(_pchip(n) for n in READY_ORDER)
    built = "".join(_pchip(n,"b") for n in BUILT_ORDER)
    n_total = len(BUILT)
    return f"""
<section id="progress">
  <h2>Build progress <small>{n_total} clean-room ports done so far — every one has a numerical test suite + ARM build</small></h2>
  <div class="progress">
    <div class="pgroup">
      <div class="glabel"><b style="color:var(--green)">✅ Ready to test now</b> — aarch64 tarballs delivered. Move Manager (<code>:7700</code>) → Modules → Install Custom → Upload, then pick in a chain slot. Each card below has step-by-step "put it through its paces" tests.</div>
      <div class="pchips">{ready}</div>
    </div>
    <div class="pgroup">
      <div class="glabel"><b style="color:var(--blue)">🔨 Built &amp; unit-tested</b> — on dev branches, not in the delivered tarballs yet (MIDI-FX + the CC modulation bank). Build/deploy the noted branch to try them.</div>
      <div class="pchips">{built}</div>
    </div>
    <div class="glabel" style="margin-top:12px">All four "ready" modules are <b>numerically verified offline</b> (DSP shape, stability, spectral correctness) but <b>not yet heard on hardware</b> — these tests are the listening pass that confirms timbre.</div>
    <div class="glabel" style="margin-top:10px;border-top:1px solid var(--line);padding-top:10px">
      <b style="color:var(--yellow)">⚠ UI reality vs. the mockups below:</b> these ported modules ship as <b>chainable DSP only</b> — they have no custom display. On Move they surface through Schwung's standard <b>menu-driven knob editor</b>: a <b>preset selector</b> at the top, then 8 mappable knobs + sub-menus (Waveforms / Morph / Envelope, or Peak Tuning / Ring Times, etc.). The ASCII screens on the cards below are the <b>original norns full-screen concept</b> — a design reference, <i>not</i> what you see on the Move.
    </div>
  </div>
</section>"""

# Featured cards grouped by form
FORM_ORDER = ["midi_fx","sound_generator","audio_fx","tool","overtake"]
def featured_section():
    out = []
    for form in FORM_ORDER:
        items = [f for f in FEATURED if f["form"] == form]
        if not items: continue
        flabel, fdesc = FORM_META[form]
        out.append(f'<h3 class="formhead" id="form-{form}">{esc(flabel)} <small>{esc(fdesc)}</small></h3>')
        for f in items:
            s = sheet(f["name"])
            bname, bemoji, _ = BUCKET_META[f["bucket"]]
            pcls, plabel, pdesc = f["port"]
            url = s["url"]
            link = f'<a href="{esc(url)}" target="_blank" rel="noopener">source ↗</a>' if url else ""
            crow = f'<div class="row"><span class="k">crow/JF</span><span class="v">{f["crow"]}</span></div>' if f.get("crow") and f["crow"]!="—" else ""
            tags = " ".join(badge(t,"tag") for t in s["tags"][:6])
            out.append(f"""
            <div class="card" id="feat-{esc(f['name'])}">
              <div class="cardhead">
                <div class="title">{esc(f['name'])} <span class="auth">by {esc(s['author']) or '—'}</span></div>
                <div class="badges">{status_badge(f['name'])} {badge(bemoji+' '+bname,'b-'+f['bucket'])} {badge(plabel,'p-'+pcls)} {badge('effort: '+f['effort'],'eff')} {('<span class=chain>⛓ chainable</span>' if f['chainable'] else '')}</div>
              </div>
              <div class="tags">{tags} {link}</div>
              <div class="cardbody">
                <div class="design">
                  <div class="row"><span class="k">Additive</span><span class="v">{f['additive']}</span></div>
                  <div class="row"><span class="k">vs Schwung</span><span class="v">{f['why']}</span></div>
                  <div class="row"><span class="k">Mapping</span><span class="v">{f['mapping']}</span></div>
                  <div class="row"><span class="k">Portability</span><span class="v">{esc(plabel)} — {esc(pdesc)}</span></div>
                  {crow}
                </div>
                <div class="mockwrap"><div class="mockcap">◇ norns-original concept — Move uses the knob-editor menu</div><pre class="mock">{esc(f['mockup'])}</pre></div>
              </div>
              {testing_block(f['name'])}
            </div>""")
    return "\n".join(out)

def triage_rows():
    out = []
    for r in sorted(TRIAGE, key=lambda x: ({PORT:0,ADAPT:1,NICHE:2,DUP:3,SKIP:4}[x["bucket"]], x["name"].lower())):
        bname, bemoji, _ = BUCKET_META[r["bucket"]]
        pcls, plabel, _ = r["port"]
        form = FORM_META[r["form"]][0] if r["form"] else "—"
        star = "★" if r["featured"] else ""
        url = r["url"]
        nm = f'<a href="{esc(url)}" target="_blank" rel="noopener">{esc(r["name"])}</a>' if url else esc(r["name"])
        out.append(
          f'<tr data-bucket="{r["bucket"]}" data-form="{r["form"] or ""}" data-port="{pcls}" data-feat="{1 if r["featured"] else 0}">'
          f'<td class="c-name">{star} {nm}</td>'
          f'<td class="c-tags">{esc(", ".join(r["tags"]))}</td>'
          f'<td class="c-bucket">{badge(bemoji+" "+bname,"b-"+r["bucket"])}</td>'
          f'<td class="c-form">{esc(form)}</td>'
          f'<td class="c-port">{badge(plabel,"p-"+pcls)}</td>'
          f'<td class="c-note">{esc(r["note"])}</td>'
          f'</tr>')
    return "\n".join(out)

legend_buckets = "".join(
    f'<div class="leg"><span class="badge b-{k}">{v[1]} {v[0]}</span> {esc(v[2])}</div>'
    for k,v in BUCKET_META.items())
legend_ports = "".join(
    f'<div class="leg"><span class="badge p-{p[0]}">{esc(p[1])}</span> {esc(p[2])}</div>'
    for p in [P_LUA,P_SCS,P_SOFT,P_CV,P_NA])

stat_cards = "".join(
    f'<div class="stat"><div class="num">{bstat.get(k,0)}</div><div class="lbl">{BUCKET_META[k][1]} {BUCKET_META[k][0]}</div></div>'
    for k in [PORT,ADAPT,NICHE,DUP,SKIP])

HTML = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>norns → Schwung · adaptation targets</title>
<style>
:root{{--bg:#0e1116;--panel:#161b22;--panel2:#1c232c;--ink:#e6edf3;--mut:#9aa7b4;--line:#2b333d;
--green:#3fb950;--yellow:#d2a106;--blue:#58a6ff;--gray:#6e7681;--black:#484f58;
--lua:#3fb950;--scs:#d29922;--soft:#a371f7;--cv:#58a6ff;--na:#6e7681;}}
*{{box-sizing:border-box}}
body{{margin:0;background:var(--bg);color:var(--ink);font:15px/1.55 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif}}
a{{color:var(--blue);text-decoration:none}} a:hover{{text-decoration:underline}}
header{{padding:34px 28px 22px;border-bottom:1px solid var(--line);background:linear-gradient(180deg,#11161d,#0e1116)}}
h1{{margin:0 0 6px;font-size:26px;letter-spacing:.2px}}
.sub{{color:var(--mut);max-width:62ch}}
.wrap{{max-width:1180px;margin:0 auto;padding:0 20px}}
section{{padding:26px 0;border-bottom:1px solid var(--line)}}
h2{{font-size:20px;margin:0 0 4px}}
h2 small,.sub small{{color:var(--mut);font-weight:400}}
.method{{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:16px 18px;margin-top:14px;color:var(--mut)}}
.method b{{color:var(--ink)}}
.legendgrid{{display:grid;grid-template-columns:1fr 1fr;gap:8px 28px;margin-top:12px}}
.leg{{font-size:13px;color:var(--mut)}}
.stats{{display:flex;gap:12px;flex-wrap:wrap;margin-top:16px}}
.stat{{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:12px 18px;min-width:120px;text-align:center}}
.stat .num{{font-size:28px;font-weight:700}} .stat .lbl{{color:var(--mut);font-size:13px}}
.badge{{display:inline-block;font-size:11px;font-weight:600;padding:2px 7px;border-radius:20px;border:1px solid var(--line);white-space:nowrap}}
.tag{{color:var(--mut);background:#0d1117}}
.b-port{{color:#0e1116;background:var(--green)}} .b-adapt{{color:#0e1116;background:var(--yellow)}}
.b-niche{{color:#0e1116;background:var(--blue)}} .b-dup{{color:var(--ink);background:#30363d}}
.b-skip{{color:var(--mut);background:#21262d}}
.p-lua{{color:var(--lua);border-color:var(--lua)}} .p-sc-stock{{color:var(--scs);border-color:var(--scs)}}
.p-softcut{{color:var(--soft);border-color:var(--soft)}} .p-cv{{color:var(--cv);border-color:var(--cv)}}
.p-na{{color:var(--na);border-color:var(--na)}} .eff{{color:var(--mut)}}
.chain{{font-size:11px;color:var(--soft)}}
.st-ready{{color:#0e1116;background:var(--green)}} .st-built{{color:#0e1116;background:var(--blue)}}
.testing{{border-top:1px solid var(--line);background:#0c1a13;padding:14px 16px}}
.testhead{{font-weight:700;color:var(--green);font-size:13px;letter-spacing:.3px;margin-bottom:6px}}
.testlead{{color:var(--mut);font-size:13px;margin-bottom:8px}}
.testlead code,.teststeps li code{{background:#0d1117;padding:1px 5px;border-radius:4px;font-size:12px}}
.teststeps{{margin:0;padding-left:20px}}
.teststeps li{{font-size:13.5px;margin:7px 0;color:var(--ink)}}
.teststeps li b{{color:#7ee08a}}
.progress{{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:16px 18px;margin-top:14px}}
.pgroup{{margin:10px 0}}
.pgroup .glabel{{font-size:13px;color:var(--mut);margin-bottom:8px}}
.pchips{{display:flex;gap:8px;flex-wrap:wrap}}
.pchip{{display:inline-flex;flex-direction:column;gap:1px;background:var(--panel2);border:1px solid var(--line);
border-radius:9px;padding:7px 12px;min-width:96px}}
.pchip:hover{{border-color:var(--green);text-decoration:none}}
.pchip .pn{{font-weight:700;font-size:14px;color:var(--ink)}} .pchip .ps{{font-size:11px;color:var(--mut)}}
.pchip.b .pn{{}} .pchip.b:hover{{border-color:var(--blue)}}
.formhead{{margin:26px 0 10px;font-size:17px;border-left:3px solid var(--blue);padding-left:10px}}
.formhead small{{color:var(--mut);font-weight:400;font-size:13px}}
.card{{background:var(--panel);border:1px solid var(--line);border-radius:12px;margin:14px 0;overflow:hidden}}
.cardhead{{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;padding:14px 16px 8px;flex-wrap:wrap}}
.title{{font-size:18px;font-weight:700}} .auth{{color:var(--mut);font-size:13px;font-weight:400}}
.badges{{display:flex;gap:6px;flex-wrap:wrap;align-items:center}}
.tags{{padding:0 16px 10px;display:flex;gap:6px;flex-wrap:wrap;align-items:center}}
.cardbody{{display:grid;grid-template-columns:1fr 420px;gap:0;border-top:1px solid var(--line)}}
.design{{padding:14px 16px}}
.row{{display:grid;grid-template-columns:92px 1fr;gap:12px;padding:5px 0;border-bottom:1px dashed #232a33}}
.row:last-child{{border-bottom:0}}
.row .k{{color:var(--mut);font-size:12px;text-transform:uppercase;letter-spacing:.4px;padding-top:2px}}
.row .v{{font-size:14px}} .row .v code{{background:#0d1117;padding:1px 5px;border-radius:4px;font-size:12px}}
.mockwrap{{border-left:1px solid var(--line);display:flex;flex-direction:column}}
.mockcap{{background:#0a0d12;color:var(--mut);font-size:10px;letter-spacing:.3px;padding:6px 16px 0;
font-style:italic}}
.mock{{margin:0;background:#0a0d12;padding:8px 16px 14px;color:#9fe0a8;
font:12px/1.35 ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;white-space:pre;overflow-x:auto}}
@media(max-width:840px){{.cardbody{{grid-template-columns:1fr}}.mockwrap{{border-left:0;border-top:1px solid var(--line)}}.legendgrid{{grid-template-columns:1fr}}}}
.controls{{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0;align-items:center}}
.controls input,.controls select{{background:var(--panel);border:1px solid var(--line);color:var(--ink);
padding:8px 10px;border-radius:8px;font-size:13px}}
.controls input{{min-width:220px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th,td{{text-align:left;padding:7px 9px;border-bottom:1px solid #1f262e;vertical-align:top}}
th{{position:sticky;top:0;background:var(--panel2);cursor:pointer;user-select:none;font-size:12px;color:var(--mut);
text-transform:uppercase;letter-spacing:.3px}}
th:hover{{color:var(--ink)}}
tr:hover td{{background:#12171e}}
.c-name{{white-space:nowrap;font-weight:600}} .c-tags{{color:var(--mut);max-width:220px}}
.c-note{{color:var(--mut);max-width:430px}}
.count{{color:var(--mut);font-size:13px;margin-left:auto}}
footer{{padding:30px 20px;color:var(--mut);font-size:13px;text-align:center}}
.toc{{display:flex;gap:14px;flex-wrap:wrap;margin-top:10px;font-size:13px}}
</style></head>
<body>
<header><div class="wrap">
  <h1>norns → Schwung · adaptation targets</h1>
  <div class="sub">Triage of <b>{len(SCRIPTS)} norns scripts</b> against Schwung's <b>80-module</b> catalog, scored for <b>additivity</b>, <b>portability</b>, and <b>Move form-fit</b> — with {len(FEATURED)} source-verified deep-dives and ASCII mockups for the Move's 128×64 display, 4×8 pads, and 8 encoders.</div>
  <div class="toc">
    <a href="#method">Methodology</a><a href="#progress">Build progress ({len(BUILT)})</a><a href="#featured">Featured ({len(FEATURED)})</a>
    <a href="#form-midi_fx">MIDI FX</a><a href="#form-sound_generator">Synths</a><a href="#form-audio_fx">Audio FX</a>
    <a href="#form-tool">Tools</a><a href="#form-overtake">Overtake</a><a href="#triage">Full triage ({len(SCRIPTS)})</a>
  </div>
</div></header>

<div class="wrap">

<section id="method">
  <h2>Methodology <small>how each script was judged</small></h2>
  <div class="method">
    A norns script splits into a <b>Lua layer</b> (sequencing, MIDI, UI, params — the workflow) and an optional
    <b>SuperCollider engine</b> (the DSP). That split <i>is</i> the portability story:
    pure-Lua scripts are logic reimplements in QuickJS/C; SC engines need a native DSP <code>.so</code> (or scsynth via your
    <code>sc-plugins-arm64</code> path); softcut scripts need a varispeed buffer DSP; crow/JF logic is kept as MIDI and the
    physical CV jacks are dropped. The <b>{len(FEATURED)} featured</b> targets were source-inspected (engine/softcut/crow/grid
    confirmed). The full <b>{len(SCRIPTS)}-row table is curated + heuristic</b> — deduped against Schwung's catalog and bucketed
    by tags/description, <i>not</i> source-verified per row. Norns gives 3 keys + 3 encoders + 128×64; the Move gives 8 encoders
    + a 4×8 pad grid + transport keys — so most adaptations are an <b>expansion</b>, not a squeeze.
    <div class="legendgrid" style="margin-top:14px"><div><b>Verdict buckets</b>{legend_buckets}</div><div><b>Portability</b>{legend_ports}</div></div>
  </div>
  <div class="stats">{stat_cards}</div>
</section>

{progress_section()}

<section id="featured">
  <h2>Featured deep-dives <small>{len(FEATURED)} high-value targets, balanced across all five form factors</small></h2>
  {featured_section()}
</section>

<section id="triage">
  <h2>Full triage <small>all {len(SCRIPTS)} scripts — sortable & filterable</small></h2>
  <div class="controls">
    <input id="q" type="search" placeholder="search name / tags / note…">
    <select id="fb"><option value="">all verdicts</option>
      <option value="port">🟢 Port</option><option value="adapt">🟡 Adapt</option>
      <option value="niche">🔵 Niche</option><option value="dup">⚪ Duplicate</option><option value="skip">⚫ Skip</option></select>
    <select id="ff"><option value="">all forms</option>
      <option value="midi_fx">MIDI FX</option><option value="sound_generator">Sound Generator</option>
      <option value="audio_fx">Audio FX</option><option value="tool">Tool</option><option value="overtake">Overtake</option></select>
    <select id="fp"><option value="">all portability</option>
      <option value="lua">Lua-only</option><option value="sc-stock">SC engine</option>
      <option value="softcut">Softcut</option><option value="cv">CV-reinterp</option><option value="na">N/A</option></select>
    <label style="color:var(--mut);font-size:13px"><input id="fonly" type="checkbox"> featured only</label>
    <span class="count" id="count"></span>
  </div>
  <table id="tbl"><thead><tr>
    <th data-c="0">Script</th><th data-c="1">norns tags</th><th data-c="2">Verdict</th>
    <th data-c="3">Move form</th><th data-c="4">Portability</th><th data-c="5">Adaptation note</th>
  </tr></thead><tbody>
  {triage_rows()}
  </tbody></table>
</section>

<section>
  <h2>Scoring rubric</h2>
  <div class="method">
    <b>Additivity</b> — does Schwung (or a community module) already do this? Duplicates are demoted regardless of quality.<br>
    <b>Portability</b> — Lua-only ➜ SC engine ➜ Softcut ➜ CV-reinterpret ➜ N/A (increasing effort/risk).<br>
    <b>Form-fit</b> — which Move idiom it becomes: MIDI FX / Sound Generator / Audio FX / Tool / Overtake.<br>
    <b>Workflow density</b> — does it bring a whole <i>way of working</i> (a sequencer paradigm, a tuning system, a looper) rather than one knob?<br>
    <b>Grid/crow dependence</b> — none / adaptable-to-encoders / needs-overtake-grid / dead-without-hardware.
  </div>
</section>

</div>
<footer>Generated from norns_scripts_discourse.xlsx · {len(SCRIPTS)} scripts · {len(FEATURED)} deep-dives · build_schwung_targets.py</footer>

<script>
const tbl=document.getElementById('tbl'),tb=tbl.tBodies[0],rows=[...tb.rows];
const q=document.getElementById('q'),fb=document.getElementById('fb'),ff=document.getElementById('ff'),
fp=document.getElementById('fp'),fonly=document.getElementById('fonly'),count=document.getElementById('count');
function apply(){{
  const s=q.value.toLowerCase(),b=fb.value,f=ff.value,p=fp.value,o=fonly.checked;let n=0;
  rows.forEach(r=>{{
    const txt=r.textContent.toLowerCase();
    const ok=(!s||txt.includes(s))&&(!b||r.dataset.bucket===b)&&(!f||r.dataset.form===f)&&
             (!p||r.dataset.port===p)&&(!o||r.dataset.feat==='1');
    r.style.display=ok?'':'none'; if(ok)n++;
  }});
  count.textContent=n+' / '+rows.length+' shown';
}}
[q,fb,ff,fp].forEach(e=>e.addEventListener('input',apply));fonly.addEventListener('change',apply);
const order={{port:0,adapt:1,niche:2,dup:3,skip:4}};
let sortc=-1,asc=true;
tbl.tHead.querySelectorAll('th').forEach(th=>th.addEventListener('click',()=>{{
  const c=+th.dataset.c; asc=(sortc===c)?!asc:true; sortc=c;
  rows.sort((a,b)=>{{
    let x=a.cells[c].textContent.trim().toLowerCase(),y=b.cells[c].textContent.trim().toLowerCase();
    if(c===2){{x=order[a.dataset.bucket];y=order[b.dataset.bucket];}}
    return (x>y?1:x<y?-1:0)*(asc?1:-1);
  }});
  rows.forEach(r=>tb.appendChild(r));
}}));
apply();
</script>
</body></html>"""

with open(OUT,"w") as fh:
    fh.write(HTML)
print("wrote", OUT, f"({len(HTML)//1024} KB)  scripts={len(SCRIPTS)} featured={len(FEATURED)}")
print("buckets:", dict(bstat))
